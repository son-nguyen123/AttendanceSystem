"""Convert the older Factory 1 attendance workbook into the current frame.

The legacy Factory 1 files are usually old ``.xls`` Output 1 workbooks.  We
read their semantic payroll summary first, then let the normal payroll writer
create the current formula-driven summary.  This deliberately does not copy
bank-account cells: the converted file is a payroll/history source, not a bank
registry import.
"""

from __future__ import annotations

import math
import shutil
from pathlib import Path
from tempfile import TemporaryDirectory

import xlrd
from openpyxl import Workbook, load_workbook

from app.services.data_mapper import (
    _normalize_employee_code,
    _normalize_label,
    _owner_records_by_code,
    _prepare_mapping_source_workbook,
)
from app.services.payroll_store import normalize_employee_name
from app.services.payroll_workbook import export_payroll_workbook
from app.services.workbook_processor import export_processed_workbook


_OVERTIME_LABELS = {
    "so h tang ca",
    "so gio tang ca",
    "gio lam them",
    "h tang",
}


def export_factory1_legacy_output2(source_path: Path, output_path: Path) -> Path:
    """Convert one old Factory 1 workbook to the current formula frame."""

    with TemporaryDirectory(prefix="factory1-legacy-") as temp_dir:
        prepared_path = Path(temp_dir) / "legacy.xlsx"
        if source_path.suffix.lower() == ".xls":
            _convert_xls_to_xlsx(source_path, prepared_path)
        else:
            shutil.copy2(source_path, prepared_path)

        # Some old .xlsx files contain tens of thousands of style-only rows
        # below the actual payroll table. Compact the XML before openpyxl sees
        # it; otherwise it allocates a cell object for every blank row twice.
        if source_path.suffix.lower() == ".xls":
            readable_path = prepared_path
        else:
            readable_path, _ = _prepare_mapping_source_workbook(source_path, Path(temp_dir))

        workbook = load_workbook(readable_path, data_only=False)
        values_workbook = load_workbook(readable_path, data_only=True)
        source_sheet = _select_sheet(workbook)
        values_sheet = values_workbook[source_sheet.title]
        records = _owner_records_by_code(source_sheet, values_sheet)
        if not records:
            raise ValueError("KhÃ´ng tÃ¬m tháº¥y dá»¯ liá»‡u nhÃ¢n viÃªn trong báº£ng cÅ© XÆ°á»Ÿng 1")

        # Reuse the semantic summary reader for period-specific inputs such as
        # bonus, penalty and advance.  Calculated daily/hourly/final values are
        # intentionally not copied below because the new frame writes formulas.
        # In particular, the old combined "Ứng Lương + Phạt" field is read by
        # that helper as ``advance_or_penalty`` so its full amount is entered in
        # the new "Ứng Lương" cell.  The formula itself is unchanged: the new
        # frame simply receives the same monthly input value.
        from app.services.factory2_workbook import _read_legacy_payroll_values

        summary_values = _read_legacy_payroll_values(
            readable_path,
            formula_workbook=workbook,
            values_workbook=values_workbook,
        )
        overtime = _legacy_overtime_by_code(source_sheet, values_sheet, records)
        workbook.close()
        values_workbook.close()
        source_values: dict[str, dict] = {}
        for code, record in records.items():
            old_values = summary_values.get(code, {})
            # Keep identity/profile values from the old month.  Do not pass
            # old daily/hourly/final totals: the new frame must write formulas
            # for the new month's calculations.
            item = {
                "name": normalize_employee_name(record.name or ""),
                "start_work_note": record.start_work_note or "",
                "note": record.note or "",
                "monthly_salary": _number(record.salary),
                "bonus": _number(record.bonus) if record.bonus is not None else _number(old_values.get("bonus")),
                "overtime_hours": _number(old_values.get("overtime_hours")) if old_values.get("overtime_hours") is not None else overtime.get(code, 0),
                # An explicit empty value prevents the normal payroll writer
                # from filling a bank number from the local bank registry.
                "bank_account": "",
            }
            for key in ("penalty_rate", "nq_penalty", "advance_or_penalty"):
                value = _number(old_values.get(key))
                if value is not None:
                    item[key] = value
            source_values[code] = item

        # The normal payroll export first creates Output 1 and then reads it
        # again to write Output 2.  Reuse that processed intermediate here so
        # a legacy conversion does not load the same large workbook twice.
        processed_path = Path(temp_dir) / "processed.xlsx"
        export_processed_workbook(readable_path, processed_path, factory="factory1")
        return export_payroll_workbook(
            processed_path,
            output_path,
            include_saved_data=True,
            factory="factory1",
            source_payroll_values=source_values,
            source_is_processed=True,
        )


def _select_sheet(workbook):
    best = None
    best_count = -1
    for worksheet in workbook.worksheets:
        count = sum(1 for row in range(1, worksheet.max_row + 1) if str(worksheet.cell(row, 1).value or "").strip() == "Att. Time")
        if count > best_count:
            best = worksheet
            best_count = count
    if best is None or best_count <= 0:
        raise ValueError("KhÃ´ng tÃ¬m tháº¥y sheet cháº¥m cÃ´ng cÅ© XÆ°á»Ÿng 1")
    return best


def _legacy_overtime_by_code(source_sheet, values_sheet, records: dict[str, object]) -> dict[str, int | float]:
    """Read ``Sá»‘ H táº±ng ca``/``Giá» lÃ m thÃªm`` from old summary rows."""

    result: dict[str, int | float] = {}
    for row in range(1, source_sheet.max_row + 1):
        overtime_columns = [
            col
            for col in range(1, source_sheet.max_column + 1)
            if _normalize_label(source_sheet.cell(row, col).value) in _OVERTIME_LABELS
        ]
        if not overtime_columns:
            continue
        code_columns = [
            col
            for col in range(1, source_sheet.max_column + 1)
            if _normalize_label(source_sheet.cell(row, col).value) in {"ma", "mÃ£"}
        ]
        if not code_columns:
            continue
        code_col = code_columns[0]
        for candidate_row in range(row + 1, min(source_sheet.max_row, row + 24) + 1):
            code = _normalize_employee_code(values_sheet.cell(candidate_row, code_col).value)
            if not code or code not in records:
                continue
            for overtime_col in overtime_columns:
                value = _number(values_sheet.cell(candidate_row, overtime_col).value)
                if value is not None:
                    result.setdefault(code, value)
                    break
    return result


def _number(value: object) -> int | float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            return None
        return int(value) if float(value).is_integer() else float(value)
    text = str(value).strip().replace(" ", "")
    if not text or text in {"-", "?"}:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def _convert_xls_to_xlsx(source_path: Path, output_path: Path) -> None:
    """Convert values from BIFF8 ``.xls`` while dropping style-only tail rows."""

    book = xlrd.open_workbook(str(source_path), formatting_info=False)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_index, sheet_name in enumerate(book.sheet_names()):
        source_sheet = book.sheet_by_index(sheet_index)
        target = workbook.create_sheet(_safe_sheet_title(sheet_name, sheet_index))
        # Old attendance workbooks often carry thousands of formatted but
        # empty rows below the actual table. ``xlrd`` exposes those rows in
        # ``nrows``; trim the meaningful tail before copying values so the
        # conversion does not scan blank cells twice.
        last_row = source_sheet.nrows
        while last_row > 0 and not any(value not in (None, "") for value in source_sheet.row_values(last_row - 1)):
            last_row -= 1
        for row_index in range(last_row):
            row_values = source_sheet.row_values(row_index)
            for col_index, raw_value in enumerate(row_values):
                if raw_value in (None, ""):
                    continue
                cell = source_sheet.cell(row_index, col_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, book.datemode)
                elif cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK, xlrd.XL_CELL_ERROR}:
                    value = None
                if value not in (None, ""):
                    target.cell(row=row_index + 1, column=col_index + 1).value = value
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _safe_sheet_title(value: str, index: int) -> str:
    title = "".join(char for char in str(value or "Sheet") if char not in "[]:*?/\\")[:31] or f"Sheet{index + 1}"
    return title
