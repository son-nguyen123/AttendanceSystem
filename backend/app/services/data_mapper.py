from __future__ import annotations

import copy
from math import ceil
import re
import unicodedata
from contextlib import ExitStack
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any, Literal
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.payroll_store import normalize_employee_name


SUMMARY_TOTAL_LABELS = {"tong gio cong", "tong giờ công"}
SUMMARY_CODE_LABELS = {"ma", "mã"}
SUMMARY_MIN_COL = 32
SUMMARY_ROWS = 7
MONEY_FORMAT = "#,##0"
FLEXIBLE_NUMBER_FORMAT = "General"
WHITE = "FFFFFF"
NOTE_FILL = "FFF4CC"
START_WORK_FILL = "DDEBF7"
YELLOW = "FFFF00"
GREEN = "C4D79B"
RED = "C00000"
BLACK = "000000"
FONT_NAME = "Arial"
MIN_TRAILING_STYLE_ROWS_TO_COMPACT = 512

_WORKSHEET_XML_PREFIX = "xl/worksheets/"
_ROW_ELEMENT_RE = re.compile(br"<row\b[^>]*(?:/>|>.*?</row>)", re.DOTALL)
_ROW_NUMBER_RE = re.compile(br"\br=[\"'](\d+)[\"']")
_SEMANTIC_CELL_CONTENT_RE = re.compile(br"<(?:v|f|is)(?:\s|>)")
_DIMENSION_RE = re.compile(br"(<dimension\b[^>]*\bref=[\"'])([^\"']+)([\"'])")

SALARY_LABELS = {"muc luong"}
BONUS_LABELS = {"thuong"}
NAME_LABELS = {"ten", "ten nhan vien", "ho va ten"}
NOTE_LABELS = {"ghi chu", "ghi chu ho so", "note"}
LEGACY_DEDUCTION_LABELS = {"ung luong + phat"}
ADVANCE_LABELS = {"ung luong"}
NQ_PENALTY_LABELS = {"phat nq"}
NQ_RATE_LABELS = {"muc tien phat nq tren gio cong (d)", "muc tien phat nq tren gio cong"}


@dataclass
class SummaryBlock:
    code: str
    header_row: int
    result_row: int
    total_col: int
    code_col: int
    data_start_col: int
    data_end_col: int


@dataclass(frozen=True)
class OwnerRecord:
    code: str
    name: str | None = None
    bank_account: str | None = None
    start_work_note: str | None = None
    salary: Any = None
    bonus: Any = None
    note: str | None = None
    has_previous_deduction: bool = False
    final_salary_header: str | None = None


MappingMode = Literal["output1", "output2"]


def inspect_bank_accounts_for_mapping(
    current_path: Path,
    previous_path: Path,
    *,
    factory: str = "factory1",
    smart_mapping: bool = True,
) -> dict[str, Any]:
    """Compare the mapping target with the saved bank registry.

    The previous workbook is only a candidate source.  The bank registry is
    the authoritative source used when Output 2 is generated, so this helper
    deliberately reports differences instead of silently copying an account
    from the old workbook.
    """
    from app.services.bank_account_store import get_saved_account_number, normalize_account_number

    # Use the same compaction as the real mapping. Owner-edited workbooks can
    # retain tens of thousands of style-only rows; loading that inflated sheet
    # twice here otherwise leaves the UI stuck at "Đang gán" before mapping.
    with TemporaryDirectory(prefix="attendance-owner-inspect-") as temp_dir, ExitStack() as stack:
        temp_root = Path(temp_dir)
        prepared_current_path, ignored_current_rows = _prepare_mapping_source_workbook(
            current_path,
            temp_root / "current",
        )
        prepared_previous_path, ignored_previous_rows = _prepare_mapping_source_workbook(
            previous_path,
            temp_root / "previous",
        )
        current_wb = load_workbook(prepared_current_path, data_only=False)
        stack.callback(current_wb.close)
        previous_wb = load_workbook(prepared_previous_path, data_only=False)
        stack.callback(previous_wb.close)
        previous_values_wb = load_workbook(prepared_previous_path, data_only=True)
        stack.callback(previous_values_wb.close)

        current_ws = current_wb.active
        previous_ws = previous_wb.active
        previous_values_ws = previous_values_wb[previous_ws.title]
        current_blocks = _summary_blocks_by_code(current_ws)
        if not current_blocks:
            raise ValueError("File tháng mới chưa có vùng Tổng giờ công / Mã / Tên nhân viên để kiểm tra")

        previous_records = _owner_records_by_code(previous_ws, previous_values_ws) if smart_mapping else {}
        current_codes = sorted(current_blocks, key=_code_sort_key)
        missing: list[dict[str, str]] = []
        changed: list[dict[str, str]] = []
        for code in current_codes:
            saved = normalize_account_number(get_saved_account_number(factory, code))
            candidate = normalize_account_number(
                (previous_records.get(code).bank_account if previous_records.get(code) else "")
            )
            source_name = str(previous_records.get(code).name or "").strip() if previous_records.get(code) else ""
            item = {
                "employee_code": code,
                "name": source_name,
                "saved_account": saved,
                "candidate_account": candidate,
            }
            if not saved:
                missing.append(item)
            elif candidate and candidate != saved:
                changed.append(item)

        return {
            "factory": factory,
            "current_count": len(current_codes),
            "missing_bank_accounts": missing,
            "changed_bank_accounts": changed,
            "missing_count": len(missing),
            "changed_count": len(changed),
            "ignored_trailing_style_rows": ignored_current_rows + ignored_previous_rows,
        }


def map_owner_data_to_current_workbook(
    current_path: Path,
    previous_path: Path,
    output_path: Path,
    mode: MappingMode = "output2",
    smart_mapping: bool = True,
    factory: str = "factory1",
) -> dict[str, Any]:
    with TemporaryDirectory(prefix="attendance-owner-map-") as temp_dir, ExitStack() as stack:
        prepared_previous_path, ignored_trailing_style_rows = _prepare_mapping_source_workbook(
            previous_path,
            Path(temp_dir),
        )
        current_wb = load_workbook(current_path)
        stack.callback(current_wb.close)
        previous_wb = load_workbook(prepared_previous_path)
        stack.callback(previous_wb.close)
        previous_values_wb = load_workbook(prepared_previous_path, data_only=True)
        stack.callback(previous_values_wb.close)

        current_ws = current_wb.active
        previous_ws = previous_wb.active
        previous_values_ws = previous_values_wb[previous_ws.title]

        previous_records = (
            _owner_records_by_code(previous_ws, previous_values_ws)
            if smart_mapping
            else {}
        )
        previous_blocks = (
            {}
            if smart_mapping
            else _summary_blocks_by_code(previous_ws)
        )
        current_blocks = _summary_blocks_by_code(current_ws)
        if not current_blocks:
            raise ValueError("File tháng mới chưa có vùng Tổng giờ công / Mã / Tên nhân viên để gán")
        if smart_mapping and not previous_records:
            raise ValueError(
                "File tháng cũ chưa có dữ liệu nhân viên có thể nhận diện theo Mã và tiêu đề cột"
            )
        if not smart_mapping and not previous_blocks:
            raise ValueError("File tháng cũ chưa có vùng dữ liệu theo mã nhân viên")

        previous_codes = set(previous_records if smart_mapping else previous_blocks)
        current_codes = set(current_blocks)
        matched_codes = sorted(previous_codes & current_codes, key=_code_sort_key)
        new_codes = sorted(current_codes - previous_codes, key=_code_sort_key)
        inactive_codes = sorted(previous_codes - current_codes, key=_code_sort_key)

        deduction_review_codes: list[str] = []
        legacy_template = _best_template(previous_blocks) if previous_blocks else None
        for code, target_block in current_blocks.items():
            matched_source = (
                previous_records.get(code)
                if smart_mapping
                else previous_blocks.get(code)
            )
            if mode == "output2":
                needs_deduction_review = _write_reformed_owner_area(
                    previous_ws,
                    previous_values_ws,
                    current_ws,
                    matched_source,
                    target_block,
                )
                if needs_deduction_review:
                    deduction_review_codes.append(code)
            elif not smart_mapping and legacy_template is not None:
                _copy_owner_area(
                    previous_ws,
                    current_ws,
                    matched_source or legacy_template,
                    target_block,
                    copy_values=matched_source is not None,
                    mode=mode,
                )
            # In smart mode Output 1 is already the current attendance
            # workbook, so it stays untouched.

        if mode == "output2":
            _write_monthly_grand_total(current_ws, list(current_blocks.values()))

        # Bank accounts are deliberately managed by the Bank screen/Word
        # import. Never carry an account number from the owner mapping file;
        # Output 2 receives only the factory-specific bank registry below.
        _clear_mapped_bank_account_cells(current_ws, list(current_blocks.values()), mode)
        if mode == "output2":
            _apply_saved_bank_accounts(current_ws, list(current_blocks.values()), factory, mode)

        _enable_formula_recalculation(current_wb)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        current_wb.save(output_path)
        return {
            "mode": mode,
            "smart_mapping": smart_mapping,
            "matched_count": len(matched_codes),
            "new_count": len(new_codes),
            "inactive_count": len(inactive_codes),
            "matched_codes": matched_codes,
            "new_codes": new_codes,
            "inactive_codes": inactive_codes,
            "deduction_review_count": len(deduction_review_codes),
            "deduction_review_codes": sorted(deduction_review_codes, key=_code_sort_key),
            "ignored_trailing_style_rows": ignored_trailing_style_rows,
        }


def _prepare_mapping_source_workbook(source_path: Path, temp_dir: Path) -> tuple[Path, int]:
    """Remove only excessive trailing rows that contain style but no value/formula.

    Some Excel files keep tens of thousands of formatted empty rows in sheet XML.
    Loading those rows twice with openpyxl makes mapping look frozen. The source
    file is never modified; a compact temporary copy is used for reading.
    """

    compacted_parts: dict[str, bytes] = {}
    ignored_rows = 0
    try:
        with ZipFile(source_path, "r") as source_zip:
            for name in source_zip.namelist():
                if not name.startswith(_WORKSHEET_XML_PREFIX) or not name.endswith(".xml"):
                    continue
                original_xml = source_zip.read(name)
                compacted_xml, removed = _compact_trailing_style_rows(original_xml)
                if removed:
                    compacted_parts[name] = compacted_xml
                    ignored_rows += removed
    except (OSError, ValueError):
        return source_path, 0

    if not compacted_parts:
        return source_path, 0

    temp_dir.mkdir(parents=True, exist_ok=True)
    compacted_path = temp_dir / f"mapping-source{source_path.suffix.lower()}"
    with ZipFile(source_path, "r") as source_zip, ZipFile(
        compacted_path,
        "w",
        compression=ZIP_DEFLATED,
    ) as target_zip:
        for item in source_zip.infolist():
            data = compacted_parts.get(item.filename)
            if data is None:
                data = source_zip.read(item.filename)
            target_zip.writestr(item, data)
    return compacted_path, ignored_rows


def _compact_trailing_style_rows(worksheet_xml: bytes) -> tuple[bytes, int]:
    rows: list[tuple[re.Match[bytes], int]] = []
    last_meaningful_row = 0
    max_row = 0
    for match in _ROW_ELEMENT_RE.finditer(worksheet_xml):
        row_number_match = _ROW_NUMBER_RE.search(match.group(0))
        if not row_number_match:
            continue
        row_number = int(row_number_match.group(1))
        rows.append((match, row_number))
        max_row = max(max_row, row_number)
        if _SEMANTIC_CELL_CONTENT_RE.search(match.group(0)):
            last_meaningful_row = max(last_meaningful_row, row_number)

    if (
        not last_meaningful_row
        or max_row - last_meaningful_row < MIN_TRAILING_STYLE_ROWS_TO_COMPACT
    ):
        return worksheet_xml, 0

    trailing_rows = [(match, row_number) for match, row_number in rows if row_number > last_meaningful_row]
    if not trailing_rows:
        return worksheet_xml, 0

    chunks: list[bytes] = []
    cursor = 0
    for match, _ in trailing_rows:
        chunks.append(worksheet_xml[cursor : match.start()])
        cursor = match.end()
    chunks.append(worksheet_xml[cursor:])
    compacted_xml = b"".join(chunks)
    compacted_xml = _shrink_worksheet_dimension(compacted_xml, last_meaningful_row)
    return compacted_xml, len(trailing_rows)


def _shrink_worksheet_dimension(worksheet_xml: bytes, last_row: int) -> bytes:
    def replace_dimension(match: re.Match[bytes]) -> bytes:
        reference = match.group(2)
        parts = reference.split(b":", 1)
        end_reference = parts[-1]
        compacted_end = re.sub(br"\d+$", str(last_row).encode("ascii"), end_reference)
        compacted_reference = (
            parts[0] + b":" + compacted_end
            if len(parts) == 2
            else compacted_end
        )
        return match.group(1) + compacted_reference + match.group(3)

    return _DIMENSION_RE.sub(replace_dimension, worksheet_xml, count=1)


def _owner_records_by_code(source_ws, source_values_ws=None) -> dict[str, OwnerRecord]:
    """Read old owner data by semantic headers and employee code.

    The source layout is intentionally not assumed to use seven-row blocks or
    fixed Excel columns. A header row establishes the meaning of each column;
    employee codes found below it establish the record rows.
    """

    values_ws = source_values_ws or source_ws
    header_sections: list[tuple[int, dict[str, int], str | None]] = []
    for row in range(1, source_ws.max_row + 1):
        fields, final_salary_header = _semantic_header_fields(source_ws, row)
        if "code" not in fields:
            continue
        business_fields = set(fields) - {"code", "total"}
        if not business_fields:
            continue
        header_sections.append((row, fields, final_salary_header))

    records: dict[str, OwnerRecord] = {}
    for index, (header_row, fields, final_salary_header) in enumerate(header_sections):
        next_header_row = (
            header_sections[index + 1][0]
            if index + 1 < len(header_sections)
            else source_ws.max_row + 1
        )
        section_end_row = min(next_header_row - 1, header_row + 24)
        code_col = fields["code"]
        for result_row in range(header_row + 1, section_end_row + 1):
            code = _normalize_employee_code(values_ws.cell(result_row, code_col).value)
            if not code:
                code = _normalize_employee_code(source_ws.cell(result_row, code_col).value)
            if not code:
                continue
            if code in records:
                raise ValueError(f"Phát hiện mã nhân viên bị trùng trong file tháng cũ: {code}")

            name = _semantic_name_value(
                source_ws,
                values_ws,
                result_row,
                code_col,
                fields,
            )
            salary = _semantic_field_value(
                source_ws,
                values_ws,
                result_row,
                fields.get("salary"),
            )
            bonus = _semantic_field_value(
                source_ws,
                values_ws,
                result_row,
                fields.get("bonus"),
            )
            bank_account = _semantic_bank_account_value(
                source_ws,
                values_ws,
                header_row,
                result_row,
                code_col,
                fields,
            )
            start_work_note, note = _semantic_start_work_and_note(
                source_ws,
                values_ws,
                result_row,
                section_end_row,
                code_col,
                fields,
            )
            has_previous_deduction = any(
                _semantic_amount_present(
                    source_ws,
                    values_ws,
                    result_row,
                    fields.get(field_name),
                )
                for field_name in ("legacy_deduction", "advance", "nq_penalty", "nq_rate")
            )
            records[code] = OwnerRecord(
                code=code,
                name=name,
                bank_account=bank_account,
                start_work_note=start_work_note,
                salary=salary,
                bonus=bonus,
                note=note,
                has_previous_deduction=has_previous_deduction,
                final_salary_header=final_salary_header,
            )
    return records


def _semantic_header_fields(ws, row: int) -> tuple[dict[str, int], str | None]:
    fields: dict[str, int] = {}
    final_salary_header = None
    for col in range(1, ws.max_column + 1):
        raw_value = ws.cell(row, col).value
        label = _normalize_label(raw_value)
        if not label:
            continue
        field_name = _semantic_field_name(label)
        if field_name and field_name not in fields:
            fields[field_name] = col
        if label.startswith("luong thang") and final_salary_header is None:
            final_salary_header = str(raw_value)
    return fields, final_salary_header


def _semantic_field_name(label: str) -> str | None:
    if label in SUMMARY_CODE_LABELS:
        return "code"
    if label in SUMMARY_TOTAL_LABELS:
        return "total"
    if label in NAME_LABELS:
        return "name"
    if label in SALARY_LABELS:
        return "salary"
    if label in BONUS_LABELS:
        return "bonus"
    if label in NOTE_LABELS:
        return "note"
    if label in LEGACY_DEDUCTION_LABELS:
        return "legacy_deduction"
    if label in ADVANCE_LABELS:
        return "advance"
    if label in NQ_PENALTY_LABELS:
        return "nq_penalty"
    if label in NQ_RATE_LABELS:
        return "nq_rate"
    if label.startswith("luong thang"):
        return "final_salary"
    return None


def _normalize_employee_code(value: object) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else None
    text = str(value).strip()
    if not text or text.startswith("="):
        return None
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    if not re.fullmatch(r"(?=.*\d)[A-Za-z0-9_-]{1,20}", text):
        return None
    return text


def _semantic_name_value(source_ws, values_ws, row: int, code_col: int, fields: dict[str, int]) -> str | None:
    name_col = fields.get("name")
    if name_col is not None:
        value = _plain_text_value(values_ws.cell(row, name_col).value)
        if value:
            return value

    first_business_col = min(
        (col for name, col in fields.items() if name not in {"code", "total"} and col > code_col),
        default=min(source_ws.max_column + 1, code_col + 6),
    )
    end_col = min(source_ws.max_column, max(code_col + 1, first_business_col - 1))
    for col in range(code_col + 1, end_col + 1):
        value = _plain_text_value(values_ws.cell(row, col).value)
        if value and _semantic_field_name(_normalize_label(value)) is None:
            return value
    return None


def _semantic_field_value(source_ws, values_ws, row: int, col: int | None):
    if col is None:
        return None
    cached = values_ws.cell(row, col).value
    if cached is not None:
        return cached
    raw = source_ws.cell(row, col)
    return None if raw.data_type == "f" else raw.value


def _semantic_amount_present(source_ws, values_ws, row: int, col: int | None) -> bool:
    if col is None:
        return False
    if _has_meaningful_amount(values_ws.cell(row, col).value):
        return True
    raw = source_ws.cell(row, col)
    return raw.data_type == "f" and values_ws.cell(row, col).value is None


def _semantic_bank_account_value(
    source_ws,
    values_ws,
    header_row: int,
    result_row: int,
    code_col: int,
    fields: dict[str, int],
) -> str | None:
    """Read the account number kept one row above the employee name in legacy files."""
    name_col = fields.get("name") or code_col + 1
    for row in range(header_row + 1, result_row):
        value = _bank_account_text(values_ws.cell(row, name_col).value)
        if value:
            return value
    return None


def _bank_account_text(value: object) -> str | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            return None
        text = str(int(value))
    else:
        text = re.sub(r"\s+", "", str(value))
    return text if re.fullmatch(r"\d{8,20}", text) else None


def _semantic_start_work_and_note(
    source_ws,
    values_ws,
    result_row: int,
    section_end_row: int,
    code_col: int,
    fields: dict[str, int],
) -> str | None:
    name_col = fields.get("name") or code_col + 1
    note_row = result_row + 1
    start_work_note = _plain_text_value(values_ws.cell(note_row, name_col).value) if note_row <= section_end_row else None
    note_parts: list[str] = []
    if start_work_note and " | " in start_work_note:
        start_work_note, inline_note = start_work_note.split(" | ", maxsplit=1)
        if inline_note.strip():
            note_parts.append(inline_note.strip())

    # Legacy files put the comment in any column to the right of the start
    # work field (AJ, AM, ...). Preserve it separately from Bắt đầu làm.
    if note_row <= section_end_row:
        for col in range(name_col + 1, source_ws.max_column + 1):
            value = _plain_text_value(values_ws.cell(note_row, col).value)
            if value and value not in note_parts:
                note_parts.append(value)

    if start_work_note or note_parts:
        return start_work_note, " | ".join(note_parts) or None

    start_col = code_col + 1
    end_col = max(fields.values(), default=min(source_ws.max_column, code_col + 12))
    for row in range(result_row + 1, section_end_row + 1):
        for col in range(start_col, min(source_ws.max_column, end_col) + 1):
            value = _plain_text_value(values_ws.cell(row, col).value)
            if not value or len(value) < 3:
                continue
            if _semantic_field_name(_normalize_label(value)) is None:
                return None, value
    return None, None


def _plain_text_value(value: object) -> str | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text or text.startswith("=") or text == "?":
        return None
    return text


def _summary_blocks_by_code(ws) -> dict[str, SummaryBlock]:
    blocks: dict[str, SummaryBlock] = {}
    for row in range(1, ws.max_row + 1):
        total_col = _find_summary_total_col(ws, row)
        if not total_col:
            continue
        code_col = _find_code_col(ws, row, total_col)
        if not code_col:
            continue
        result_row = row + 5
        code = str(ws.cell(result_row, code_col).value or "").strip()
        if not code:
            continue
        if code in blocks:
            raise ValueError(f"Phát hiện mã nhân viên bị trùng trong file: {code}")
        data_start_col = code_col + 1
        data_end_col = _find_data_end_col(ws, row, result_row, data_start_col)
        blocks[code] = SummaryBlock(
            code=code,
            header_row=row,
            result_row=result_row,
            total_col=total_col,
            code_col=code_col,
            data_start_col=data_start_col,
            data_end_col=data_end_col,
        )
    return blocks


def _find_summary_total_col(ws, row: int) -> int | None:
    for col in range(SUMMARY_MIN_COL, ws.max_column + 1):
        if _normalize_label(ws.cell(row, col).value) in SUMMARY_TOTAL_LABELS:
            return col
    return None


def _find_code_col(ws, row: int, total_col: int) -> int | None:
    for col in range(total_col + 1, min(ws.max_column, total_col + 4) + 1):
        if _normalize_label(ws.cell(row, col).value) in SUMMARY_CODE_LABELS:
            return col
    return None


def _find_data_end_col(ws, header_row: int, result_row: int, start_col: int) -> int:
    end_col = max(start_col, ws.max_column)
    while end_col > start_col:
        has_content = any(
            ws.cell(row, end_col).value not in (None, "")
            for row in range(header_row, header_row + SUMMARY_ROWS)
        )
        if has_content:
            return end_col
        end_col -= 1
    return start_col


def _best_template(blocks: dict[str, SummaryBlock]) -> SummaryBlock:
    return max(blocks.values(), key=lambda block: block.data_end_col - block.data_start_col)


def _write_reformed_owner_area(
    source_ws,
    source_values_ws,
    target_ws,
    source: OwnerRecord | SummaryBlock | None,
    target: SummaryBlock,
) -> bool:
    total_col = target.total_col
    penalty_rate_col = total_col + 1
    code_col = total_col + 2
    name_col = total_col + 3
    salary_col = total_col + 4
    daily_salary_col = total_col + 5
    hourly_salary_col = total_col + 6
    work_days_col = total_col + 7
    overtime_col = total_col + 8
    bonus_col = total_col + 9
    nq_penalty_col = total_col + 10
    advance_col = total_col + 11
    final_salary_col = total_col + 12
    result_row = target.result_row
    note_row = target.header_row + SUMMARY_ROWS - 1

    current_name = normalize_employee_name(_first_non_empty_text(
        target_ws,
        result_row,
        target.code_col + 1,
        max(target.code_col + 1, target.data_end_col),
    ))
    source_name = None
    source_start_work_note = None
    source_salary = None
    source_bonus = None
    source_note = None
    needs_deduction_review = False
    if isinstance(source, OwnerRecord):
        source_name = normalize_employee_name(source.name)
        source_start_work_note = source.start_work_note
        source_salary = source.salary
        source_bonus = source.bonus
        source_note = source.note
        needs_deduction_review = source.has_previous_deduction
    elif source is not None:
        source_name = normalize_employee_name(_first_non_empty_text(
            source_values_ws,
            source.result_row,
            source.code_col + 1,
            max(source.code_col + 1, source.data_end_col),
        ))
        source_salary = _fixed_field_value(
            source_ws,
            source_values_ws,
            source,
            SALARY_LABELS,
        )
        source_bonus = _fixed_field_value(
            source_ws,
            source_values_ws,
            source,
            BONUS_LABELS,
        )
        source_note = _first_non_empty_text(
            source_values_ws,
            source.header_row + SUMMARY_ROWS - 1,
            source.data_start_col,
            source.data_end_col,
        )
        needs_deduction_review = _source_has_previous_deduction(
            source_ws,
            source_values_ws,
            source,
        )

    _unmerge_overlapping(
        target_ws,
        target.header_row,
        total_col,
        note_row,
        max(final_salary_col, target.data_end_col),
    )
    _clear_owner_tail(
        target_ws,
        target.header_row,
        total_col,
        note_row,
        max(final_salary_col, target.data_end_col),
    )
    _format_reformed_owner_area(
        target_ws,
        target.header_row,
        note_row,
        total_col,
        final_salary_col,
    )

    headers = {
        total_col: "Tổng giờ công",
        penalty_rate_col: "Mức tiền phạt NQ trên giờ công (đ)",
        code_col: "Mã",
        name_col: "Tên nhân viên / Ghi chú",
        salary_col: "Mức Lương",
        daily_salary_col: "Lương 1 Ngày Công",
        hourly_salary_col: "Lương 1 Giờ Công",
        work_days_col: "Số Ngày Đi Làm",
        overtime_col: "Giờ làm thêm",
        bonus_col: "Thưởng",
        nq_penalty_col: "Phạt NQ",
        advance_col: "Ứng Lương",
        final_salary_col: _final_salary_header(source_ws, source),
    }
    for col, value in headers.items():
        target_ws.cell(target.header_row, col).value = value

    first_day = get_column_letter(1)
    last_day = get_column_letter(total_col - 1)
    total_letter = get_column_letter(total_col)
    penalty_rate_letter = get_column_letter(penalty_rate_col)
    salary_letter = get_column_letter(salary_col)
    daily_salary_letter = get_column_letter(daily_salary_col)
    hourly_salary_letter = get_column_letter(hourly_salary_col)
    work_days_letter = get_column_letter(work_days_col)
    overtime_letter = get_column_letter(overtime_col)
    bonus_letter = get_column_letter(bonus_col)
    nq_penalty_letter = get_column_letter(nq_penalty_col)
    advance_letter = get_column_letter(advance_col)

    target_ws.cell(result_row, total_col).value = f"=SUM({first_day}{result_row}:{last_day}{result_row})"
    target_ws.cell(result_row, penalty_rate_col).value = None
    target_ws.cell(result_row, code_col).value = target.code
    target_ws.cell(result_row, name_col).value = normalize_employee_name(source_name or current_name)
    target_ws.cell(result_row, salary_col).value = source_salary
    target_ws.cell(result_row, daily_salary_col).value = (
        f'=IF({salary_letter}{result_row}>0,{salary_letter}{result_row}/26,'
        f'IF({hourly_salary_letter}{result_row}>0,{hourly_salary_letter}{result_row}*8,""))'
    )
    target_ws.cell(result_row, hourly_salary_col).value = (
        f'=IF({salary_letter}{result_row}>0,{salary_letter}{result_row}/208,'
        f'IF({daily_salary_letter}{result_row}>0,{daily_salary_letter}{result_row}/8,""))'
    )
    target_ws.cell(result_row, work_days_col).value = (
        f"=SUM({first_day}{result_row}:{last_day}{result_row})/8"
    )
    target_ws.cell(result_row, overtime_col).value = (
        f"=SUM({first_day}{result_row + 1}:{last_day}{result_row + 1})"
    )
    target_ws.cell(result_row, bonus_col).value = source_bonus
    target_ws.cell(result_row, nq_penalty_col).value = (
        f'=IF(ISNUMBER({penalty_rate_letter}{result_row}),'
        f"{penalty_rate_letter}{result_row}*{total_letter}{result_row},0)"
    )
    advance_cell = target_ws.cell(result_row, advance_col)
    advance_cell.value = "?" if needs_deduction_review else None
    if needs_deduction_review:
        advance_cell.comment = Comment(
            "Tháng trước mã này có dữ liệu tại cột Ứng Lương/Phạt. "
            "Số cũ không được tự chuyển sang tháng mới. Chủ cần kiểm tra và nhập lại nếu tháng này có phát sinh.",
            "AttendanceSystem",
        )
        advance_cell.fill = PatternFill("solid", fgColor=GREEN)
        advance_cell.font = Font(name=FONT_NAME, bold=True, size=12, color=RED)
    target_ws.cell(result_row, final_salary_col).value = (
        f'=IF({daily_salary_letter}{result_row}>0,{daily_salary_letter}{result_row},'
        f'IF({salary_letter}{result_row}>0,{salary_letter}{result_row}/26,'
        f'IF({hourly_salary_letter}{result_row}>0,{hourly_salary_letter}{result_row}*8,0)))'
        f"*{work_days_letter}{result_row}"
        f"+({overtime_letter}{result_row}*{hourly_salary_letter}{result_row}*1.5)"
        f"-IF(ISNUMBER({advance_letter}{result_row}),{advance_letter}{result_row},0)"
        f"-{nq_penalty_letter}{result_row}"
        f"+IF(ISNUMBER({bonus_letter}{result_row}),{bonus_letter}{result_row},0)"
    )

    bank_row = result_row - 1
    # Bank accounts are managed by the Bank screen/Word import. A mapped
    # workbook is never authoritative for this field.
    target_ws.cell(bank_row, name_col).value = None
    target_ws.cell(bank_row, name_col).number_format = "@"
    target_ws.cell(bank_row, name_col).alignment = Alignment(horizontal="left", vertical="center")
    for col in range(name_col, final_salary_col + 1):
        note_part = target_ws.cell(note_row, col)
        note_part.value = None
        note_part.fill = PatternFill("solid", fgColor=NOTE_FILL)
    target_ws.merge_cells(
        start_row=note_row,
        start_column=name_col + 1,
        end_row=note_row,
        end_column=final_salary_col,
    )
    start_work_cell = target_ws.cell(note_row, name_col)
    start_work_cell.value = source_start_work_note
    start_work_cell.fill = PatternFill("solid", fgColor=START_WORK_FILL)
    start_work_cell.font = Font(name=FONT_NAME, size=9, bold=True, color=BLACK)
    start_work_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell = target_ws.cell(note_row, name_col + 1)
    note_cell.value = source_note
    note_cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    note_cell.font = Font(name=FONT_NAME, size=9, color=RED)
    note_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )
    _set_note_row_height(target_ws, note_row, source_start_work_note, source_note, name_col, final_salary_col)

    for col in {
        salary_col,
        daily_salary_col,
        hourly_salary_col,
        bonus_col,
        nq_penalty_col,
        advance_col,
        final_salary_col,
    }:
        target_ws.cell(result_row, col).number_format = MONEY_FORMAT
    for col in {total_col, penalty_rate_col, work_days_col, overtime_col}:
        target_ws.cell(result_row, col).number_format = FLEXIBLE_NUMBER_FORMAT

    return needs_deduction_review


def _mapped_bank_account_col(block: SummaryBlock, mode: MappingMode) -> int:
    """Return the bank-account column for the workbook layout being written.

    Output 1 keeps the current attendance layout, where the account cell is
    immediately to the right of the code column. Output 2 inserts the NQ-rate
    column before the code, so its account cell is three columns after the
    total-hours column. Using ``block.code_col + 1`` for both layouts writes
    the account into Output 2's actual code column.
    """
    return block.total_col + 3 if mode == "output2" else block.code_col + 1


def _clear_mapped_bank_account_cells(
    ws,
    blocks: list[SummaryBlock],
    mode: MappingMode,
) -> None:
    for block in blocks:
        cell = ws.cell(block.result_row - 1, _mapped_bank_account_col(block, mode))
        cell.value = None
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_saved_bank_accounts(
    ws,
    blocks: list[SummaryBlock],
    factory: str,
    mode: MappingMode,
) -> None:
    from app.services.bank_account_store import get_saved_account_number

    for block in blocks:
        cell = ws.cell(block.result_row - 1, _mapped_bank_account_col(block, mode))
        cell.value = get_saved_account_number(factory, block.code)
        cell.number_format = "@"
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _set_note_row_height(ws, row: int, start_value: object, note_value: object, start_col: int, end_col: int) -> None:
    """Estimate the height needed for separate start-work and comment cells."""
    start_width = float(ws.column_dimensions[get_column_letter(start_col)].width or 10)
    note_width = sum(
        float(ws.column_dimensions[get_column_letter(col)].width or 10)
        for col in range(start_col + 1, end_col + 1)
    )
    line_count = max(
        _wrapped_line_count(start_value, start_width),
        _wrapped_line_count(note_value, note_width),
    )
    required_height = max(20, 15 * line_count + 5)
    ws.row_dimensions[row].height = max(float(ws.row_dimensions[row].height or 0), required_height)


def _wrapped_line_count(value: object, width: float) -> int:
    chars_per_line = max(12, int(width * 0.72))
    return max((ceil(len(line) / chars_per_line) or 1 for line in (str(value or "").splitlines() or [""])), default=1)


def _write_monthly_grand_total(ws, blocks: list[SummaryBlock]) -> None:
    if not blocks:
        return

    ordered = sorted(blocks, key=lambda block: block.header_row)
    final_salary_col = ordered[0].total_col + 12
    label_start_col = max(1, ordered[0].total_col - 27)
    total_row = max(block.header_row + SUMMARY_ROWS for block in ordered)
    label = _period_label_from_blocks(ws, ordered)

    _unmerge_overlapping(
        ws,
        total_row,
        label_start_col,
        total_row,
        final_salary_col,
    )
    for col in range(label_start_col, final_salary_col + 1):
        cell = ws.cell(total_row, col)
        cell.value = None
        cell.fill = PatternFill("solid", fgColor=YELLOW)
        cell.border = Border(
            left=Side(style="thin", color=BLACK),
            right=Side(style="thin", color=BLACK),
            top=Side(style="thin", color=BLACK),
            bottom=Side(style="thin", color=BLACK),
        )
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=BLACK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(
        start_row=total_row,
        start_column=label_start_col,
        end_row=total_row,
        end_column=final_salary_col - 1,
    )
    ws.cell(total_row, label_start_col).value = f"Tổng tháng {label}" if label else "Tổng tháng"
    salary_letter = get_column_letter(final_salary_col)
    salary_rows = ",".join(f"{salary_letter}{block.result_row}" for block in ordered)
    ws.cell(total_row, final_salary_col).value = f"=SUM({salary_rows})"
    ws.cell(total_row, final_salary_col).number_format = MONEY_FORMAT
    ws.row_dimensions[total_row].height = max(float(ws.row_dimensions[total_row].height or 0), 20)


def _period_label_from_blocks(ws, blocks: list[SummaryBlock]) -> str:
    for block in blocks:
        text = str(ws.cell(block.header_row - 1, 3).value or "")
        match = re.search(r"(20\d{2})-(\d{1,2})-\d{1,2}", text)
        if match:
            return f"{int(match.group(2))}/{match.group(1)}"
    return ""


def _format_reformed_owner_area(
    ws,
    header_row: int,
    note_row: int,
    total_col: int,
    final_salary_col: int,
) -> None:
    thin = Side(style="thin", color=BLACK)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(header_row, note_row + 1):
        for col in range(total_col, final_salary_col + 1):
            cell = ws.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
                shrink_to_fit=row == header_row,
            )
            if col <= total_col + 1:
                fill_color = WHITE
            elif col <= total_col + 3:
                fill_color = YELLOW
            else:
                fill_color = YELLOW if row == header_row else GREEN
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(
                name=FONT_NAME,
                size=9,
                bold=row == header_row,
                color=RED if row == header_row and col in {total_col, total_col + 1, total_col + 2} else BLACK,
            )

    result_row = header_row + 5
    for col in range(total_col, final_salary_col + 1):
        cell = ws.cell(result_row, col)
        cell.font = Font(
            name=FONT_NAME,
            size=10,
            bold=col in {total_col, total_col + 1, total_col + 2, total_col + 3, final_salary_col},
            color=RED if col in {total_col, total_col + 1, total_col + 2} else BLACK,
        )

    widths = {
        total_col: 16,
        total_col + 1: 38,
        total_col + 2: 8,
        total_col + 3: 28,
        total_col + 4: 14,
        total_col + 5: 16,
        total_col + 6: 16,
        total_col + 7: 14,
        total_col + 8: 11,
        total_col + 9: 12,
        total_col + 10: 12,
        total_col + 11: 14,
        total_col + 12: 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(
            float(ws.column_dimensions[get_column_letter(col)].width or 0),
            width,
        )
    ws.row_dimensions[header_row].height = max(
        float(ws.row_dimensions[header_row].height or 0),
        28,
    )


def _fixed_field_value(
    source_ws,
    source_values_ws,
    source: SummaryBlock,
    labels: set[str],
):
    col = _find_header_col(source_ws, source, labels)
    if col is None:
        return None
    cached = source_values_ws.cell(source.result_row, col).value
    if cached is not None:
        return cached
    raw = source_ws.cell(source.result_row, col)
    return None if raw.data_type == "f" else raw.value


def _source_has_previous_deduction(
    source_ws,
    source_values_ws,
    source: SummaryBlock,
) -> bool:
    for labels in (
        LEGACY_DEDUCTION_LABELS,
        ADVANCE_LABELS,
        NQ_PENALTY_LABELS,
        NQ_RATE_LABELS,
    ):
        col = _find_header_col(source_ws, source, labels)
        if col is None:
            continue
        raw_cell = source_ws.cell(source.result_row, col)
        cached_cell = source_values_ws.cell(source.result_row, col)
        if _has_meaningful_amount(cached_cell.value):
            return True
        if raw_cell.data_type == "f" and cached_cell.value is None:
            return True
    return False


def _find_header_col(
    ws,
    block: SummaryBlock,
    labels: set[str],
) -> int | None:
    for col in range(block.total_col, block.data_end_col + 1):
        if _normalize_label(ws.cell(block.header_row, col).value) in labels:
            return col
    return None


def _first_non_empty_text(ws, row: int, start_col: int, end_col: int) -> str | None:
    for col in range(start_col, end_col + 1):
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip() and not value.startswith("="):
            return value.strip()
    return None


def _has_meaningful_amount(value: object) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return abs(float(value)) > 1e-9
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text or text in {"-", "0", "0.0", "0.00"}:
        return False
    try:
        return abs(float(text)) > 1e-9
    except ValueError:
        return True


def _final_salary_header(source_ws, source: OwnerRecord | SummaryBlock | None) -> str:
    if isinstance(source, OwnerRecord):
        return source.final_salary_header or "Lương Tháng"
    if source is not None:
        for col in range(source.total_col, source.data_end_col + 1):
            value = source_ws.cell(source.header_row, col).value
            if "luong thang" in _normalize_label(value):
                return str(value)
    return "Lương Tháng"


def _copy_owner_area(
    source_ws,
    target_ws,
    source: SummaryBlock,
    target: SummaryBlock,
    copy_values: bool,
    mode: MappingMode,
) -> None:
    source_start_col = source.total_col
    source_end_col = source.data_start_col if mode == "output1" else source.data_end_col
    target_start_col = target.total_col
    target_end_col = target_start_col + (source_end_col - source_start_col)
    col_offset = target_start_col - source_start_col
    row_offset = target.header_row - source.header_row
    _unmerge_overlapping(target_ws, target.header_row, target_start_col, target.header_row + SUMMARY_ROWS - 1, target_end_col)

    for source_col in range(source_start_col, source_end_col + 1):
        target_col = source_col + col_offset
        source_letter = get_column_letter(source_col)
        target_letter = get_column_letter(target_col)
        target_ws.column_dimensions[target_letter].width = source_ws.column_dimensions[source_letter].width
        for source_row in range(source.header_row, source.header_row + SUMMARY_ROWS):
            target_row = source_row + row_offset
            source_cell = source_ws.cell(source_row, source_col)
            target_cell = target_ws.cell(target_row, target_col)
            _copy_cell_style(source_cell, target_cell)
            target_cell.value = _mapped_value(
                source_cell,
                target_cell,
                target_code=target.code,
                source_header_row=source.header_row,
                source_result_row=source.result_row,
                source_code_col=source.code_col,
                copy_values=copy_values,
            )

    if target.data_end_col > target_end_col:
        _clear_owner_tail(target_ws, target.header_row, target_end_col + 1, target.header_row + SUMMARY_ROWS - 1, target.data_end_col)

    for merged in list(source_ws.merged_cells.ranges):
        if not _range_overlaps(
            merged.min_row,
            merged.min_col,
            merged.max_row,
            merged.max_col,
            source.header_row,
            source_start_col,
            source.header_row + SUMMARY_ROWS - 1,
            source_end_col,
        ):
            continue
        target_ws.merge_cells(
            start_row=merged.min_row + row_offset,
            start_column=merged.min_col + col_offset,
            end_row=merged.max_row + row_offset,
            end_column=merged.max_col + col_offset,
        )


def _mapped_value(
    source_cell,
    target_cell,
    target_code: str,
    source_header_row: int,
    source_result_row: int,
    source_code_col: int,
    copy_values: bool,
):
    value = source_cell.value
    if value is None:
        return None
    if source_cell.row == source_header_row:
        return value
    if source_cell.data_type == "f" and isinstance(value, str):
        return Translator(value, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
    if not copy_values:
        if source_cell.column == source_code_col and source_cell.row == source_result_row:
            return target_code
        return None
    return value


def _copy_cell_style(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy.copy(source_cell.protection)
    else:
        target_cell.font = Font()
        target_cell.fill = PatternFill(fill_type=None)
        target_cell.border = Border()
    if source_cell.hyperlink:
        target_cell._hyperlink = copy.copy(source_cell.hyperlink)
    if source_cell.comment:
        target_cell.comment = copy.copy(source_cell.comment)


def _clear_owner_tail(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    if min_col > max_col:
        return
    _unmerge_overlapping(ws, min_row, min_col, max_row, max_col)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.hyperlink = None
            cell.comment = None


def _unmerge_overlapping(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if _range_overlaps(merged.min_row, merged.min_col, merged.max_row, merged.max_col, min_row, min_col, max_row, max_col):
            ws.unmerge_cells(str(merged))


def _range_overlaps(a_min_row: int, a_min_col: int, a_max_row: int, a_max_col: int, b_min_row: int, b_min_col: int, b_max_row: int, b_max_col: int) -> bool:
    return not (a_max_row < b_min_row or a_min_row > b_max_row or a_max_col < b_min_col or a_min_col > b_max_col)


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("đ", "d")
    return " ".join(text.split())


def _code_sort_key(code: str) -> tuple[int, str]:
    return (0, f"{int(code):010d}") if code.isdigit() else (1, code)


def _enable_formula_recalculation(workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    for attr, value in (
        ("calcMode", "auto"),
        ("fullCalcOnLoad", True),
        ("forceFullCalc", True),
    ):
        try:
            setattr(calculation, attr, value)
        except Exception:
            continue
