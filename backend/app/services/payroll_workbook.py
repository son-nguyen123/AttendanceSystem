from copy import copy
from math import ceil
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.block_detector import detect_employee_blocks
from app.services.payroll_store import PayrollEntry, calculate_daily_salary, calculate_hourly_salary, calculate_monthly_salary, get_payroll_entry, normalize_employee_name
from app.services.workbook_processor import _center_attendance_numbers, _format_attendance_punch_row, _format_attendance_time_row, _format_title_area, export_processed_workbook


PAYROLL_START_COL = 32
PAYROLL_END_COL = 44
MONEY_FORMAT = '#,##0'
FLEXIBLE_NUMBER_FORMAT = "General"

YELLOW = "FFFF00"
GREEN = "C4D79B"
WHITE = "FFFFFF"
NOTE_FILL = "FFF4CC"
START_WORK_FILL = "DDEBF7"
RED = "C00000"
FONT_NAME = "Arial"


def preview_payroll(
    source_path: Path,
    review_overrides: list[dict] | None = None,
    profile_codes: set[str] | None = None,
    factory: str = "factory1",
) -> dict:
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_path = Path(temp_file.name)

    try:
        export_processed_workbook(source_path, temp_path, review_overrides=review_overrides, factory=factory)
        wb = load_workbook(temp_path, data_only=False)
        ws = _select_attendance_sheet(wb)
        blocks = detect_employee_blocks(ws)
        employees = [_build_employee_preview(ws, block, profile_codes=profile_codes, factory=factory) for block in blocks]
        return {"sheet_name": ws.title, "employees": employees}
    finally:
        temp_path.unlink(missing_ok=True)


def export_payroll_workbook(
    source_path: Path,
    output_path: Path,
    review_overrides: list[dict] | None = None,
    include_saved_data: bool = True,
    profile_codes: set[str] | None = None,
    factory: str = "factory1",
    source_payroll_values: dict[str, dict] | None = None,
    source_is_processed: bool = False,
) -> Path:
    temp_output1 = None
    if not source_is_processed:
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
            temp_output1 = Path(temp_file.name)

    try:
        processed_path = source_path
        if not source_is_processed:
            export_processed_workbook(source_path, temp_output1, review_overrides=review_overrides, factory=factory)
            processed_path = temp_output1
        wb = load_workbook(processed_path, data_only=False)
        ws = _select_attendance_sheet(wb)
        blocks = detect_employee_blocks(ws)

        for block in blocks:
            _format_attendance_time_row(ws, block)
            _format_attendance_punch_row(ws, block)
            _center_attendance_numbers(ws, block)
            source_values = (source_payroll_values or {}).get(block.employee_code)
            preview = _build_employee_preview(
                ws,
                block,
                include_saved_data=include_saved_data,
                profile_codes=profile_codes,
                factory=factory,
                source_values=source_values,
            )
            _write_payroll_block(ws, block, preview, source_values=source_values)

        _write_monthly_grand_total(ws, blocks)
        _set_payroll_column_widths(ws)
        _format_title_area(ws, PAYROLL_END_COL)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path
    finally:
        if temp_output1 is not None:
            temp_output1.unlink(missing_ok=True)


def apply_payroll_to_workbook(
    source_path: Path,
    output_path: Path,
    profile_codes: set[str] | None = None,
    factory: str = "factory1",
) -> Path:
    keep_vba = source_path.suffix.lower() == ".xlsm"
    wb = load_workbook(source_path, data_only=False, keep_vba=keep_vba)
    ws = _select_attendance_sheet(wb)
    blocks = detect_employee_blocks(ws)
    if not blocks:
        raise ValueError("Khong tim thay block cham cong trong file")

    for block in blocks:
        _format_attendance_time_row(ws, block)
        _format_attendance_punch_row(ws, block)
        _center_attendance_numbers(ws, block)
        preview = _build_employee_preview(ws, block, profile_codes=profile_codes, factory=factory)
        _write_payroll_block(ws, block, preview)

    _write_monthly_grand_total(ws, blocks)
    _set_payroll_column_widths(ws)
    _format_title_area(ws, PAYROLL_END_COL)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_monthly_grand_total(ws, blocks) -> None:
    if not blocks:
        return

    ordered = sorted(blocks, key=lambda block: block.header_row)
    total_row = max(block.header_row + 8 for block in ordered)
    label_start_col = max(1, PAYROLL_START_COL - 27)
    label = _month_label(ws.cell(row=ordered[0].header_row, column=3).value)

    _unmerge_overlapping(ws, total_row, label_start_col, total_row, PAYROLL_END_COL)
    border = _thin_border()
    for col in range(label_start_col, PAYROLL_END_COL + 1):
        cell = ws.cell(total_row, col)
        cell.value = None
        cell.fill = PatternFill("solid", fgColor=YELLOW)
        cell.border = border
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _safe_merge(ws, total_row, label_start_col, total_row, PAYROLL_END_COL - 1)
    ws.cell(total_row, label_start_col).value = f"Tổng tháng {label}" if label else "Tổng tháng"
    final_salary_letter = get_column_letter(PAYROLL_END_COL)
    salary_rows = ",".join(f"{final_salary_letter}{block.result_row}" for block in ordered)
    ws.cell(total_row, PAYROLL_END_COL).value = f"=SUM({salary_rows})"
    ws.cell(total_row, PAYROLL_END_COL).number_format = MONEY_FORMAT
    ws.row_dimensions[total_row].height = max(float(ws.row_dimensions[total_row].height or 0), 20)


def _select_attendance_sheet(wb):
    best_sheet = None
    best_count = -1
    for ws in wb.worksheets:
        count = sum(1 for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == "Att. Time")
        if count > best_count:
            best_sheet = ws
            best_count = count

    if best_sheet is None or best_count <= 0:
        raise ValueError("Không tìm thấy sheet chấm công có dòng Att. Time")

    return best_sheet


def _build_employee_preview(
    ws,
    block,
    include_saved_data: bool = True,
    profile_codes: set[str] | None = None,
    factory: str = "factory1",
    source_values: dict | None = None,
) -> dict:
    entry = get_payroll_entry(block.employee_code, factory)
    if profile_codes is not None and block.employee_code not in profile_codes:
        entry = PayrollEntry()
    source_values = source_values or {}
    total_hours = _prefer_source_number(source_values, "total_hours", _sum_work_hours(ws, block.result_row)) or 0
    monthly_salary = _prefer_source_number(source_values, "monthly_salary", calculate_monthly_salary(entry) if include_saved_data else None)
    daily_salary = _prefer_source_number(source_values, "daily_salary", calculate_daily_salary(entry) if include_saved_data else 0)
    hourly_salary = _prefer_source_number(source_values, "hourly_salary", calculate_hourly_salary(entry) if include_saved_data else 0)
    work_days = _prefer_source_number(source_values, "work_days", total_hours / 8)
    overtime_hours = _prefer_source_number(source_values, "overtime_hours", _sum_work_hours(ws, block.result_row + 1)) or 0
    bonus = _prefer_source_number(source_values, "bonus", entry.bonus if include_saved_data else None)
    # A frame conversion keeps period-specific values from that exact source
    # workbook. Normal monthly exports still leave them empty because no
    # source_payroll_values are supplied.
    advance_or_penalty = _prefer_source_number(source_values, "advance_or_penalty", None)
    nq_penalty = _prefer_source_number(source_values, "nq_penalty", 0)
    month_salary = (
        (daily_salary or 0) * (work_days or 0)
        + overtime_hours * (hourly_salary or 0) * 1.5
        + (bonus or 0)
        - nq_penalty
        - (advance_or_penalty or 0)
    )

    source_final_salary = _prefer_source_number(source_values, "final_salary", None)

    def source_number_or_rounded(key: str, value: int | float | None) -> int | float | None:
        if key in source_values and value is not None:
            return value
        return _round_optional_number(value)

    bank_account = (
        str(source_values.get("bank_account") or "").strip()
        if "bank_account" in source_values
        else _saved_bank_account(factory, block.employee_code) if include_saved_data else ""
    )

    return {
        "employee_code": block.employee_code,
        "name": normalize_employee_name(_prefer_source_text(source_values, "name", entry.name if include_saved_data else "")),
        "bank_account": bank_account,
        "start_work_note": _prefer_source_text(source_values, "start_work_note", entry.start_work_note if include_saved_data else ""),
        "note": _prefer_source_text(source_values, "note", entry.note if include_saved_data else ""),
        "header_row": block.header_row,
        "result_row": block.result_row,
        "note_row": block.header_row + 7,
        "total_hours": source_number_or_rounded("total_hours", total_hours),
        "monthly_salary": source_number_or_rounded("monthly_salary", monthly_salary),
        "daily_salary_input": entry.daily_salary if include_saved_data else None,
        "daily_salary": source_number_or_rounded("daily_salary", daily_salary),
        "hourly_salary": source_number_or_rounded("hourly_salary", hourly_salary),
        "standard_work_days": entry.standard_work_days if include_saved_data else 26,
        "work_days": source_number_or_rounded("work_days", work_days),
        "overtime_hours": source_number_or_rounded("overtime_hours", overtime_hours),
        "bonus": bonus,
        "penalty_rate": _prefer_source_number(source_values, "penalty_rate", None),
        "nq_penalty": source_number_or_rounded("nq_penalty", nq_penalty),
        "advance_or_penalty": advance_or_penalty,
        "final_salary": (
            source_final_salary
            if source_final_salary is not None
            else _round_number(month_salary)
        ),
    }


def _write_payroll_block(ws, block, preview: dict, source_values: dict | None = None) -> None:
    h = block.header_row
    payroll_header_row = h + 1
    result_row = block.result_row
    note_row = h + 7
    month_label = _month_label(ws.cell(row=h, column=3).value)
    total_col = 32
    penalty_rate_col = 33
    code_col = 34
    name_col = 35
    salary_col = 36
    daily_salary_col = 37
    hourly_salary_col = 38
    work_days_col = 39
    overtime_col = 40
    bonus_col = 41
    nq_penalty_col = 42
    advance_col = 43
    final_salary_col = 44

    _unmerge_overlapping(ws, h, PAYROLL_START_COL, note_row, PAYROLL_END_COL)
    _copy_boundary_style(ws, h, note_row)
    ws.row_dimensions[payroll_header_row].height = max(float(ws.row_dimensions[payroll_header_row].height or 0), 28)

    headers = {
        total_col: "Tổng giờ công",
        penalty_rate_col: "Mức tiền phạt NQ trên giờ công (đ)",
        code_col: "Mã",
        name_col: "Tên nhân viên / Ghi chú",
        salary_col: "Mức Lương",
        daily_salary_col: "Lương 1 Ngày Công",
        hourly_salary_col: "Lương 1 Giờ Công",
        work_days_col: "Số Ngày Đi Làm",
        overtime_col: "Giờ làm thêm",
        bonus_col: "Thưởng",
        nq_penalty_col: "Phạt NQ",
        advance_col: "Ứng Lương",
        final_salary_col: f"Lương Tháng {month_label}",
    }

    for col in range(PAYROLL_START_COL, PAYROLL_END_COL + 1):
        for row in range(h, note_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if row == h or col <= penalty_rate_col:
                fill_color = WHITE
            elif col <= name_col:
                fill_color = YELLOW
            else:
                fill_color = YELLOW if row == payroll_header_row else GREEN
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(name=FONT_NAME, size=9)

    for col in range(PAYROLL_START_COL, PAYROLL_END_COL + 1):
        cell = ws.cell(row=h, column=col)
        cell.value = None
        cell.fill = PatternFill("solid", fgColor=WHITE)

    for col, title in headers.items():
        cell = ws.cell(row=payroll_header_row, column=col)
        cell.value = title
        if col <= penalty_rate_col:
            cell.fill = PatternFill("solid", fgColor=WHITE)
        else:
            cell.fill = PatternFill("solid", fgColor=YELLOW)
        cell.font = Font(
            name=FONT_NAME,
            bold=True,
            size=8,
            color=RED if col in {total_col, penalty_rate_col, code_col} else "000000",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)

    first_day = get_column_letter(1)
    last_day = get_column_letter(total_col - 1)
    total_letter = get_column_letter(total_col)
    penalty_rate_letter = get_column_letter(penalty_rate_col)
    salary_letter = get_column_letter(salary_col)
    daily_salary_letter = get_column_letter(daily_salary_col)
    hourly_salary_letter = get_column_letter(hourly_salary_col)
    work_days_letter = get_column_letter(work_days_col)
    overtime_letter = get_column_letter(overtime_col)
    bonus_letter = get_column_letter(bonus_col)
    nq_penalty_letter = get_column_letter(nq_penalty_col)
    advance_letter = get_column_letter(advance_col)

    source_values = source_values or {}

    def has_source_value(key: str) -> bool:
        return key in source_values and source_values[key] not in (None, "")

    values = {
        # A legacy conversion is a historical snapshot. Preserve the exact
        # total from that workbook when it exists; normal monthly exports keep
        # the live formula driven by the attendance row.
        total_col: (
            preview["total_hours"]
            if has_source_value("total_hours")
            else f"=SUM({first_day}{result_row}:{last_day}{result_row})"
        ),
        penalty_rate_col: preview.get("penalty_rate"),
        code_col: block.employee_code,
        name_col: preview["name"],
        salary_col: preview["monthly_salary"],
        daily_salary_col: (
            # The salary column is the input for the new frame.  Even when a
            # legacy workbook supplied an old daily-rate number, keep this
            # cell formula-driven so the converted frame follows the same
            # rules as Xưởng 1 when the salary is edited later.
            f'=IF({salary_letter}{result_row}>0,{salary_letter}{result_row}/26,"")'
        ),
        hourly_salary_col: (
            f'=IF({salary_letter}{result_row}>0,{salary_letter}{result_row}/208,"")'
        ),
        work_days_col: (
            # Work days are derived from the attendance row, not copied as a
            # stale number from the old vertical summary.
            f"=SUM({first_day}{result_row}:{last_day}{result_row})/8"
        ),
        overtime_col: (
            preview["overtime_hours"]
            if has_source_value("overtime_hours")
            else f"=SUM({first_day}{result_row + 1}:{last_day}{result_row + 1})"
        ),
        bonus_col: preview["bonus"],
        nq_penalty_col: (
            preview["nq_penalty"]
            if has_source_value("nq_penalty")
            else f'=IF(ISNUMBER({penalty_rate_letter}{result_row}),'
                 f"{penalty_rate_letter}{result_row}*{total_letter}{result_row},0)"
        ),
        advance_col: preview["advance_or_penalty"],
        final_salary_col: (
            preview["final_salary"]
            if has_source_value("final_salary")
            else (
                f'=IF({daily_salary_letter}{result_row}>0,{daily_salary_letter}{result_row},'
                f'IF({salary_letter}{result_row}>0,{salary_letter}{result_row}/26,0))'
                f"*{work_days_letter}{result_row}"
                f"+({overtime_letter}{result_row}*{hourly_salary_letter}{result_row}*1.5)"
                f"-IF(ISNUMBER({advance_letter}{result_row}),{advance_letter}{result_row},0)"
                f"-{nq_penalty_letter}{result_row}"
                f"+IF(ISNUMBER({bonus_letter}{result_row}),{bonus_letter}{result_row},0)"
            )
        ),
    }

    for col, value in values.items():
        cell = ws.cell(row=result_row, column=col)
        cell.value = value
        cell.font = Font(
            name=FONT_NAME,
            bold=col in {total_col, penalty_rate_col, code_col, name_col, final_salary_col},
            size=10 if col in {name_col, final_salary_col} else 9,
            color=RED if col in {total_col, penalty_rate_col, code_col} else "000000",
        )
        if col in {
            salary_col,
            daily_salary_col,
            hourly_salary_col,
            bonus_col,
            nq_penalty_col,
            advance_col,
            final_salary_col,
        }:
            cell.number_format = MONEY_FORMAT
        elif col in {total_col, penalty_rate_col, work_days_col, overtime_col}:
            cell.number_format = FLEXIBLE_NUMBER_FORMAT

    bank_account_cell = ws.cell(row=result_row - 1, column=name_col)
    bank_account_cell.value = str(preview.get("bank_account") or "")
    bank_account_cell.number_format = "@"
    bank_account_cell.alignment = Alignment(horizontal="left", vertical="center")

    # The legacy layout reserves AI for Bắt đầu làm and AJ:AR for a separate
    # comment. Do not merge them: a long comment must never hide the start
    # work information.
    for col in range(total_col, name_col):
        ws.cell(row=note_row, column=col).value = None
    for col in range(name_col, final_salary_col + 1):
        note_part = ws.cell(row=note_row, column=col)
        note_part.value = None
        note_part.fill = PatternFill("solid", fgColor=NOTE_FILL)
    start_work_cell = ws.cell(row=note_row, column=name_col)
    start_work_cell.value = _format_start_work_note(preview.get("start_work_note"))
    start_work_cell.fill = PatternFill("solid", fgColor=START_WORK_FILL)
    start_work_cell.font = Font(name=FONT_NAME, size=9, bold=True, color="000000")
    start_work_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _safe_merge(ws, note_row, name_col + 1, note_row, final_salary_col)
    note_cell = ws.cell(row=note_row, column=name_col + 1)
    note_cell.value = str(preview.get("note") or "").strip()
    note_cell.fill = PatternFill("solid", fgColor=NOTE_FILL)
    note_cell.font = Font(name=FONT_NAME, size=9, color=RED)
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _set_note_row_height(ws, note_row, start_work_cell.value, note_cell.value, name_col, final_salary_col)


def _set_note_row_height(ws, row: int, start_value: object, note_value: object, start_col: int, end_col: int) -> None:
    """Excel does not auto-fit merged cells, so estimate both note areas."""
    start_width = float(ws.column_dimensions[get_column_letter(start_col)].width or 10)
    note_width = sum(
        float(ws.column_dimensions[get_column_letter(col)].width or 10)
        for col in range(start_col + 1, end_col + 1)
    )
    line_count = max(_wrapped_line_count(start_value, start_width), _wrapped_line_count(note_value, note_width))
    required_height = max(20, 15 * line_count + 5)
    ws.row_dimensions[row].height = max(float(ws.row_dimensions[row].height or 0), required_height)


def _wrapped_line_count(value: object, width: float) -> int:
    chars_per_line = max(12, int(width * 0.72))
    return max((ceil(len(line) / chars_per_line) or 1 for line in (str(value or "").splitlines() or [""])), default=1)


def _sum_work_hours(ws, row: int) -> float:
    total = 0.0
    for col in range(1, 32):
        value = _restore_number(ws.cell(row=row, column=col).value)
        if isinstance(value, (int, float)):
            total += float(value)
    return total


def _restore_number(value: object) -> int | float | None:
    number = _parse_number(value)
    return _round_number(number) if number is not None else None


def _parse_number(value: object) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value

    text = str(value or "").strip()
    if not text or text == "?":
        return None

    normalized = text.replace(" ", "")
    if normalized.count(",") > 1 and "." not in normalized:
        normalized = normalized.replace(",", "")
    elif normalized.count(".") > 1 and "," not in normalized:
        normalized = normalized.replace(".", "")
    elif "," in normalized and "." in normalized:
        # Vietnamese files commonly use dots for thousands and comma for the
        # decimal separator. Keep the last separator as the decimal marker.
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    else:
        normalized = normalized.replace(",", ".")
    try:
        number = float(normalized)
    except ValueError:
        return None
    return number


def _prefer_source_number(source: dict, key: str, fallback: int | float | None) -> int | float | None:
    if key not in source:
        return fallback
    value = _parse_number(source.get(key))
    return fallback if value is None else value


def _prefer_source_text(source: dict, key: str, fallback: str) -> str:
    if key in source:
        return str(source.get(key) or "").strip()
    return fallback


def _round_optional_number(value: float | None) -> int | float | None:
    if value is None:
        return None
    return _round_number(value)


def _round_number(value: float) -> int | float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _format_start_work_note(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.lower()
    if normalized.startswith("bắt đầu") or normalized.startswith("bat dau"):
        return text
    return f"Bắt đầu làm {text}"


def _saved_bank_account(factory: str, employee_code: object) -> str:
    # Import lazily because the bank module also owns optional Drive helpers.
    from app.services.bank_account_store import get_saved_account_number

    return get_saved_account_number(factory, employee_code)


def _month_label(value: object) -> str:
    text = str(value or "")
    if len(text) >= 7 and text[4] == "-":
        year = text[0:4]
        month = str(int(text[5:7]))
        return f"{month}/{year}"
    return ""


def _copy_boundary_style(ws, start_row: int, end_row: int) -> None:
    source_col = 31
    for row in range(start_row, end_row + 1):
        source = ws.cell(row=row, column=source_col)
        for col in range(PAYROLL_START_COL, PAYROLL_END_COL + 1):
            target = ws.cell(row=row, column=col)
            target._style = copy(source._style)
            if source.has_style:
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.number_format = source.number_format
                target.protection = copy(source.protection)


def _safe_merge(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    target = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == target:
            return
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def _unmerge_overlapping(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if not (
            merged_range.max_row < start_row
            or merged_range.min_row > end_row
            or merged_range.max_col < start_col
            or merged_range.min_col > end_col
        ):
            ws.unmerge_cells(str(merged_range))


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_payroll_column_widths(ws) -> None:
    widths = {
        "AF": 16,
        "AG": 38,
        "AH": 8,
        "AI": 28,
        "AJ": 14,
        "AK": 16,
        "AL": 16,
        "AM": 14,
        "AN": 11,
        "AO": 12,
        "AP": 12,
        "AQ": 14,
        "AR": 18,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
