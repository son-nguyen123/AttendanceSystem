from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.payroll_store import normalize_employee_code, normalize_employee_name


SUMMARY_MIN_COL = 32


def read_final_copy_overview(path: Path, month: int) -> list[dict[str, Any]]:
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, data_only=True, keep_vba=keep_vba)
    try:
        rows: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows.extend(_read_sheet(worksheet, month))
        return rows
    finally:
        workbook.close()


def _read_sheet(ws, month: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for header_row in range(1, ws.max_row + 1):
        total_col = _find_label_col(ws, header_row, SUMMARY_MIN_COL, ws.max_column, {"tong gio cong"})
        if not total_col:
            continue
        code_col = _find_label_col(ws, header_row, total_col + 1, min(total_col + 4, ws.max_column), {"ma"})
        if not code_col:
            continue

        result_row = header_row + 5
        code = _employee_code(ws.cell(result_row, code_col).value)
        if not code:
            continue

        name_col = code_col + 1
        work_days_col = _find_work_days_col(ws, header_row, name_col + 1)
        total_hours = _number(ws.cell(result_row, total_col).value)
        work_days = _number(ws.cell(result_row, work_days_col).value) if work_days_col else None
        if work_days is None and total_hours is not None:
            work_days = total_hours / 8
        if total_hours is None and work_days is not None:
            total_hours = work_days * 8

        rows.append(
            {
                "employee_code": code,
                "employee_name": _text(ws.cell(result_row, name_col).value),
                "month": month,
                "total_hours": _round_number(total_hours or 0),
                "work_days": _round_number(work_days or 0),
                "late_count": 0,
                "issue_count": 0,
                "source": "final_copy",
            }
        )
    return rows


def _find_work_days_col(ws, row: int, start_col: int) -> int | None:
    for col in range(start_col, ws.max_column + 1):
        if _normalize_label(ws.cell(row, col).value) == "so ngay di lam":
            return col
    return None


def _find_label_col(ws, row: int, start_col: int, end_col: int, labels: set[str]) -> int | None:
    for col in range(start_col, end_col + 1):
        if _normalize_label(ws.cell(row, col).value) in labels:
            return col
    return None


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return " ".join(text.split())


def _employee_code(value: object) -> str:
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return normalize_employee_code(value)


def _text(value: object) -> str:
    return normalize_employee_name(value)


def _number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip()
    if not text:
        return None
    cleaned = text.replace(",", "").replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _round_number(value: float) -> int | float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded
