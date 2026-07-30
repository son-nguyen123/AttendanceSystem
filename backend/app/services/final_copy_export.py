from __future__ import annotations

import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill


SUMMARY_MIN_COL = 32
SUMMARY_ROWS = 7
TOTAL_LABELS = {"tong gio cong"}
CODE_LABELS = {"ma"}
PRIVATE_TAIL_LABELS = {
    "muc luong",
    "luong 1 ngay cong",
    "luong 1 gio cong",
    "so ngay di lam",
    "thuong",
    "ung luong + phat",
}


def export_final_copy_output1(source_path: Path, output_path: Path) -> Path:
    keep_vba = source_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(source_path, data_only=False, keep_vba=keep_vba)
    try:
        for worksheet in workbook.worksheets:
            _strip_private_tail(worksheet)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return output_path
    finally:
        workbook.close()


def _strip_private_tail(ws) -> None:
    for header_row in range(1, ws.max_row + 1):
        total_col = _find_label_col(ws, header_row, SUMMARY_MIN_COL, ws.max_column, TOTAL_LABELS)
        if not total_col:
            continue
        code_col = _find_label_col(ws, header_row, total_col + 1, min(total_col + 4, ws.max_column), CODE_LABELS)
        if not code_col:
            continue

        name_col = code_col + 1
        tail_start = _find_private_tail_start(ws, header_row, name_col)
        if not tail_start:
            tail_start = _merged_range_end_col(ws, header_row, name_col) + 1
        if tail_start <= name_col or tail_start > ws.max_column:
            continue
        _clear_area(ws, header_row, tail_start, header_row + SUMMARY_ROWS - 1, ws.max_column)


def _find_private_tail_start(ws, row: int, start_col: int) -> int | None:
    for col in range(start_col, ws.max_column + 1):
        label = _normalize_label(ws.cell(row, col).value)
        if label in PRIVATE_TAIL_LABELS or label.startswith("luong thang"):
            return col
    return None


def _find_label_col(ws, row: int, start_col: int, end_col: int, labels: set[str]) -> int | None:
    for col in range(start_col, end_col + 1):
        if _normalize_label(ws.cell(row, col).value) in labels:
            return col
    return None


def _merged_range_end_col(ws, row: int, col: int) -> int:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.max_col
    return col


def _clear_area(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    _unmerge_overlapping(ws, min_row, min_col, max_row, max_col)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.hyperlink = None
            cell.comment = None


def _unmerge_overlapping(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if not (
            merged.max_row < min_row
            or merged.min_row > max_row
            or merged.max_col < min_col
            or merged.min_col > max_col
        ):
            ws.unmerge_cells(str(merged))


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return " ".join(text.split())
