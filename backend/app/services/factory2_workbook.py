from collections import OrderedDict
from dataclasses import dataclass, field
from datetime import date, datetime
from calendar import monthrange
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.attendance import DayComputation
from app.services.attendance_calculator import calculate_day
from app.services.payroll_workbook import export_payroll_workbook
from app.services.payroll_store import PayrollEntry, calculate_daily_salary, calculate_hourly_salary, calculate_monthly_salary, get_payroll_entry, normalize_employee_name
from app.services.punch_parser import parse_punches
from app.services.workbook_processor import export_processed_workbook


YELLOW = "FFFF00"
GRAY = "D9D9D9"
RED = "C00000"
WHITE = "FFFFFF"
FONT_NAME = "Arial"
MONEY_FORMAT = "#,##0"
INTEGER_NUMBER_FORMAT = "#,##0"
NUMBER_FORMAT = "#,##0.##"
MAX_FACTORY2_COLUMNS = 200
MAX_TRAILING_EMPTY_ROWS = 200


@dataclass
class Factory2Day:
    day: date
    punches: list[str]
    computation: DayComputation | None = None


@dataclass
class Factory2Employee:
    employee_code: str
    source_name: str = ""
    days: list[Factory2Day] = field(default_factory=list)


def analyze_factory2_workbook(path: Path) -> dict:
    employees, sheet_name, period, employee_counts = _read_active_employees(path)
    rows: list[dict] = []
    manual_checks: list[dict] = []
    result_cells = 0
    missing_cells = 0
    late_cells = 0

    for employee in employees:
        block_results = []
        for day in employee.days:
            if day.computation is None:
                block_results.append(_empty_day_result(day.day))
                continue

            item = day.computation
            if item.work_value is not None:
                result_cells += 1
            if item.missing_count is not None:
                missing_cells += 1
            if item.late_minutes is not None:
                late_cells += 1
            if item.manual_checks:
                manual_checks.append(
                    {
                        "employee_code": employee.employee_code,
                        "day": item.day,
                        "cell": "",
                        "punches": item.punches,
                        "messages": item.manual_checks,
                    }
                )
            block_results.append(
                {
                    "day": item.day,
                    "column": str(item.day),
                    "punches": item.punches,
                    "work_value": item.work_value,
                    "missing_count": item.missing_count,
                    "late_minutes": item.late_minutes,
                }
            )

        rows.append(
            {
                "employee_code": employee.employee_code,
                "header_row": 0,
                "punch_row": 0,
                "missing_row": 0,
                "late_row": 0,
                "result_row": 0,
                "results": block_results,
            }
        )

    return {
        "sheet_name": sheet_name,
        "period": period,
        "summary": {
            "blocks": len(employees),
            "source_employee_count": employee_counts["source_employee_count"],
            "empty_employee_count": employee_counts["empty_employee_count"],
            "result_cells": result_cells,
            "missing_cells": missing_cells,
            "late_cells": late_cells,
            "manual_check_count": len(manual_checks),
        },
        "blocks": rows,
        "manual_checks": manual_checks,
    }


def preview_factory2_payroll(
    source_path: Path,
    review_overrides: list[dict] | None = None,
    profile_codes: set[str] | None = None,
    factory: str = "factory2",
) -> dict:
    employees, sheet_name, _period, _employee_counts = _read_active_employees(
        source_path,
        review_overrides=review_overrides,
    )
    return {
        "sheet_name": sheet_name,
        "employees": [_build_employee_preview(employee, profile_codes=profile_codes, factory=factory) for employee in employees],
    }


def export_factory2_output1(
    source_path: Path,
    output_path: Path,
    review_overrides: list[dict] | None = None,
    factory: str = "factory2",
) -> Path:
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_source = Path(temp_file.name)

    try:
        _write_factory1_shaped_source(source_path, temp_source)
        return export_processed_workbook(temp_source, output_path, review_overrides=review_overrides, factory=factory)
    finally:
        temp_source.unlink(missing_ok=True)


def export_factory2_output2(
    source_path: Path,
    output_path: Path,
    review_overrides: list[dict] | None = None,
    include_saved_data: bool = True,
    profile_codes: set[str] | None = None,
    factory: str = "factory2",
    carry_source_payroll_data: bool = False,
) -> Path:
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
        temp_source = Path(temp_file.name)

    try:
        _write_factory1_shaped_source(source_path, temp_source)
        return export_payroll_workbook(
            temp_source,
            output_path,
            review_overrides=review_overrides,
            include_saved_data=include_saved_data,
            profile_codes=profile_codes,
            factory=factory,
            source_payroll_values=_read_legacy_payroll_values(source_path) if carry_source_payroll_data else None,
        )
    finally:
        temp_source.unlink(missing_ok=True)


def write_factory2_standard_source(source_path: Path, output_path: Path) -> Path:
    return _write_factory1_shaped_source(source_path, output_path)


def _read_legacy_payroll_values(
    source_path: Path,
    *,
    formula_workbook=None,
    values_workbook=None,
) -> dict[str, dict[str, Any]]:
    """Read same-period owner values from any recognizable table in the source.

    Raw vertical attendance files normally have no payroll fields. Older
    owner-edited forms may contain a separate or repeated summary table; those
    values must survive a frame conversion because the converted Output 2 is
    intended to become the final copy for the same month.
    """
    owns_formula_wb = formula_workbook is None
    owns_values_wb = values_workbook is None
    formula_wb = formula_workbook or load_workbook(source_path, data_only=False)
    values_wb = values_workbook or load_workbook(source_path, data_only=True)
    try:
        result: dict[str, dict[str, Any]] = {}
        for source_ws in formula_wb.worksheets:
            values_ws = values_wb[source_ws.title]
            for header_row, fields in _legacy_payroll_sections(source_ws):
                extras = _legacy_adjacent_summary_values(values_ws, header_row)
                data_row = _legacy_payroll_data_row(values_ws, header_row, fields)
                if data_row is None:
                    continue

                code = _clean_code(values_ws.cell(data_row, fields["code"]).value)
                if not code:
                    continue

                item = result.setdefault(code, {})
                item.update({key: value for key, value in extras.items() if value not in (None, "")})
                for key, col in fields.items():
                    if key == "code":
                        continue
                    value = values_ws.cell(data_row, col).value
                    if value in (None, ""):
                        value = source_ws.cell(data_row, col).value
                    if value not in (None, ""):
                        item[key] = value
        return result
    finally:
        if owns_formula_wb:
            formula_wb.close()
        if owns_values_wb:
            values_wb.close()


def _legacy_payroll_header_fields(ws, row: int) -> dict[str, int]:
    fields: dict[str, int] = {}
    for col in range(1, min(ws.max_column, MAX_FACTORY2_COLUMNS) + 1):
        label = _plain_text(ws.cell(row, col).value)
        if not label:
            continue
        if label in {"ma", "ma nv", "ma nhan vien"}:
            # A legacy employee block may also have a standalone "Mã:" label
            # near column A.  The payroll summary's code is the one after
            # "Tổng giờ công", so prefer that later semantic column.
            if (
                "code" not in fields
                or (
                    "total_hours" in fields
                    and col > fields["total_hours"]
                    and fields["code"] < fields["total_hours"]
                )
            ):
                fields["code"] = col
        elif label in {"ten", "ten nv", "ten nhan vien", "ho va ten", "ten nhan vien ghi chu"} and "name" not in fields:
            fields["name"] = col
        elif label.startswith("muc tien phat nq") and "penalty_rate" not in fields:
            fields["penalty_rate"] = col
        else:
            payroll_field = _legacy_payroll_field(label)
            if payroll_field and payroll_field not in fields:
                fields[payroll_field] = col
    return fields


def _legacy_payroll_field(label: str) -> str | None:
    """Map labels from the old per-employee summary blocks to Output 2 fields."""
    if label in {"muc luong", "luong thang muc luong"}:
        return "monthly_salary"
    if label.startswith("luong 1 ngay"):
        return "daily_salary"
    if label.startswith("luong 1 gio"):
        return "hourly_salary"
    if label.startswith("so ngay di lam") or label.startswith("so ngay i lam") or label == "ngay cong":
        return "work_days"
    if label.startswith("tong h lam") or label.startswith("tong gio lam") or label.startswith("tong gio cong"):
        return "total_hours"
    if label.startswith("h tang") or label.startswith("so h tang ca") or label.startswith("gio tang ca") or label.startswith("gio lam them"):
        return "overtime_hours"
    if label.startswith("thuong"):
        return "bonus"
    if label in {"ung luong + phat", "ung luong va phat"}:
        # Legacy Factory 1 stores these two deductions in one input column.
        # Keep the complete amount in the new advance input so the existing
        # payroll formula produces the same result without changing formulas.
        return "advance_or_penalty"
    if label.startswith("phat nq"):
        return "nq_penalty"
    if label.startswith("ung luong") or label.startswith("ung luong phat"):
        return "advance_or_penalty"
    if label.startswith("luong thang"):
        return "final_salary"
    if label == "so tai khoan":
        return "bank_account"
    if label.startswith("bat dau lam"):
        return "start_work_note"
    if label.startswith("ghi chu"):
        return "note"
    return None


def _legacy_payroll_sections(ws) -> list[tuple[int, dict[str, int]]]:
    """Find both normal summary tables and the old Xưởng 2 block format.

    The old form places the employee code in the first cell of the summary
    header row instead of writing a literal ``Mã`` label.  Its real data is on
    the next row, so a label-only scanner misses the entire payroll section.
    """
    sections: list[tuple[int, dict[str, int]]] = []
    payroll_keys = {
        "monthly_salary", "daily_salary", "hourly_salary", "work_days",
        "overtime_hours", "bonus", "nq_penalty", "advance_or_penalty", "final_salary",
    }
    for row in range(1, min(ws.max_row, 100_000) + 1):
        fields = _legacy_payroll_header_fields(ws, row)
        if "code" not in fields:
            first = ws.cell(row, 1).value
            second = _plain_text(ws.cell(row, 2).value)
            if _clean_code(first) and second in {"ten", "ten nv", "ten nhan vien", "ho va ten", "ho ten"}:
                fields["code"] = 1
                fields.setdefault("name", 2)
        if "code" not in fields or not payroll_keys.intersection(fields):
            continue
        # A raw attendance header may contain Mã NV but never has payroll
        # fields; requiring at least two payroll labels filters it out.
        if len(payroll_keys.intersection(fields)) < 2:
            continue
        sections.append((row, fields))
    return sections


def _legacy_payroll_data_row(ws, header_row: int, fields: dict[str, int]) -> int | None:
    code_col = fields["code"]
    for row in range(header_row + 1, min(ws.max_row, header_row + 6) + 1):
        code = _clean_code(ws.cell(row, code_col).value)
        if not code:
            continue
        values = [ws.cell(row, col).value for key, col in fields.items() if key != "code"]
        if any(value not in (None, "") for value in values):
            return row
    return None


def _legacy_adjacent_summary_values(ws, header_row: int) -> dict[str, Any]:
    """Read values stored on the metadata row above an old summary header.

    Older Xưởng 2 sheets put ``Bắt đầu làm ...`` and the free-form note on the
    same row as ``Tổng h làm`` instead of giving them dedicated header cells.
    Keep those values when converting the vertical layout to Output 2.
    """
    result: dict[str, Any] = {}
    start_row = max(1, header_row - 3)
    for row in range(start_row, header_row):
        for col in range(1, min(ws.max_column, MAX_FACTORY2_COLUMNS) + 1):
            raw_value = ws.cell(row, col).value
            label = _plain_text(raw_value)
            if not label:
                continue
            field = _legacy_payroll_field(label)
            if field == "start_work_note":
                # A value such as "Bắt đầu làm T3/2023" is itself the data,
                # while a bare header "Bắt đầu làm" takes the next cell.
                if label not in {"bat dau lam", "bat dau vao lam"}:
                    result.setdefault("start_work_note", str(raw_value).strip())
                    continue
            if field not in {"total_hours", "work_days", "overtime_hours", "start_work_note", "note"}:
                continue
            for value_col in range(col + 1, min(ws.max_column, col + 4) + 1):
                value = ws.cell(row, value_col).value
                if value not in (None, ""):
                    result.setdefault(field, value)
                    break

        # In the common legacy form, the note is the free text immediately
        # after the numeric total-hours value (there is no "Ghi chú" label).
        for col in range(1, min(ws.max_column, MAX_FACTORY2_COLUMNS) + 1):
            if _legacy_payroll_field(_plain_text(ws.cell(row, col).value)) != "total_hours":
                continue
            numeric_col = next(
                (
                    candidate
                    for candidate in range(col + 1, min(ws.max_column, col + 4) + 1)
                    if isinstance(ws.cell(row, candidate).value, (int, float))
                    and not isinstance(ws.cell(row, candidate).value, bool)
                ),
                None,
            )
            if numeric_col is None:
                continue
            for note_col in range(numeric_col + 1, min(ws.max_column, numeric_col + 4) + 1):
                candidate = ws.cell(row, note_col).value
                if isinstance(candidate, str) and candidate.strip():
                    result.setdefault("note", candidate.strip())
                    break
    return result


def _export_factory2_output2_vertical(
    source_path: Path,
    output_path: Path,
    review_overrides: list[dict] | None = None,
) -> Path:
    employees, sheet_name, period, _employee_counts = _read_active_employees(
        source_path,
        review_overrides=review_overrides,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(sheet_name)
    ws.sheet_view.showGridLines = True

    _setup_output_sheet(ws, period)
    row = 3
    for employee in employees:
        row = _write_employee_block(ws, row, employee, period)
        row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path


def _write_factory1_shaped_source(source_path: Path, output_path: Path) -> Path:
    employees, sheet_name, period, _employee_counts = _read_active_employees(source_path)
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(sheet_name)
    ws.sheet_view.showGridLines = True

    _write_factory1_shaped_title(ws, period)
    start_row = 3
    for index, employee in enumerate(employees):
        block_row = start_row + index * 8
        _write_factory1_shaped_block(ws, block_row, employee, period)

    _set_factory1_shaped_widths(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path


def _write_factory1_shaped_title(ws, period: dict[str, int | str | None]) -> None:
    _safe_merge(ws, 1, 1, 2, 31)
    title = f"Bang cong Xuong 2 {period.get('label') or ''}".strip()
    cell = ws.cell(row=1, column=1)
    cell.value = title
    cell.font = Font(name=FONT_NAME, bold=True, size=20)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28


def _write_factory1_shaped_block(ws, header_row: int, employee: Factory2Employee, period: dict[str, int | str | None]) -> None:
    day_row = header_row + 1
    employee_row = header_row + 2
    punch_row = header_row + 3
    missing_row = header_row + 4
    late_row = header_row + 5
    result_row = header_row + 6
    note_row = header_row + 7
    days_by_number = {item.day.day: item for item in employee.days}
    month = period.get("month")
    year = period.get("year")
    period_start = f"{year}-{int(month):02d}-01" if isinstance(month, int) and isinstance(year, int) else ""
    period_end = (
        f"{year}-{int(month):02d}-{monthrange(year, month)[1]:02d}"
        if isinstance(month, int) and isinstance(year, int)
        else ""
    )
    period_range = f"{period_start} ~ {period_end}" if period_start and period_end else period_start

    ws.cell(row=header_row, column=1).value = "Att. Time"
    ws.cell(row=header_row, column=3).value = period_range
    ws.cell(row=header_row, column=10).value = f"Tabulation {date.today().isoformat()}"
    _safe_merge(ws, header_row, 1, header_row, 2)
    _safe_merge(ws, header_row, 3, header_row, 8)
    _safe_merge(ws, header_row, 10, header_row, 15)
    _safe_merge(ws, header_row, 16, header_row, 31)

    for col in range(1, 32):
        day_cell = ws.cell(row=day_row, column=col)
        day_cell.value = col
        day_cell.font = Font(name=FONT_NAME, bold=True, size=9)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=employee_row, column=1).value = "Mã:"
    ws.cell(row=employee_row, column=3).value = employee.employee_code
    ws.cell(row=employee_row, column=9).value = f"Tên: {employee.source_name if employee.source_name != employee.employee_code else ''}".strip()
    ws.cell(row=employee_row, column=19).value = "Phòng Ban: Công ty"
    _safe_merge(ws, employee_row, 1, employee_row, 2)
    _safe_merge(ws, employee_row, 3, employee_row, 8)
    _safe_merge(ws, employee_row, 9, employee_row, 15)
    _safe_merge(ws, employee_row, 19, employee_row, 22)

    for col in range(1, 32):
        day = days_by_number.get(col)
        computation = day.computation if day else None
        ws.cell(row=punch_row, column=col).value = "\n".join(day.punches) if day and day.punches else None
        ws.cell(row=missing_row, column=col).value = computation.missing_count if computation else None
        ws.cell(row=late_row, column=col).value = computation.late_minutes if computation else None
        ws.cell(row=result_row, column=col).value = computation.work_value if computation else None

    row_heights = {
        header_row: 14.5,
        day_row: 28,
        employee_row: 18,
        punch_row: 100,
        missing_row: 14.5,
        late_row: 14.5,
        result_row: 14.5,
        note_row: 14.5,
    }
    for row in range(header_row, note_row + 1):
        ws.row_dimensions[row].height = row_heights.get(row, 14.5)
        for col in range(1, 32):
            cell = ws.cell(row=row, column=col)
            cell.border = _thin_border()
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
            cell.fill = PatternFill("solid", fgColor=WHITE)
            if row in {missing_row, late_row, result_row}:
                cell.number_format = _number_format_for(cell.value)

    for col in range(1, 32):
        if _is_sunday(period, col):
            for row in range(day_row, note_row + 1):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=YELLOW)

    for row in (header_row, employee_row):
        for col in range(1, 32):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _set_factory1_shaped_widths(ws) -> None:
    for col in range(1, 32):
        ws.column_dimensions[get_column_letter(col)].width = 6.2


def _is_sunday(period: dict[str, int | str | None], day: int) -> bool:
    month = period.get("month")
    year = period.get("year")
    if not isinstance(month, int) or not isinstance(year, int):
        return False
    try:
        return date(year, month, day).weekday() == 6
    except ValueError:
        return False


def _read_active_employees(
    path: Path,
    review_overrides: list[dict] | None = None,
) -> tuple[
    list[Factory2Employee],
    str,
    dict[str, int | str | None],
    dict[str, int],
]:
    wb = load_workbook(path, data_only=True)
    try:
        ws, header_row, columns = _select_factory2_sheet(wb)
        overrides = _review_overrides_by_employee_day(review_overrides or [])
        employees_by_code: OrderedDict[str, Factory2Employee] = OrderedDict()
        period_month = None
        period_year = None

        trailing_empty_rows = 0
        last_source_row = min(ws.max_row, 100_000)
        relevant_columns = [columns["code"], columns["date"], *columns["punches"]]
        if "name" in columns:
            relevant_columns.append(columns["name"])

        for row in range(header_row + 1, last_source_row + 1):
            code = _clean_code(ws.cell(row=row, column=columns["code"]).value)
            has_relevant_value = any(ws.cell(row=row, column=col).value not in (None, "") for col in relevant_columns)
            if not has_relevant_value:
                trailing_empty_rows += 1
                if trailing_empty_rows >= MAX_TRAILING_EMPTY_ROWS:
                    break
                continue
            trailing_empty_rows = 0
            if not code:
                continue

            day_value = ws.cell(row=row, column=columns["date"]).value
            current_date = _coerce_date(day_value)
            if current_date is None:
                continue

            period_month = period_month or current_date.month
            period_year = period_year or current_date.year
            employee = employees_by_code.setdefault(
                code,
                Factory2Employee(
                    employee_code=code,
                    source_name=_clean_text(ws.cell(row=row, column=columns.get("name", columns["code"])).value),
                ),
            )

            punches = _row_punches(ws, row, columns["punches"])
            computation = _compute_day(current_date, punches)
            if computation is not None:
                override = overrides.get((code, current_date.day), {})
                computation = _apply_override(computation, override)
            employee.days.append(Factory2Day(day=current_date, punches=punches, computation=computation))

        active_employees = [employee for employee in employees_by_code.values() if any(day.punches for day in employee.days)]
        period = {
            "month": period_month,
            "year": period_year,
            "label": f"{period_month:02d}/{period_year}" if period_month and period_year else "",
        }
        employee_counts = {
            "source_employee_count": len(employees_by_code),
            "empty_employee_count": len(employees_by_code) - len(active_employees),
        }
        return active_employees, ws.title, period, employee_counts
    finally:
        wb.close()


def _select_factory2_sheet(wb):
    for ws in wb.worksheets:
        for row in range(1, min(ws.max_row, 20) + 1):
            headers = [_plain_text(ws.cell(row=row, column=col).value) for col in range(1, min(ws.max_column, MAX_FACTORY2_COLUMNS) + 1)]
            if not any("ma" in value and "nv" in value for value in headers):
                continue
            if not any(value.startswith("ngay") or value.startswith("nga") for value in headers):
                continue

            columns = _factory2_columns(ws, row)
            if columns:
                return ws, row, columns
    raise ValueError("Khong tim thay sheet raw xuong 2 co cot Ma NV / Ngay / Lan")


def _factory2_columns(ws, header_row: int) -> dict | None:
    columns: dict[str, object] = {"punches": []}
    for col in range(1, min(ws.max_column, MAX_FACTORY2_COLUMNS) + 1):
        text = _plain_text(ws.cell(row=header_row, column=col).value)
        if not text:
            continue
        if "ma" in text and "nv" in text and "code" not in columns:
            columns["code"] = col
        elif ("ten" in text or text.startswith("te") or "name" in text) and "name" not in columns:
            columns["name"] = col
        elif (text.startswith("ngay") or text.startswith("nga") or text == "date") and "date" not in columns:
            columns["date"] = col
        elif text.startswith("lan") or text.startswith("la"):
            columns["punches"].append(col)

    if "code" not in columns or "date" not in columns or not columns["punches"]:
        return None
    return columns


def _row_punches(ws, row: int, punch_columns: list[int]) -> list[str]:
    punches: list[str] = []
    for col in punch_columns:
        for punch in parse_punches(ws.cell(row=row, column=col).value):
            if punch not in punches:
                punches.append(punch)
    return sorted(punches)


def _compute_day(current_date: date, punches: list[str]) -> DayComputation | None:
    if not punches:
        return None

    calculated = calculate_day(punches)
    return DayComputation(
        day=current_date.day,
        column=current_date.day,
        column_letter=str(current_date.day),
        raw_value=", ".join(punches),
        punches=punches,
        work_value=calculated.work_value,
        missing_count=calculated.missing_count,
        late_minutes=calculated.late_minutes,
        manual_checks=calculated.manual_checks,
    )


def _apply_override(item: DayComputation, override: dict) -> DayComputation:
    if not override:
        return item
    return DayComputation(
        day=item.day,
        column=item.column,
        column_letter=item.column_letter,
        raw_value=item.raw_value,
        punches=item.punches,
        work_value=override.get("work_value", item.work_value),
        missing_count=override.get("missing_count", item.missing_count),
        late_minutes=override.get("late_minutes", item.late_minutes),
        manual_checks=item.manual_checks,
    )


def _empty_day_result(value: date) -> dict:
    return {
        "day": value.day,
        "column": str(value.day),
        "punches": [],
        "work_value": None,
        "missing_count": None,
        "late_minutes": None,
    }


def _build_employee_preview(
    employee: Factory2Employee,
    profile_codes: set[str] | None = None,
    factory: str = "factory2",
) -> dict:
    entry = get_payroll_entry(employee.employee_code, factory)
    if profile_codes is not None and employee.employee_code not in profile_codes:
        entry = PayrollEntry()
    total_hours = _employee_total_hours(employee)
    monthly_salary = calculate_monthly_salary(entry)
    daily_salary = calculate_daily_salary(entry)
    hourly_salary = calculate_hourly_salary(entry)
    work_days = total_hours / 8
    final_salary = total_hours * hourly_salary + entry.bonus
    from app.services.bank_account_store import get_saved_account_number
    fallback_name = normalize_employee_name(employee.source_name) if employee.source_name != employee.employee_code else ""

    return {
        "employee_code": employee.employee_code,
        "name": normalize_employee_name(entry.name or fallback_name),
        "bank_account": get_saved_account_number(factory, employee.employee_code),
        "start_work_note": entry.start_work_note,
        "note": entry.note,
        "header_row": None,
        "result_row": None,
        "note_row": None,
        "total_hours": _round_number(total_hours),
        "monthly_salary": _round_optional_number(monthly_salary),
        "daily_salary_input": entry.daily_salary,
        "daily_salary": _round_number(daily_salary),
        "hourly_salary": _round_number(hourly_salary),
        "standard_work_days": entry.standard_work_days,
        "work_days": _round_number(work_days),
        "bonus": entry.bonus,
        "advance_or_penalty": None,
        "final_salary": _round_number(final_salary),
    }


def _write_employee_block(ws, start_row: int, employee: Factory2Employee, period: dict) -> int:
    title_row = start_row
    header_row = start_row + 1
    first_day_row = start_row + 2
    summary_name_row = first_day_row + len(employee.days)
    summary_header_row = summary_name_row + 1
    summary_value_row = summary_name_row + 2

    preview = _build_employee_preview(employee)
    month_label = period.get("label") or ""
    title = preview["name"] or employee.source_name or employee.employee_code
    _safe_merge(ws, title_row, 1, title_row, 13)
    ws.cell(row=title_row, column=1).value = title
    ws.cell(row=title_row, column=1).fill = PatternFill("solid", fgColor=YELLOW)
    ws.cell(row=title_row, column=1).font = Font(name=FONT_NAME, bold=True, size=10)
    ws.cell(row=title_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Mã",
        "Họ Và Tên",
        "Ngày",
        "Công",
        "Quên bấm",
        "Đi trễ",
        "Lần 1",
        "Lần 2",
        "Lần 3",
        "Lần 4",
        "Lần 5",
        "Lần 6",
        "Lần 7",
    ]
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=index)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=YELLOW if 4 <= index <= 6 else WHITE)
        cell.font = Font(name=FONT_NAME, bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    for offset, day in enumerate(employee.days):
        row = first_day_row + offset
        is_sunday = day.day.weekday() == 6
        fill = PatternFill("solid", fgColor=GRAY if is_sunday else WHITE)
        output_fill = PatternFill("solid", fgColor=YELLOW if day.punches else GRAY if is_sunday else WHITE)
        values = [
            employee.employee_code,
            preview["name"] or employee.source_name,
            day.day,
            day.computation.work_value if day.computation else None,
            day.computation.missing_count if day.computation else None,
            day.computation.late_minutes if day.computation else None,
            *day.punches[:7],
        ]
        values += [None] * (13 - len(values))
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = _thin_border()
            cell.fill = output_fill if 4 <= col <= 6 else fill
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col == 3:
                cell.number_format = "m/d/yyyy"
            elif col in {4, 5, 6}:
                cell.number_format = _number_format_for(value)

    _write_summary_rows(ws, summary_name_row, summary_header_row, summary_value_row, employee, preview, month_label)
    for row in range(start_row, summary_value_row + 1):
        ws.row_dimensions[row].height = 18
    ws.row_dimensions[summary_header_row].height = 27
    return summary_value_row


def _write_summary_rows(ws, name_row: int, header_row: int, value_row: int, employee: Factory2Employee, preview: dict, month_label: str) -> None:
    _safe_merge(ws, name_row, 1, name_row, 13)
    name_cell = ws.cell(row=name_row, column=1)
    name_cell.value = preview["name"] or employee.source_name or employee.employee_code
    name_cell.fill = PatternFill("solid", fgColor=YELLOW)
    name_cell.font = Font(name=FONT_NAME, bold=True, size=10)
    name_cell.alignment = Alignment(horizontal="left", vertical="center")
    name_cell.border = _thin_border()

    labels = [
        "Mã",
        "Họ Và Tên",
        "Tổng h làm",
        "Số Ngày đi làm",
        "Mức Lương",
        "Lương 1 Ngày",
        "Lương 1 Giờ",
        "H tăng",
        "Thưởng",
        "Ứng lương",
        f"Lương Tháng {month_label}",
        "",
        "",
    ]
    values = [
        employee.employee_code,
        preview["name"] or employee.source_name,
        preview["total_hours"],
        preview["work_days"],
        preview["monthly_salary"],
        preview["daily_salary"],
        preview["hourly_salary"],
        "",
        preview["bonus"],
        preview["advance_or_penalty"],
        preview["final_salary"],
        preview["start_work_note"],
        preview["note"],
    ]
    for col in range(1, 14):
        for row, value, bold in ((header_row, labels[col - 1], True), (value_row, values[col - 1], False)):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.fill = PatternFill("solid", fgColor=YELLOW)
            cell.font = Font(name=FONT_NAME, bold=bold or col in {2, 11}, size=8 if bold else 9, color=RED if col == 3 else "000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
            cell.border = _thin_border()
            if row == value_row and col in {5, 6, 7, 9, 10, 11}:
                cell.number_format = MONEY_FORMAT
            elif row == value_row and col in {3, 4}:
                cell.number_format = _number_format_for(value)


def _setup_output_sheet(ws, period: dict) -> None:
    title = f"Bảng công Xưởng 2 {period.get('label') or ''}".strip()
    _safe_merge(ws, 1, 1, 1, 13)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).font = Font(name=FONT_NAME, bold=True, size=16)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    widths = {
        1: 8,
        2: 24,
        3: 12,
        4: 8,
        5: 9,
        6: 8,
        7: 9,
        8: 9,
        9: 9,
        10: 9,
        11: 9,
        12: 9,
        13: 9,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _employee_total_hours(employee: Factory2Employee) -> float:
    total = 0.0
    for day in employee.days:
        value = day.computation.work_value if day.computation else None
        if isinstance(value, (int, float)):
            total += float(value)
    return total


def _review_overrides_by_employee_day(items: list[dict]) -> dict[tuple[str, int], dict]:
    result: dict[tuple[str, int], dict] = {}
    for item in items:
        employee_code = str(item.get("employee_code", "")).strip()
        day = item.get("day")
        if not employee_code or not isinstance(day, int):
            continue

        target = result.setdefault((employee_code, day), {})
        if "missing_count" in item:
            target["missing_count"] = item.get("missing_count")
        if "late_minutes" in item:
            target["late_minutes"] = item.get("late_minutes")
        if "work_value" in item:
            target["work_value"] = item.get("work_value")
    return result


def _plain_text(value: object) -> str:
    text = _clean_text(value).lower()
    # Vietnamese "đ" does not decompose under NFD, so map it explicitly
    # before stripping combining marks (otherwise "bắt đầu" becomes
    # "bat au" and the legacy field matcher cannot recognize it).
    text = text.replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _clean_text(value: object) -> str:
    return str(value or "").strip()


def _clean_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _safe_sheet_title(value: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", " ", value or "Xuong 2")
    return title[:31] or "Xuong 2"


def _safe_merge(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if not (
            merged_range.max_row < start_row
            or merged_range.min_row > end_row
            or merged_range.max_col < start_col
            or merged_range.min_col > end_col
        ):
            ws.unmerge_cells(str(merged_range))
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _number_format_for(value: object) -> str:
    if isinstance(value, (int, float)) and float(value).is_integer():
        return INTEGER_NUMBER_FORMAT
    return NUMBER_FORMAT


def _round_optional_number(value: float | None) -> int | float | None:
    if value is None:
        return None
    return _round_number(value)


def _round_number(value: float) -> int | float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded
