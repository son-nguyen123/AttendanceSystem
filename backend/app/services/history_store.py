import json
import shutil
import sqlite3
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.services.block_detector import detect_employee_blocks
from app.services.final_copy_overview import read_final_copy_overview
from app.services.period_detector import detect_period_from_workbook
from app.services.payroll_store import (
    get_payroll_entry,
    load_payroll_data,
    merge_missing_payroll_entries,
    normalize_employee_code,
)
from app.services.owner_profile_sync import sync_latest_final_copy_profile
from app.services.payroll_workbook import export_payroll_workbook, preview_payroll
from app.services.workbook_processor import analyze_workbook, export_processed_workbook


STORAGE_DIR = Path(__file__).resolve().parents[2] / "storage"
HISTORY_DIR = STORAGE_DIR / "history"
DB_PATH = STORAGE_DIR / "attendance_history.db"


def init_history_db() -> None:
    STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    with _connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS attendance_periods (
                id TEXT PRIMARY KEY,
                factory TEXT NOT NULL DEFAULT 'factory1',
                month INTEGER NOT NULL,
                year INTEGER NOT NULL,
                label TEXT NOT NULL,
                source_filename TEXT NOT NULL,
                source_path TEXT NOT NULL,
                output1_path TEXT NOT NULL,
                output2_path TEXT NOT NULL,
                sheet_name TEXT NOT NULL,
                block_count INTEGER NOT NULL DEFAULT 0,
                result_cells INTEGER NOT NULL DEFAULT 0,
                missing_cells INTEGER NOT NULL DEFAULT 0,
                late_cells INTEGER NOT NULL DEFAULT 0,
                manual_check_count INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS employee_monthly_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_id TEXT NOT NULL,
                employee_code TEXT NOT NULL,
                employee_name TEXT NOT NULL DEFAULT '',
                total_hours REAL NOT NULL DEFAULT 0,
                work_days REAL NOT NULL DEFAULT 0,
                monthly_salary REAL,
                daily_salary REAL NOT NULL DEFAULT 0,
                hourly_salary REAL NOT NULL DEFAULT 0,
                standard_work_days REAL NOT NULL DEFAULT 26,
                bonus REAL NOT NULL DEFAULT 0,
                advance_or_penalty REAL NOT NULL DEFAULT 0,
                final_salary REAL NOT NULL DEFAULT 0,
                note TEXT NOT NULL DEFAULT '',
                FOREIGN KEY(period_id) REFERENCES attendance_periods(id) ON DELETE CASCADE,
                UNIQUE(period_id, employee_code)
            );

            CREATE TABLE IF NOT EXISTS employee_daily_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_id TEXT NOT NULL,
                employee_code TEXT NOT NULL,
                day INTEGER NOT NULL,
                punches_json TEXT NOT NULL DEFAULT '[]',
                work_value TEXT,
                missing_count TEXT,
                late_minutes INTEGER,
                manual_checks_json TEXT NOT NULL DEFAULT '[]',
                FOREIGN KEY(period_id) REFERENCES attendance_periods(id) ON DELETE CASCADE,
                UNIQUE(period_id, employee_code, day)
            );

            CREATE INDEX IF NOT EXISTS idx_period_month_year
                ON attendance_periods(factory, year, month);
            CREATE INDEX IF NOT EXISTS idx_employee_monthly_code
                ON employee_monthly_records(employee_code);
            CREATE INDEX IF NOT EXISTS idx_employee_daily_lookup
                ON employee_daily_records(employee_code, period_id, day);
            """
        )
        period_columns = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(attendance_periods)").fetchall()
        }
        if "factory" not in period_columns:
            conn.execute(
                "ALTER TABLE attendance_periods ADD COLUMN factory TEXT NOT NULL DEFAULT 'factory1'"
            )

        columns = {
            row["name"]
            for row in conn.execute("PRAGMA table_info(employee_daily_records)").fetchall()
        }
        if "review_notes_json" not in columns:
            conn.execute(
                "ALTER TABLE employee_daily_records ADD COLUMN review_notes_json TEXT NOT NULL DEFAULT '[]'"
            )
        _bootstrap_payroll_data_from_history(conn)


def save_session_to_history(
    source_path: Path,
    source_filename: str,
    month: int | None = None,
    year: int | None = None,
    label: str | None = None,
    review_overrides: list[dict] | None = None,
    factory: str = "factory1",
) -> dict:
    init_history_db()
    detected = detect_period_from_workbook(source_path)
    selected_month = month or _as_int(detected.get("month"))
    selected_year = year or _as_int(detected.get("year"))
    if not selected_month or not selected_year:
        raise ValueError("Chưa xác định được tháng/năm của file")
    if selected_month < 1 or selected_month > 12:
        raise ValueError("Tháng phải nằm trong khoảng 1-12")

    period_id = uuid4().hex
    period_dir = HISTORY_DIR / period_id
    period_dir.mkdir(parents=True, exist_ok=True)

    suffix = source_path.suffix or ".xlsx"
    original_path = period_dir / f"original{suffix}"
    output1_path = period_dir / "attendance_output.xlsx"
    output2_path = period_dir / "payroll_private_output.xlsx"

    overrides = review_overrides or []
    factory = _normalize_factory(factory)
    profile_sync = sync_latest_final_copy_profile(source_path, factory)
    profile_codes = {
        str(code).strip()
        for code in profile_sync.get("profile_codes", [])
        if str(code).strip()
    } if profile_sync.get("status") == "ok" else set()

    shutil.copy2(source_path, original_path)
    export_processed_workbook(source_path, output1_path, review_overrides=overrides, factory=factory)
    export_payroll_workbook(
        source_path, output2_path, review_overrides=overrides,
        profile_codes=profile_codes, factory=factory,
    )

    analysis = analyze_workbook(source_path)
    _apply_review_overrides_to_analysis(analysis, overrides)
    payroll_preview = preview_payroll(
        source_path, review_overrides=overrides, profile_codes=profile_codes, factory=factory,
    )
    payroll_by_code = {item["employee_code"]: item for item in payroll_preview.get("employees", [])}
    manual_by_employee_day = _manual_checks_by_employee_day(analysis.get("manual_checks", []))
    review_notes_by_employee_day = _review_notes_by_employee_day(overrides)

    now = datetime.now().isoformat(timespec="seconds")
    period_label = label.strip() if label and label.strip() else f"Tháng {selected_month:02d}/{selected_year}"

    with _connect() as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute(
            """
            INSERT INTO attendance_periods (
                id, factory, month, year, label, source_filename, source_path, output1_path, output2_path,
                sheet_name, block_count, result_cells, missing_cells, late_cells, manual_check_count,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                period_id,
                factory,
                selected_month,
                selected_year,
                period_label,
                source_filename,
                str(original_path),
                str(output1_path),
                str(output2_path),
                analysis.get("sheet_name", ""),
                analysis["summary"]["blocks"],
                analysis["summary"]["result_cells"],
                analysis["summary"]["missing_cells"],
                analysis["summary"]["late_cells"],
                analysis["summary"]["manual_check_count"],
                now,
                now,
            ),
        )

        merge_missing_payroll_entries(
            {
                normalize_employee_code(block.get("employee_code")): {
                    "name": payroll_by_code.get(normalize_employee_code(block.get("employee_code")), {}).get("name", "")
                }
                for block in analysis.get("blocks", [])
            },
            factory=factory,
        )

        for block in analysis.get("blocks", []):
            employee_code = normalize_employee_code(block["employee_code"])
            payroll = payroll_by_code.get(employee_code, {})
            employee_name = get_payroll_entry(employee_code, factory).name or payroll.get("name") or ""
            conn.execute(
                """
                INSERT INTO employee_monthly_records (
                    period_id, employee_code, employee_name, total_hours, work_days,
                    monthly_salary, daily_salary, hourly_salary, standard_work_days,
                    bonus, advance_or_penalty, final_salary, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    period_id,
                    employee_code,
                    employee_name,
                    _num(payroll.get("total_hours")),
                    _num(payroll.get("work_days")),
                    payroll.get("monthly_salary"),
                    _num(payroll.get("daily_salary")),
                    _num(payroll.get("hourly_salary")),
                    _num(payroll.get("standard_work_days"), 26),
                    _num(payroll.get("bonus")),
                    _num(payroll.get("advance_or_penalty")),
                    _num(payroll.get("final_salary")),
                    payroll.get("note") or "",
                ),
            )

            for result in block.get("results", []):
                manual_messages = manual_by_employee_day.get((employee_code, result["day"]), [])
                review_notes = review_notes_by_employee_day.get((employee_code, result["day"]), [])
                conn.execute(
                    """
                    INSERT INTO employee_daily_records (
                        period_id, employee_code, day, punches_json, work_value,
                        missing_count, late_minutes, manual_checks_json, review_notes_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        period_id,
                        employee_code,
                        result["day"],
                        json.dumps(result.get("punches", []), ensure_ascii=False),
                        _text_or_none(result.get("work_value")),
                        _text_or_none(result.get("missing_count")),
                        result.get("late_minutes"),
                        json.dumps(manual_messages, ensure_ascii=False),
                        json.dumps(review_notes, ensure_ascii=False),
                    ),
                )

    detail = get_period_detail(period_id)
    # A saved history period is an archive, never a profile-data source. The
    # only source allowed to refresh reusable employee data is a final copy
    # explicitly selected by the owner.
    detail["profile_sync"] = profile_sync
    return detail


def list_periods(
    employee_code: str | None = None,
    month: int | None = None,
    year: int | None = None,
    factory: str | None = None,
) -> list[dict]:
    init_history_db()
    filters = []
    params: list[object] = []
    if factory:
        filters.append("p.factory = ?")
        params.append(_normalize_factory(factory))
    join_employee = bool(employee_code)
    if join_employee:
        filters.append("em.employee_code LIKE ?")
        params.append(f"%{employee_code.strip()}%")
    if month:
        filters.append("p.month = ?")
        params.append(month)
    if year:
        filters.append("p.year = ?")
        params.append(year)

    sql = """
        SELECT DISTINCT p.*
        FROM attendance_periods p
    """
    if join_employee:
        sql += " JOIN employee_monthly_records em ON em.period_id = p.id"
    if filters:
        sql += " WHERE " + " AND ".join(filters)
    sql += " ORDER BY p.year DESC, p.month DESC, p.created_at DESC"

    with _connect() as conn:
        rows = conn.execute(sql, params).fetchall()
    return [_period_row_to_dict(row) for row in rows]


def get_period_detail(period_id: str) -> dict:
    init_history_db()
    with _connect() as conn:
        period = conn.execute("SELECT * FROM attendance_periods WHERE id = ?", (period_id,)).fetchone()
        if period is None:
            raise KeyError(period_id)
        employees = conn.execute(
            """
            SELECT *
            FROM employee_monthly_records
            WHERE period_id = ?
            ORDER BY CAST(employee_code AS INTEGER), employee_code
            """,
            (period_id,),
        ).fetchall()
        daily_rows = conn.execute(
            """
            SELECT *
            FROM employee_daily_records
            WHERE period_id = ?
            ORDER BY CAST(employee_code AS INTEGER), employee_code, day
            """,
            (period_id,),
        ).fetchall()

    daily_by_code: dict[str, list[dict]] = {}
    for row in daily_rows:
        item = dict(row)
        item["punches"] = json.loads(item.pop("punches_json") or "[]")
        item["manual_checks"] = json.loads(item.pop("manual_checks_json") or "[]")
        item["review_notes"] = json.loads(item.pop("review_notes_json") or "[]")
        item["work_value"] = _restore_number(item["work_value"])
        item["missing_count"] = _restore_number(item["missing_count"])
        daily_by_code.setdefault(item["employee_code"], []).append(item)

    employee_items = []
    for row in employees:
        item = dict(row)
        item["employee_name"] = _current_employee_name(item["employee_code"], item["employee_name"], period["factory"])
        item["daily_records"] = daily_by_code.get(item["employee_code"], [])
        employee_items.append(item)

    return {"period": _period_row_to_dict(period), "employees": employee_items}


def update_employee_monthly_record(period_id: str, employee_code: str, updates: dict) -> dict:
    init_history_db()
    normalized_code = normalize_employee_code(employee_code)
    if not normalized_code:
        raise KeyError(employee_code)

    now = datetime.now().isoformat(timespec="seconds")
    daily_updates = updates.get("daily_records", [])
    if not isinstance(daily_updates, list):
        daily_updates = []

    with _connect() as conn:
        period = conn.execute("SELECT * FROM attendance_periods WHERE id = ?", (period_id,)).fetchone()
        if period is None:
            raise KeyError(period_id)

        existing = conn.execute(
            """
            SELECT *
            FROM employee_monthly_records
            WHERE period_id = ? AND employee_code = ?
            """,
            (period_id, normalized_code),
        ).fetchone()
        if existing is None:
            raise KeyError(normalized_code)

        for item in daily_updates:
            day = item.get("day")
            if not isinstance(day, int):
                continue
            conn.execute(
                """
                UPDATE employee_daily_records
                SET work_value = ?, missing_count = ?, late_minutes = ?, review_notes_json = ?
                WHERE period_id = ? AND employee_code = ? AND day = ?
                """,
                (
                    _text_or_none(item.get("work_value")),
                    _text_or_none(item.get("missing_count")),
                    item.get("late_minutes"),
                    json.dumps(item.get("review_notes", []), ensure_ascii=False),
                    period_id,
                    normalized_code,
                    day,
                ),
            )

        daily_rows = conn.execute(
            """
            SELECT day, work_value
            FROM employee_daily_records
            WHERE period_id = ? AND employee_code = ?
            """,
            (period_id, normalized_code),
        ).fetchall()
        total_hours = sum(_num(row["work_value"]) for row in daily_rows)
        work_days = total_hours / 8 if total_hours else 0
        hourly_salary = _num(updates.get("hourly_salary"), _num(existing["hourly_salary"]))
        daily_salary = hourly_salary * 8 if hourly_salary else _num(existing["daily_salary"])
        standard_work_days = 26
        monthly_salary = daily_salary * standard_work_days if daily_salary else _num(existing["monthly_salary"])
        bonus = _num(updates.get("bonus"), _num(existing["bonus"]))
        advance_or_penalty = _num(updates.get("advance_or_penalty"), _num(existing["advance_or_penalty"]))
        final_salary = total_hours * hourly_salary + bonus - advance_or_penalty
        values = {
            "employee_name": str(updates.get("employee_name", existing["employee_name"]) or "").strip(),
            "total_hours": _round_number(total_hours),
            "work_days": _round_number(work_days),
            "monthly_salary": _round_number(monthly_salary),
            "daily_salary": _round_number(daily_salary),
            "hourly_salary": _round_number(hourly_salary),
            "standard_work_days": _round_number(standard_work_days),
            "bonus": bonus,
            "advance_or_penalty": advance_or_penalty,
            "final_salary": _round_number(final_salary),
            "note": str(updates.get("note", existing["note"]) or "").strip(),
        }
        conn.execute(
            """
            UPDATE employee_monthly_records
            SET employee_name = ?, total_hours = ?, work_days = ?, monthly_salary = ?, daily_salary = ?,
                hourly_salary = ?, standard_work_days = ?, bonus = ?, advance_or_penalty = ?, final_salary = ?, note = ?
            WHERE period_id = ? AND employee_code = ?
            """,
            (
                values["employee_name"],
                values["total_hours"],
                values["work_days"],
                values["monthly_salary"],
                values["daily_salary"],
                values["hourly_salary"],
                values["standard_work_days"],
                values["bonus"],
                values["advance_or_penalty"],
                values["final_salary"],
                values["note"],
                period_id,
                normalized_code,
            ),
        )
        conn.execute(
            "UPDATE attendance_periods SET updated_at = ? WHERE id = ?",
            (now, period_id),
        )

    _rewrite_history_output_files(dict(period), normalized_code, daily_updates, values)
    return get_period_detail(period_id)


def search_employee_history(
    employee_code: str,
    month: int | None = None,
    year: int | None = None,
    factory: str | None = None,
) -> list[dict]:
    init_history_db()
    filters = ["em.employee_code LIKE ?"]
    params: list[object] = [f"%{employee_code.strip()}%"]
    if factory:
        filters.append("p.factory = ?")
        params.append(_normalize_factory(factory))
    if month:
        filters.append("p.month = ?")
        params.append(month)
    if year:
        filters.append("p.year = ?")
        params.append(year)

    with _connect() as conn:
        rows = conn.execute(
            f"""
            SELECT
                p.id AS period_id, p.factory, p.month, p.year, p.label, p.created_at,
                em.employee_code, em.employee_name, em.total_hours, em.work_days,
                em.final_salary, em.note
            FROM employee_monthly_records em
            JOIN attendance_periods p ON p.id = em.period_id
            WHERE {" AND ".join(filters)}
            ORDER BY p.year DESC, p.month DESC, p.created_at DESC, em.employee_code
            """,
            params,
        ).fetchall()
    results = []
    for row in rows:
        item = dict(row)
        item["employee_name"] = _current_employee_name(item["employee_code"], item["employee_name"], item["factory"])
        results.append(item)
    return results


def list_known_employee_codes(factory: str | None = None) -> list[str]:
    init_history_db()
    filters = []
    params: list[object] = []
    if factory:
        filters.append("p.factory = ?")
        params.append(_normalize_factory(factory))
    where_sql = f"WHERE {' AND '.join(filters)}" if filters else ""
    with _connect() as conn:
        rows = conn.execute(
            f"""
            SELECT DISTINCT employee_code
            FROM employee_monthly_records em
            JOIN attendance_periods p ON p.id = em.period_id
            {where_sql}
            ORDER BY CAST(employee_code AS INTEGER), employee_code
            """,
            params,
        ).fetchall()
    return [str(row["employee_code"]) for row in rows]


def get_latest_period_snapshot(factory: str | None = None) -> dict:
    init_history_db()
    factory_filter = _normalize_factory(factory) if factory else None
    with _connect() as conn:
        if factory_filter:
            period = conn.execute(
                """
                SELECT *
                FROM attendance_periods
                WHERE factory = ?
                ORDER BY year DESC, month DESC, created_at DESC
                LIMIT 1
                """,
                (factory_filter,),
            ).fetchone()
        else:
            period = conn.execute(
                """
                SELECT *
                FROM attendance_periods
                ORDER BY year DESC, month DESC, created_at DESC
                LIMIT 1
                """
            ).fetchone()
        if period is None:
            return {"period": None, "employee_codes": []}

        rows = conn.execute(
            """
            SELECT DISTINCT employee_code
            FROM employee_monthly_records
            WHERE period_id = ?
            ORDER BY CAST(employee_code AS INTEGER), employee_code
            """,
            (period["id"],),
        ).fetchall()

    return {
        "period": _period_row_to_dict(period),
        "employee_codes": [str(row["employee_code"]) for row in rows],
    }


def get_review_memory(month: int, year: int, factory: str | None = None) -> dict:
    init_history_db()
    factory_filter = _normalize_factory(factory) if factory else None
    with _connect() as conn:
        if factory_filter:
            period = conn.execute(
                """
                SELECT *
                FROM attendance_periods
                WHERE month = ? AND year = ? AND factory = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (month, year, factory_filter),
            ).fetchone()
        else:
            period = conn.execute(
                """
                SELECT *
                FROM attendance_periods
                WHERE month = ? AND year = ?
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (month, year),
            ).fetchone()
        if period is None:
            return {"period": None, "records": []}

        rows = conn.execute(
            """
            SELECT *
            FROM employee_daily_records
            WHERE period_id = ?
            ORDER BY CAST(employee_code AS INTEGER), employee_code, day
            """,
            (period["id"],),
        ).fetchall()

    records = []
    for row in rows:
        manual_checks = _json_list(row["manual_checks_json"])
        review_notes = _json_list(row["review_notes_json"] if "review_notes_json" in row.keys() else "[]")
        missing_count = _restore_number(row["missing_count"])
        late_minutes = row["late_minutes"]
        work_value = _restore_number(row["work_value"])
        records.append(
            {
                "employee_code": str(row["employee_code"]),
                "day": int(row["day"]),
                "punches": _json_list(row["punches_json"]),
                "work_value": work_value,
                "missing_count": missing_count,
                "late_minutes": late_minutes,
                "manual_checks": manual_checks,
                "review_notes": review_notes,
            }
        )

    return {"period": _period_row_to_dict(period), "records": records}


def get_attendance_overview(year: int | None = None, factory: str | None = None) -> dict:
    init_history_db()
    factory = _normalize_factory(factory)
    final_copies = _safe_final_copies(factory)
    final_years = {int(item["year"]) for item in final_copies if item.get("year")}
    with _connect() as conn:
        db_years = [
            int(row["year"])
            for row in conn.execute(
                "SELECT DISTINCT year FROM attendance_periods WHERE factory = ? ORDER BY year DESC",
                (factory,),
            ).fetchall()
        ]
        years = sorted(set(db_years) | final_years, reverse=True)
        if not years:
            return {
                "factory": factory,
                "years": [],
                "year": year,
                "latest_month": None,
                "employees": [],
                "summary": {
                    "active_count": 0,
                    "inactive_count": 0,
                    "employee_count": 0,
                    "total_work_days": 0,
                    "total_hours": 0,
                },
            }

        selected_year = year if year in years else years[0]
        selected_final_copies = [item for item in final_copies if int(item.get("year") or 0) == selected_year]
        final_months = {int(item["month"]) for item in selected_final_copies if item.get("month")}
        period_rows = conn.execute(
            """
            SELECT *
            FROM attendance_periods
            WHERE year = ? AND factory = ?
            ORDER BY month ASC, created_at ASC
            """,
            (selected_year, factory),
        ).fetchall()
        latest_by_month: dict[int, sqlite3.Row] = {}
        for row in period_rows:
            month = int(row["month"])
            latest_by_month[month] = row

        periods = list(latest_by_month.values())
        period_ids = [str(row["id"]) for row in periods]
        if not period_ids and not selected_final_copies:
            return {
                "factory": factory,
                "years": years,
                "year": selected_year,
                "latest_month": None,
                "employees": [],
                "summary": {
                    "active_count": 0,
                    "inactive_count": 0,
                    "employee_count": 0,
                    "total_work_days": 0,
                    "total_hours": 0,
                },
            }

        if period_ids:
            placeholders = ",".join("?" for _ in period_ids)
            monthly_rows = conn.execute(
                f"""
                SELECT em.*, p.month
                FROM employee_monthly_records em
                JOIN attendance_periods p ON p.id = em.period_id
                WHERE em.period_id IN ({placeholders})
                """,
                period_ids,
            ).fetchall()
            daily_rows = conn.execute(
                f"""
                SELECT ed.*, p.month
                FROM employee_daily_records ed
                JOIN attendance_periods p ON p.id = ed.period_id
                WHERE ed.period_id IN ({placeholders})
                """,
                period_ids,
            ).fetchall()
        else:
            monthly_rows = []
            daily_rows = []

    latest_month = max(set(latest_by_month) | final_months)
    overview_month_count = max(1, len(set(latest_by_month) | final_months))
    daily_stats: dict[tuple[str, int], dict[str, int]] = {}
    for row in daily_rows:
        key = (str(row["employee_code"]), int(row["month"]))
        stats = daily_stats.setdefault(key, {"late_count": 0, "issue_count": 0})
        if row["late_minutes"] is not None:
            stats["late_count"] += 1

        missing_count = str(row["missing_count"] or "").strip()
        manual_checks = _json_list(row["manual_checks_json"])
        review_notes = _json_list(row["review_notes_json"] if "review_notes_json" in row.keys() else "[]")
        if missing_count == "?" or manual_checks or review_notes:
            stats["issue_count"] += 1

    employee_names = _employee_names_by_code(factory)
    employees: dict[str, dict] = {}
    for row in monthly_rows:
        employee_code = str(row["employee_code"])
        current_name = employee_names.get(employee_code, "")
        item = employees.setdefault(
            employee_code,
            {
                "employee_code": employee_code,
                "employee_name": current_name,
                "months": [_empty_overview_month(month) for month in range(1, 13)],
                "total_hours": 0.0,
                "total_work_days": 0.0,
                "total_late_count": 0,
                "total_issue_count": 0,
                "active": False,
            },
        )
        if not current_name and str(row["employee_name"] or "").strip():
            item["employee_name"] = str(row["employee_name"] or "")

        month = int(row["month"])
        stats = daily_stats.get((employee_code, month), {"late_count": 0, "issue_count": 0})
        total_hours = _num(row["total_hours"])
        work_days = _num(row["work_days"])
        month_item = {
            "month": month,
            "total_hours": _round_number(total_hours),
            "work_days": _round_number(work_days),
            "late_count": stats["late_count"],
            "issue_count": stats["issue_count"],
        }
        item["months"][month - 1] = month_item

    for copy_item in selected_final_copies:
        month = int(copy_item["month"])
        # A saved history period is the trusted result produced by the app's
        # rules. The final-copy workbook is only a fallback for months where
        # no history exists, because its external layout may omit review data.
        if month in latest_by_month:
            continue
        try:
            final_rows = read_final_copy_overview(Path(str(copy_item["path"])), month)
        except Exception:
            final_rows = []
        for row in final_rows:
            employee_code = str(row["employee_code"])
            current_name = employee_names.get(employee_code, "")
            item = employees.setdefault(
                employee_code,
                {
                    "employee_code": employee_code,
                    "employee_name": current_name or str(row.get("employee_name") or ""),
                    "months": [_empty_overview_month(item_month) for item_month in range(1, 13)],
                    "total_hours": 0.0,
                    "total_work_days": 0.0,
                    "total_late_count": 0,
                    "total_issue_count": 0,
                    "active": False,
                },
            )
            if not item["employee_name"] and str(row.get("employee_name") or "").strip():
                item["employee_name"] = str(row.get("employee_name") or "")
            item["months"][month - 1] = {
                "month": month,
                "total_hours": row["total_hours"],
                "work_days": row["work_days"],
                "late_count": 0,
                "issue_count": 0,
                "source": "final_copy",
            }

    for item in employees.values():
        for month_item in item["months"]:
            item["total_hours"] += float(month_item["total_hours"])
            item["total_work_days"] += float(month_item["work_days"])
            item["total_late_count"] += int(month_item["late_count"])
            item["total_issue_count"] += int(month_item["issue_count"])

        latest_item = item["months"][latest_month - 1]
        item["active"] = float(latest_item["work_days"]) > 0 or float(latest_item["total_hours"]) > 0
        item["average_work_days"] = _round_number(item["total_work_days"] / overview_month_count)
        item["total_hours"] = _round_number(item["total_hours"])
        item["total_work_days"] = _round_number(item["total_work_days"])

    employee_items = sorted(
        employees.values(),
        key=lambda item: (
            0 if item["active"] else 1,
            -float(item["total_work_days"]),
            _employee_sort_key(item["employee_code"]),
        ),
    )
    active_count = sum(1 for item in employee_items if item["active"])
    total_work_days = sum(float(item["total_work_days"]) for item in employee_items)
    total_hours = sum(float(item["total_hours"]) for item in employee_items)

    return {
        "factory": factory,
        "years": years,
        "year": selected_year,
        "latest_month": latest_month,
        "employees": employee_items,
        "summary": {
            "active_count": active_count,
            "inactive_count": len(employee_items) - active_count,
            "employee_count": len(employee_items),
            "total_work_days": _round_number(total_work_days),
            "total_hours": _round_number(total_hours),
        },
        "source": {
            "mode": "prefer_history_fallback_final_copy",
            "factory": factory,
            "final_copy_months": sorted(final_months),
            "machine_months": sorted(latest_by_month),
            "fallback_final_copy_months": sorted(final_months - set(latest_by_month)),
        },
    }


def get_period_file(period_id: str, kind: str) -> Path:
    init_history_db()
    column_by_kind = {
        "original": "source_path",
        "output1": "output1_path",
        "output2": "output2_path",
    }
    column = column_by_kind.get(kind)
    if column is None:
        raise ValueError("Loại file không hợp lệ")

    with _connect() as conn:
        row = conn.execute(f"SELECT {column} FROM attendance_periods WHERE id = ?", (period_id,)).fetchone()
    if row is None:
        raise KeyError(period_id)

    path = Path(row[0])
    if not path.exists():
        raise FileNotFoundError(path)
    return path


def delete_period(period_id: str) -> None:
    init_history_db()
    with _connect() as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        row = conn.execute("SELECT source_path FROM attendance_periods WHERE id = ?", (period_id,)).fetchone()
        if row is None:
            raise KeyError(period_id)
        conn.execute("DELETE FROM attendance_periods WHERE id = ?", (period_id,))

    period_dir = HISTORY_DIR / period_id
    if period_dir.exists():
        shutil.rmtree(period_dir)


def _rewrite_history_output_files(
    period: dict,
    employee_code: str,
    daily_updates: list[dict],
    monthly_values: dict,
) -> None:
    daily_by_day = {
        int(item["day"]): item
        for item in daily_updates
        if isinstance(item.get("day"), int)
    }
    for path_key in ("output1_path", "output2_path"):
        path = Path(str(period.get(path_key) or ""))
        if not path.exists():
            continue

        wb = load_workbook(path, data_only=False)
        ws = _select_history_workbook_sheet(wb)
        block = next(
            (item for item in detect_employee_blocks(ws) if normalize_employee_code(item.employee_code) == employee_code),
            None,
        )
        if block is None:
            continue

        for col in range(1, 32):
            day_value = ws.cell(row=block.day_row, column=col).value
            if not isinstance(day_value, int) or day_value not in daily_by_day:
                continue

            item = daily_by_day[day_value]
            work_value = item.get("work_value")
            missing_count = item.get("missing_count")
            late_minutes = item.get("late_minutes")
            ws.cell(row=block.result_row, column=col).value = _restore_workbook_value(work_value)
            ws.cell(row=block.missing_row, column=col).value = _restore_workbook_value(missing_count)
            ws.cell(row=block.late_row, column=col).value = _restore_workbook_value(late_minutes)

        ws.cell(row=block.result_row, column=32).value = monthly_values["total_hours"]
        if path_key == "output2_path":
            note_row = block.header_row + 7
            ws.cell(row=block.result_row, column=35).value = monthly_values["employee_name"]
            ws.cell(row=block.result_row, column=36).value = monthly_values["monthly_salary"]
            ws.cell(row=block.result_row, column=37).value = monthly_values["daily_salary"]
            ws.cell(row=block.result_row, column=38).value = monthly_values["hourly_salary"]
            ws.cell(row=block.result_row, column=39).value = monthly_values["work_days"]
            ws.cell(row=block.result_row, column=41).value = monthly_values["bonus"]
            ws.cell(row=block.result_row, column=43).value = monthly_values["advance_or_penalty"]
            ws.cell(row=block.result_row, column=44).value = monthly_values["final_salary"]
            ws.cell(row=note_row, column=35).value = monthly_values["note"]

        wb.save(path)


def _select_history_workbook_sheet(wb):
    best_sheet = None
    best_count = -1
    for ws in wb.worksheets:
        count = sum(1 for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == "Att. Time")
        if count > best_count:
            best_sheet = ws
            best_count = count
    if best_sheet is None or best_count <= 0:
        raise ValueError("Không tìm thấy sheet chấm công có dòng Att. Time")
    return best_sheet


def _restore_workbook_value(value: object) -> int | float | str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        numeric = float(text.replace(",", "."))
    except ValueError:
        return text
    return int(numeric) if numeric.is_integer() else numeric


@contextmanager
def _connect() -> Iterator[sqlite3.Connection]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def _bootstrap_payroll_data_from_history(conn: sqlite3.Connection) -> None:
    rows = conn.execute(
        """
        SELECT em.employee_code, em.employee_name, em.note, p.factory
        FROM employee_monthly_records em
        JOIN attendance_periods p ON p.id = em.period_id
        WHERE TRIM(em.employee_code) != ''
        ORDER BY p.year DESC, p.month DESC, p.created_at DESC
        """
    ).fetchall()

    defaults_by_factory: dict[str, dict[str, dict]] = {"factory1": {}, "factory2": {}}
    existing_by_factory = {factory: set(load_payroll_data(factory).keys()) for factory in defaults_by_factory}
    for row in rows:
        factory = _normalize_factory(row["factory"])
        employee_code = normalize_employee_code(row["employee_code"])
        defaults = defaults_by_factory[factory]
        if not employee_code or employee_code in existing_by_factory[factory] or employee_code in defaults:
            continue
        defaults[employee_code] = {
            "name": str(row["employee_name"] or "").strip(),
            "note": str(row["note"] or "").strip(),
        }

    for factory, defaults in defaults_by_factory.items():
        merge_missing_payroll_entries(defaults, factory=factory)


def _current_employee_name(employee_code: object, fallback: object = "", factory: str = "factory1") -> str:
    fallback_name = str(fallback or "").strip()
    if fallback_name:
        return fallback_name
    current_name = get_payroll_entry(normalize_employee_code(employee_code), factory).name.strip()
    return current_name


def _safe_final_copies(factory: str | None = None) -> list[dict]:
    try:
        from app.services.cloud_sync import list_drive_final_copies

        return list_drive_final_copies(factory=factory)
    except Exception:
        return []


def _employee_names_by_code(factory: str = "factory1") -> dict[str, str]:
    return {
        normalize_employee_code(code): str(entry.get("name") or "").strip()
        for code, entry in load_payroll_data(factory).items()
        if isinstance(entry, dict)
    }


def _period_row_to_dict(row: sqlite3.Row) -> dict:
    item = dict(row)
    item["factory"] = _normalize_factory(item.get("factory"))
    item["has_output1"] = Path(item["output1_path"]).exists()
    item["has_output2"] = Path(item["output2_path"]).exists()
    return item


def _normalize_factory(value: object) -> str:
    return "factory2" if str(value or "").strip() == "factory2" else "factory1"


def _manual_checks_by_employee_day(items: list[dict]) -> dict[tuple[str, int], list[str]]:
    result: dict[tuple[str, int], list[str]] = {}
    for item in items:
        key = (item["employee_code"], item["day"])
        result.setdefault(key, []).extend(item.get("messages", []))
    return result


def _apply_review_overrides_to_analysis(analysis: dict, review_overrides: list[dict]) -> None:
    overrides = _review_overrides_by_employee_day(review_overrides)
    summary = analysis["summary"]
    summary.update({"result_cells": 0, "missing_cells": 0, "late_cells": 0})

    for block in analysis.get("blocks", []):
        employee_code = block["employee_code"]
        for result in block.get("results", []):
            override = overrides.get((employee_code, result["day"]), {})
            if "missing_count" in override:
                result["missing_count"] = override["missing_count"]
            if "late_minutes" in override:
                result["late_minutes"] = override["late_minutes"]
            if "work_value" in override:
                result["work_value"] = override["work_value"]

            if result.get("work_value") is not None and result.get("work_value") != "":
                summary["result_cells"] += 1
            if result.get("missing_count") is not None and result.get("missing_count") != "":
                summary["missing_cells"] += 1
            if result.get("late_minutes") is not None and result.get("late_minutes") != "":
                summary["late_cells"] += 1


def _review_overrides_by_employee_day(items: list[dict]) -> dict[tuple[str, int], dict]:
    result: dict[tuple[str, int], dict] = {}
    for item in items:
        employee_code = str(item.get("employee_code", "")).strip()
        day = item.get("day")
        if not employee_code or not isinstance(day, int):
            continue

        target = result.setdefault((employee_code, day), {})
        if "missing_count" in item:
            target["missing_count"] = item.get("missing_count")
        if "late_minutes" in item:
            target["late_minutes"] = item.get("late_minutes")
        if "work_value" in item:
            target["work_value"] = item.get("work_value")
    return result


def _review_notes_by_employee_day(items: list[dict]) -> dict[tuple[str, int], list[str]]:
    result: dict[tuple[str, int], list[str]] = {}
    type_labels = {"missing": "Quên bấm / chưa rõ", "late": "Đi trễ"}
    status_labels = {"ok": "Đã OK", "edited": "Đã sửa", "pending": "Chưa xác nhận"}

    for item in items:
        employee_code = str(item.get("employee_code", "")).strip()
        day = item.get("day")
        if not employee_code or not isinstance(day, int):
            continue

        notes = item.get("review_notes")
        if isinstance(notes, list) and notes:
            result.setdefault((employee_code, day), []).extend(str(note) for note in notes)
            continue

        review_type = type_labels.get(str(item.get("type") or ""), "Kiểm tra")
        status = status_labels.get(str(item.get("status") or ""), "Chưa xác nhận")
        result.setdefault((employee_code, day), []).append(f"{status}: {review_type}")
    return result


def _as_int(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _num(value: object, fallback: float = 0) -> float:
    if value is None or value == "":
        return fallback
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return fallback


def _text_or_none(value: object) -> str | None:
    if value is None:
        return None
    return str(value)


def _restore_number(value: str | None) -> int | float | str | None:
    if value is None:
        return None
    try:
        parsed = float(value)
    except ValueError:
        return value
    return int(parsed) if parsed.is_integer() else parsed


def _empty_overview_month(month: int) -> dict:
    return {
        "month": month,
        "total_hours": 0,
        "work_days": 0,
        "late_count": 0,
        "issue_count": 0,
    }


def _json_list(value: object) -> list:
    try:
        parsed = json.loads(str(value or "[]"))
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def _employee_sort_key(value: str) -> tuple[int, object]:
    try:
        return (0, int(value))
    except ValueError:
        return (1, value)


def _round_number(value: float) -> int | float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded
