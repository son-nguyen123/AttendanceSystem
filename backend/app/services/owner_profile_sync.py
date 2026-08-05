from __future__ import annotations

import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.period_detector import detect_period_from_workbook
from app.services.payroll_store import merge_payroll_profile_updates, normalize_employee_code


SUMMARY_MIN_COL = 32


def sync_latest_final_copy_profile(source_path: Path, factory: str) -> dict[str, Any]:
    """Refresh from the newest final copy without replacing manual entries."""
    try:
        # Imported lazily to avoid a cloud/profile import cycle at startup.
        from app.services.cloud_sync import list_drive_final_copies

        copies = list_drive_final_copies(factory=factory)
        if not copies:
            return _empty_result("", "no_final_copy")

        period = detect_period_from_workbook(source_path)
        current_key = (int(period.get("year") or 0), int(period.get("month") or 0))
        eligible = [
            item
            for item in copies
            if current_key == (0, 0)
            or (int(item.get("year") or 0), int(item.get("month") or 0)) <= current_key
        ]
        if not eligible:
            return _empty_result("", "no_eligible_final_copy")

        latest = max(
            eligible,
            key=lambda item: (
                int(item.get("year") or 0),
                int(item.get("month") or 0),
                str(item.get("modified_at") or ""),
            ),
        )
        path = Path(str(latest.get("path") or ""))
        if not path.exists():
            return _empty_result(str(path), "final_copy_missing")

        return sync_owner_profiles_from_workbook(
            path,
            factory=factory,
            month=int(latest.get("month") or 0) or None,
            year=int(latest.get("year") or 0) or None,
            source_kind="final_copy",
            overwrite_manual=False,
        )
    except Exception as exc:
        return {**_empty_result("", "final_copy_sync_failed"), "error": str(exc)}


def sync_owner_profiles_from_workbook(
    path: Path,
    *,
    factory: str = "factory1",
    month: int | None = None,
    year: int | None = None,
    source_kind: str = "workbook",
    overwrite_manual: bool = False,
) -> dict[str, Any]:
    if not path.exists() or not path.is_file():
        return _empty_result(str(path), "file_not_found")
    if source_kind != "final_copy":
        return _empty_result(str(path), "final_copy_only")

    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, data_only=True, keep_vba=keep_vba)
    try:
        updates: dict[str, dict[str, Any]] = {}
        for worksheet in workbook.worksheets:
            updates.update(_extract_profiles_from_sheet(worksheet))

        # Bank accounts live in their own factory-partitioned registry. Keep
        # that registry in sync with an explicitly uploaded final copy so the
        # next Output 2 can reuse the account without another manual import.
        from app.services.bank_account_store import sync_accounts_from_final_copy

        bank_sync = sync_accounts_from_final_copy(factory, updates)

        result = merge_payroll_profile_updates(
            updates,
            factory=factory,
            source_month=month,
            source_year=year,
            source_kind=source_kind,
            source_name=path.name,
            overwrite_manual=overwrite_manual,
        )
        result.update(
            {
                "status": "ok",
                "source_path": str(path),
                "source_month": month,
                "source_year": year,
                "source_kind": source_kind,
                "profile_count": len(updates),
                "profile_codes": sorted(updates),
                "fields": ["name", "bank_account", "start_work_note", "note", "monthly_salary", "daily_salary", "hourly_salary", "bonus"],
                "bank_account_sync": bank_sync,
            }
        )
        return result
    finally:
        workbook.close()


def _extract_profiles_from_sheet(ws) -> dict[str, dict[str, Any]]:
    profiles: dict[str, dict[str, Any]] = {}
    for header_row in range(1, ws.max_row + 1):
        total_col = _find_label_col(ws, header_row, SUMMARY_MIN_COL, ws.max_column, {"tong gio cong"})
        if not total_col:
            continue
        code_col = _find_label_col(ws, header_row, total_col + 1, min(total_col + 4, ws.max_column), {"ma"})
        if not code_col:
            continue

        result_row = header_row + 5
        note_row = header_row + 6
        code = _employee_code(ws.cell(result_row, code_col).value)
        if not code:
            continue

        name_col = code_col + 1
        salary_cols = _profile_columns(ws, header_row, name_col + 1)
        raw_start_work = _text(ws.cell(note_row, name_col).value)
        start_work_note, note = _split_start_work_and_note(raw_start_work)
        # AI is the dedicated Bắt đầu làm cell in both the legacy and the
        # reformed layout. Keep custom values such as "CMND ..." or
        # "Bắt đầu giữa T5/2026" there instead of misclassifying them as a
        # general comment.
        if raw_start_work and not start_work_note:
            start_work_note, note = raw_start_work, ""
        if note_row <= ws.max_row:
            for col in range(name_col + 1, ws.max_column + 1):
                comment = _text(ws.cell(note_row, col).value)
                if comment and comment not in note.split(" | "):
                    note = " | ".join(part for part in (note, comment) if part)
        profile = {
            "name": _text(ws.cell(result_row, name_col).value),
            "bank_account": _bank_account(ws, header_row, result_row, name_col),
            "start_work_note": start_work_note,
            "note": note,
        }
        for key, col in salary_cols.items():
            profile[key] = _number(ws.cell(result_row, col).value)

        if any(value not in (None, "") for value in profile.values()):
            profiles[code] = profile
    return profiles


def _bank_account(ws, header_row: int, result_row: int, name_col: int) -> str:
    for row in range(header_row + 1, result_row):
        value = ws.cell(row, name_col).value
        if value is None or isinstance(value, bool):
            continue
        if isinstance(value, (int, float)):
            if not float(value).is_integer():
                continue
            text = str(int(value))
        else:
            text = "".join(str(value).split())
        if text.isdigit() and 8 <= len(text) <= 20:
            return text
    return ""


def _profile_columns(ws, header_row: int, start_col: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for col in range(start_col, ws.max_column + 1):
        label = _normalize_label(ws.cell(header_row, col).value)
        if label == "muc luong":
            columns["monthly_salary"] = col
        elif label == "luong 1 ngay cong":
            columns["daily_salary"] = col
        elif label == "luong 1 gio cong":
            columns["hourly_salary"] = col
        elif label == "thuong":
            columns["bonus"] = col
    return columns


def _find_label_col(ws, row: int, start_col: int, end_col: int, labels: set[str]) -> int | None:
    for col in range(start_col, end_col + 1):
        if _normalize_label(ws.cell(row, col).value) in labels:
            return col
    return None


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("đ", "d")
    return " ".join(text.split())


def _text(value: object) -> str:
    return str(value or "").strip()


def _employee_code(value: object) -> str:
    if isinstance(value, bool):
        return ""
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    return normalize_employee_code(value)


def _split_start_work_and_note(value: object) -> tuple[str, str]:
    text = _text(value)
    lower = _normalize_label(text)
    if lower.startswith("bat dau lam "):
        start_work_note = text.split(" ", maxsplit=3)[-1].strip()
        if " | " in start_work_note:
            start_work_note, note = start_work_note.split(" | ", maxsplit=1)
            return start_work_note.strip(), note.strip()
        return start_work_note, ""
    return "", text


def _number(value: object) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip()
    if not text:
        return None
    cleaned = text.replace(",", "").replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _empty_result(source_path: str, reason: str) -> dict[str, Any]:
    return {
        "status": "skipped",
        "reason": reason,
        "source_path": source_path,
        "profile_count": 0,
        "updated_count": 0,
        "skipped_count": 0,
        "updated_codes": [],
        "skipped_codes": [],
        "conflict_count": 0,
        "conflict_codes": [],
        "fields": [],
    }
