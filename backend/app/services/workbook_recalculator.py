from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

from app.services.block_detector import detect_employee_blocks


INTEGER_NUMBER_FORMAT = '#,##0'
NUMBER_FORMAT = '#,##0.##'
TOTAL_HOURS_COL = 32
WORK_DAYS_COL = 40
FONT_NAME = "Arial"


def recalculate_workbook_totals(source_path: Path, output_path: Path) -> dict:
    keep_vba = source_path.suffix.lower() == ".xlsm"
    wb = load_workbook(source_path, data_only=False, keep_vba=keep_vba)
    try:
        ws = _select_attendance_sheet(wb)
        blocks = detect_employee_blocks(ws)
        if not blocks:
            raise ValueError("Khong tim thay block cham cong trong file")

        changed: list[dict] = []
        for block in blocks:
            total_hours = _round_number(_sum_work_hours(ws, block.result_row))
            old_total = _restore_number(ws.cell(row=block.result_row, column=TOTAL_HOURS_COL).value)

            total_cell = ws.cell(row=block.result_row, column=TOTAL_HOURS_COL)
            if not isinstance(total_cell, MergedCell):
                total_cell.value = total_hours
                total_cell.number_format = _number_format_for(total_hours)
                total_cell.font = Font(name=FONT_NAME, bold=True, size=10, color="C00000")

            work_days_cell = ws.cell(row=block.result_row, column=WORK_DAYS_COL)
            if _has_output2_work_days_column(ws, block) and not isinstance(work_days_cell, MergedCell):
                work_days = _round_number(float(total_hours) / 8 if total_hours else 0)
                work_days_cell.value = work_days
                work_days_cell.number_format = _number_format_for(work_days)

            if old_total != total_hours:
                changed.append(
                    {
                        "employee_code": block.employee_code,
                        "old_total_hours": old_total,
                        "new_total_hours": total_hours,
                    }
                )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return {
            "sheet_name": ws.title,
            "employee_count": len(blocks),
            "changed_count": len(changed),
            "changed": changed,
        }
    finally:
        wb.close()


def _select_attendance_sheet(wb):
    best_sheet = None
    best_count = -1
    for ws in wb.worksheets:
        count = sum(1 for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == "Att. Time")
        if count > best_count:
            best_sheet = ws
            best_count = count

    if best_sheet is None or best_count <= 0:
        raise ValueError("Khong tim thay sheet cham cong co dong Att. Time")

    return best_sheet


def _sum_work_hours(ws, row: int) -> float:
    total = 0.0
    for col in range(1, 32):
        value = _restore_number(ws.cell(row=row, column=col).value)
        if isinstance(value, (int, float)):
            total += float(value)
    return total


def _restore_number(value: object) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return _round_number(float(value))

    text = str(value or "").strip()
    if not text or text == "?":
        return None

    normalized = text.replace(" ", "").replace(",", ".")
    try:
        number = float(normalized)
    except ValueError:
        return None
    return _round_number(number)


def _round_number(value: float) -> int | float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _number_format_for(value: object) -> str:
    if isinstance(value, (int, float)) and float(value).is_integer():
        return INTEGER_NUMBER_FORMAT
    return NUMBER_FORMAT


def _has_output2_work_days_column(ws, block) -> bool:
    header_text = str(ws.cell(row=block.header_row, column=WORK_DAYS_COL).value or "").lower()
    return "ng" in header_text or "sá»‘" in header_text or "số" in header_text
