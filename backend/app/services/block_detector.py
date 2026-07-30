from openpyxl.worksheet.worksheet import Worksheet

from app.models.attendance import EmployeeBlock


def detect_employee_blocks(ws: Worksheet) -> list[EmployeeBlock]:
    blocks: list[EmployeeBlock] = []

    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "").strip() != "Att. Time":
            continue

        day_row = row + 1
        employee_row = row + 2
        punch_row = row + 3
        missing_row = row + 4
        late_row = row + 5
        result_row = row + 6

        code_label = str(ws.cell(row=employee_row, column=1).value or "").strip()
        employee_code = str(ws.cell(row=employee_row, column=3).value or "").strip()
        if code_label not in {"Mã:", "MÃ£:"} or not employee_code:
            continue

        blocks.append(
            EmployeeBlock(
                sheet_name=ws.title,
                header_row=row,
                day_row=day_row,
                employee_row=employee_row,
                punch_row=punch_row,
                missing_row=missing_row,
                late_row=late_row,
                result_row=result_row,
                employee_code=employee_code,
            )
        )

    return blocks
