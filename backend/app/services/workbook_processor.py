import shutil
from datetime import date
from pathlib import Path

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.attendance import DayComputation
from app.services.attendance_calculator import calculate_day
from app.services.block_detector import detect_employee_blocks
from app.services.bank_account_store import get_saved_account_number
from app.services.payroll_store import get_payroll_entry
from app.services.period_detector import detect_period_from_sheet
from app.services.punch_parser import parse_punches


YELLOW = "FFFF00"
GREEN = "C4D79B"
RED = "C00000"
WHITE = "FFFFFF"
MONEY_FORMAT = '#,##0'
INTEGER_NUMBER_FORMAT = '#,##0'
NUMBER_FORMAT = '#,##0.##'
FONT_NAME = "Arial"


def analyze_workbook(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    ws = _select_attendance_sheet(wb)
    blocks = detect_employee_blocks(ws)

    rows: list[dict] = []
    manual_checks: list[dict] = []
    result_cells = 0
    missing_cells = 0
    late_cells = 0

    for block in blocks:
        computations = _compute_block(ws, block)
        block_results = []
        for item in computations:
            if item.work_value is not None:
                result_cells += 1
            if item.missing_count is not None:
                missing_cells += 1
            if item.late_minutes is not None:
                late_cells += 1

            if item.manual_checks:
                manual_checks.append(
                    {
                        "employee_code": block.employee_code,
                        "day": item.day,
                        "cell": f"{item.column_letter}{block.punch_row}",
                        "punches": item.punches,
                        "messages": item.manual_checks,
                    }
                )

            block_results.append(
                {
                    "day": item.day,
                    "column": item.column_letter,
                    "punches": item.punches,
                    "work_value": item.work_value,
                    "missing_count": item.missing_count,
                    "late_minutes": item.late_minutes,
                }
            )

        rows.append(
            {
                "employee_code": block.employee_code,
                "header_row": block.header_row,
                "punch_row": block.punch_row,
                "missing_row": block.missing_row,
                "late_row": block.late_row,
                "result_row": block.result_row,
                "results": block_results,
            }
        )

    return {
        "sheet_name": ws.title,
        "period": detect_period_from_sheet(ws),
        "summary": {
            "blocks": len(blocks),
            "result_cells": result_cells,
            "missing_cells": missing_cells,
            "late_cells": late_cells,
            "manual_check_count": len(manual_checks),
        },
        "blocks": rows,
        "manual_checks": manual_checks,
    }


def export_processed_workbook(
    source_path: Path,
    output_path: Path,
    review_overrides: list[dict] | None = None,
    factory: str = "factory1",
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)

    wb = load_workbook(output_path, data_only=False)
    ws = _select_attendance_sheet(wb)
    blocks = detect_employee_blocks(ws)
    overrides = _review_overrides_by_employee_day(review_overrides or [])
    _apply_sunday_column_fills(ws, blocks)

    for block in blocks:
        _format_attendance_time_row(ws, block)
        _write_employee_info_line(ws, block, factory=factory)
        for item in _compute_block(ws, block):
            override = overrides.get((block.employee_code, item.day), {})
            missing_count = override.get("missing_count", item.missing_count)
            late_minutes = override.get("late_minutes", item.late_minutes)
            work_value = override.get("work_value", item.work_value)

            ws.cell(row=block.missing_row, column=item.column).value = (
                missing_count if missing_count is not None and missing_count != "" else None
            )
            ws.cell(row=block.late_row, column=item.column).value = (
                late_minutes if late_minutes is not None and late_minutes != "" else None
            )
            ws.cell(row=block.result_row, column=item.column).value = (
                work_value if work_value is not None and work_value != "" else None
            )
        _write_output1_summary_block(ws, block, factory=factory)

    # Output 1 intentionally stops after the public employee information
    # columns: Tổng giờ, mức phạt NQ/giờ, Mã, Tên/Ghi chú (A:AI).
    _format_title_area(ws, 35)
    _truncate_after_column(ws, 35)
    if blocks:
        ws.print_area = f"A1:AI{max(block.header_row + 7 for block in blocks)}"
    wb.save(output_path)
    return output_path


def _write_output1_summary_block(ws, block, factory: str = "factory1") -> None:
    h = block.header_row
    day_row = block.day_row
    result_row = block.result_row
    note_row = h + 7
    entry = get_payroll_entry(block.employee_code, factory)
    total_hours = _sum_work_hours(ws, result_row)

    headers = {
        32: "Tổng giờ công",
        33: "Mức tiền phạt NQ trên giờ công (đ)",
        34: "Mã",
        35: "Tên nhân viên / Ghi chú",
    }

    _clear_output1_summary_area(ws, h, note_row)
    ws.row_dimensions[day_row].height = max(float(ws.row_dimensions[day_row].height or 0), 28)

    for col in range(32, 36):
        for row in range(day_row, note_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=WHITE if col <= 33 else YELLOW)

    for col, title in headers.items():
        cell = ws.cell(row=day_row, column=col)
        cell.value = title
        cell.fill = PatternFill("solid", fgColor=WHITE if col <= 33 else YELLOW)
        cell.font = Font(name=FONT_NAME, bold=True, size=8, color=RED if col in {32, 33, 34} else "000000")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)

    total_cell = ws.cell(row=result_row, column=32)
    total_cell.value = _round_number(total_hours)
    total_cell.font = Font(name=FONT_NAME, bold=True, size=10, color=RED)
    total_cell.number_format = _number_format_for(total_cell.value)

    penalty_rate_cell = ws.cell(row=result_row, column=33)
    penalty_rate_cell.value = None
    penalty_rate_cell.font = Font(name=FONT_NAME, bold=True, size=10, color=RED)

    code_cell = ws.cell(row=result_row, column=34)
    code_cell.value = block.employee_code
    code_cell.font = Font(name=FONT_NAME, bold=True, size=10, color="000000")

    name_cell = ws.cell(row=result_row, column=35)
    name_cell.value = entry.name
    name_cell.font = Font(name=FONT_NAME, bold=True, size=10)
    name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=True)

    bank_account_cell = ws.cell(row=result_row - 1, column=35)
    bank_account_cell.value = get_saved_account_number(factory, block.employee_code)
    bank_account_cell.number_format = "@"
    bank_account_cell.alignment = Alignment(horizontal="left", vertical="center")

    start_work_note = _format_start_work_note(entry.start_work_note)
    start_work_cell = ws.cell(row=note_row, column=35)
    start_work_cell.value = start_work_note
    start_work_cell.font = Font(name=FONT_NAME, bold=True, size=10)
    start_work_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, shrink_to_fit=True)
    _set_output1_summary_column_widths(ws)


def _format_title_area(ws, end_col: int) -> None:
    title = ws.cell(row=1, column=1).value or "Báo Cáo Dữ Liệu Chấm Công"
    _unmerge_overlapping(ws, 1, 1, 2, end_col)
    _safe_merge(ws, 1, 1, 2, end_col)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = title
    title_cell.font = Font(name=FONT_NAME, size=20, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = max(float(ws.row_dimensions[1].height or 0), 28)
    ws.row_dimensions[2].height = max(float(ws.row_dimensions[2].height or 0), 28)


def _format_attendance_time_row(ws, block) -> None:
    header_row = block.header_row
    _unmerge_overlapping(ws, header_row, 1, header_row, 31)
    tabulation = str(ws.cell(row=header_row, column=10).value or "").strip()
    tabulation_date = str(ws.cell(row=header_row, column=12).value or "").strip()
    if tabulation and tabulation_date and tabulation_date not in tabulation:
        ws.cell(row=header_row, column=10).value = f"{tabulation} {tabulation_date}"
    for col in range(11, 16):
        ws.cell(row=header_row, column=col).value = None

    _safe_merge(ws, header_row, 1, header_row, 2)
    _safe_merge(ws, header_row, 3, header_row, 8)
    _safe_merge(ws, header_row, 10, header_row, 15)
    if all(ws.cell(row=header_row, column=col).value in (None, "") for col in range(16, 32)):
        _safe_merge(ws, header_row, 16, header_row, 31)

    for col in (1, 3, 10, 16):
        cell = ws.cell(row=header_row, column=col)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.font = Font(name=FONT_NAME, size=11)


def _write_employee_info_line(ws, block, factory: str = "factory1") -> None:
    name = get_payroll_entry(block.employee_code, factory).name.strip()
    _safe_merge(ws, block.employee_row, 9, block.employee_row, 15)
    name_cell = ws.cell(row=block.employee_row, column=9)
    name_cell.value = f"Tên: {name}" if name else "Tên:"
    name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    name_cell.fill = PatternFill("solid", fgColor=WHITE)


def _clear_output1_summary_area(ws, start_row: int, end_row: int) -> None:
    _unmerge_overlapping(ws, start_row, 32, end_row, 35)
    for row in range(start_row, end_row + 1):
        for col in range(32, 36):
            cell = ws.cell(row=row, column=col)
            cell.value = None
            cell.border = Border()
            cell.fill = PatternFill(fill_type=None)
            cell.font = Font(name=FONT_NAME, size=9)
            cell.alignment = Alignment()


def _sum_work_hours(ws, row: int) -> float:
    total = 0.0
    for col in range(1, 32):
        value = ws.cell(row=row, column=col).value
        if isinstance(value, (int, float)):
            total += float(value)
    return total


def _round_number(value: float) -> int | float:
    rounded = round(value, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _number_format_for(value: object) -> str:
    if isinstance(value, (int, float)) and float(value).is_integer():
        return INTEGER_NUMBER_FORMAT
    return NUMBER_FORMAT


def _format_start_work_note(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.lower()
    if normalized.startswith("bắt đầu") or normalized.startswith("bat dau"):
        return text
    return f"Bắt đầu làm {text}"


def _safe_merge(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    target = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) == target:
            return
        if _ranges_overlap(merged_range, start_row, start_col, end_row, end_col):
            ws.unmerge_cells(str(merged_range))
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def _unmerge_overlapping(ws, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if _ranges_overlap(merged_range, start_row, start_col, end_row, end_col):
            ws.unmerge_cells(str(merged_range))


def _truncate_after_column(ws, end_col: int) -> None:
    """Physically remove private columns and shrink merges crossing the cut."""
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_col <= end_col:
            continue
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        max_row = merged_range.max_row
        ws.unmerge_cells(str(merged_range))
        if min_col <= end_col:
            ws.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=end_col,
            )
    if ws.max_column > end_col:
        ws.delete_cols(end_col + 1, ws.max_column - end_col)


def _ranges_overlap(merged_range, start_row: int, start_col: int, end_row: int, end_col: int) -> bool:
    return not (
        merged_range.max_row < start_row
        or merged_range.min_row > end_row
        or merged_range.max_col < start_col
        or merged_range.min_col > end_col
    )


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _set_output1_summary_column_widths(ws) -> None:
    widths = {
        "AF": 12,
        "AG": 24,
        "AH": 8,
        "AI": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def _apply_sunday_column_fills(ws, blocks) -> None:
    period = detect_period_from_sheet(ws)
    month = period.get("month")
    year = period.get("year")
    if not isinstance(month, int) or not isinstance(year, int):
        return

    yellow_fill = PatternFill("solid", fgColor=YELLOW)
    white_fill = PatternFill("solid", fgColor=WHITE)
    for block in blocks:
        for col in range(1, 32):
            day_value = ws.cell(row=block.day_row, column=col).value
            if not _is_sunday(year, month, day_value):
                continue
            header_cell = ws.cell(row=block.header_row, column=col)
            if not isinstance(header_cell, MergedCell):
                header_cell.fill = white_fill
            for row in range(block.day_row, block.header_row + 8):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                cell.fill = yellow_fill
        _clear_employee_text_fills(ws, block, white_fill)


def _clear_employee_text_fills(ws, block, fill: PatternFill) -> None:
    for col in (9, 19):
        cell = ws.cell(row=block.employee_row, column=col)
        if not isinstance(cell, MergedCell):
            cell.fill = fill


def _is_sunday(year: int, month: int, day_value: object) -> bool:
    if not isinstance(day_value, int):
        return False
    try:
        return date(year, month, day_value).weekday() == 6
    except ValueError:
        return False


def _review_overrides_by_employee_day(items: list[dict]) -> dict[tuple[str, int], dict]:
    result: dict[tuple[str, int], dict] = {}
    for item in items:
        employee_code = str(item.get("employee_code", "")).strip()
        day = item.get("day")
        if not employee_code or not isinstance(day, int):
            continue

        target = result.setdefault((employee_code, day), {})
        if "missing_count" in item:
            target["missing_count"] = item.get("missing_count")
        if "late_minutes" in item:
            target["late_minutes"] = item.get("late_minutes")
        if "work_value" in item:
            target["work_value"] = item.get("work_value")
    return result


def _select_attendance_sheet(wb):
    best_sheet = None
    best_count = -1
    for ws in wb.worksheets:
        count = sum(1 for row in range(1, ws.max_row + 1) if ws.cell(row=row, column=1).value == "Att. Time")
        if count > best_count:
            best_sheet = ws
            best_count = count

    if best_sheet is None or best_count <= 0:
        raise ValueError("Không tìm thấy sheet chấm công có dòng Att. Time")

    return best_sheet


def _compute_block(ws, block) -> list[DayComputation]:
    computations: list[DayComputation] = []

    for column in range(1, 32):
        day_value = ws.cell(row=block.day_row, column=column).value
        if not isinstance(day_value, int):
            continue

        raw_value = ws.cell(row=block.punch_row, column=column).value
        punches = parse_punches(raw_value)
        if not punches:
            continue

        calculated = calculate_day(punches)
        computations.append(
            DayComputation(
                day=day_value,
                column=column,
                column_letter=get_column_letter(column),
                raw_value=raw_value,
                punches=punches,
                work_value=calculated.work_value,
                missing_count=calculated.missing_count,
                late_minutes=calculated.late_minutes,
                manual_checks=calculated.manual_checks,
            )
        )

    return computations
