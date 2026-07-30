from __future__ import annotations

import hashlib
import unicodedata
from dataclasses import asdict, dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Literal

from openpyxl import load_workbook

from app.services.block_detector import detect_employee_blocks
from app.services.data_mapper import (
    _owner_records_by_code,
    _prepare_mapping_source_workbook,
    _summary_blocks_by_code,
)
from app.services.factory2_workbook import analyze_factory2_workbook
from app.services.period_detector import detect_period_from_sheet
from app.services.workbook_normalizer import _employee_code_rows


WorkbookRole = Literal[
    "analysis",
    "mapping_current",
    "mapping_previous",
    "final_copy",
    "recalculate_output1",
    "recalculate_output2",
]

OUTPUT1_LABELS = {"tong gio cong", "ma"}
PRIVATE_LABELS = {
    "muc luong",
    "luong 1 ngay cong",
    "luong 1 gio cong",
    "thuong",
    "ung luong",
    "ung luong + phat",
    "phat nq",
    "muc tien phat nq tren gio cong",
    "muc tien phat nq tren gio cong (d)",
}


@dataclass(frozen=True)
class WorkbookProfile:
    detected_kind: str
    sheet_name: str
    month: int | None
    year: int | None
    employee_count: int
    employee_codes: tuple[str, ...]
    summary_count: int
    has_private_payroll: bool
    structural_errors: tuple[str, ...]
    ignored_trailing_style_rows: int = 0

    @property
    def period_label(self) -> str:
        if self.month and self.year:
            return f"{self.month:02d}/{self.year}"
        return ""

    def to_dict(self) -> dict:
        result = asdict(self)
        result["period_label"] = self.period_label
        result["accepted"] = not self.structural_errors
        return result


def inspect_workbook_for_role(
    path: Path,
    role: WorkbookRole,
    factory: str = "factory1",
) -> WorkbookProfile:
    profile = profile_workbook(path, factory=factory)
    validate_profile_for_role(profile, role)
    return profile


def profile_workbook(path: Path, factory: str = "factory1") -> WorkbookProfile:
    with TemporaryDirectory(prefix="attendance-guard-") as temp_dir:
        prepared_path, ignored_rows = _prepare_mapping_source_workbook(path, Path(temp_dir))
        workbook = load_workbook(prepared_path, data_only=False)
        try:
            worksheet = _select_attendance_sheet(workbook)
            if worksheet is None:
                return _profile_factory2_or_unknown(path, factory)

            attendance_rows = [
                row
                for row in range(1, worksheet.max_row + 1)
                if str(worksheet.cell(row, 1).value or "").strip() == "Att. Time"
            ]
            blocks = detect_employee_blocks(worksheet)
            raw_code_rows = _employee_code_rows(worksheet)
            errors: list[str] = []
            try:
                summaries = _summary_blocks_by_code(worksheet)
            except ValueError as exc:
                summaries = {}
                errors.append(str(exc))

            owner_records = {}
            semantic_error = None
            try:
                owner_records = _owner_records_by_code(worksheet, worksheet)
            except ValueError as exc:
                semantic_error = str(exc)

            labels = _all_normalized_labels(worksheet)
            has_private = any(
                label in PRIVATE_LABELS or label.startswith("luong thang")
                for label in labels
            )

            summary_rows = sorted(block.header_row for block in summaries.values())
            if summaries and not has_private:
                errors.extend(_spacing_errors(attendance_rows, "dòng Att. Time"))
                errors.extend(_spacing_errors(summary_rows, "khung Tổng giờ công"))

            if summaries and not has_private and attendance_rows and len(blocks) != len(attendance_rows):
                errors.append(
                    f"Có {len(attendance_rows)} dòng Att. Time nhưng chỉ nhận diện được "
                    f"{len(blocks)} khung nhân viên hoàn chỉnh."
                )
            if summaries and not has_private and len(summaries) != len(blocks):
                errors.append(
                    f"Có {len(blocks)} khung chấm công nhưng chỉ có {len(summaries)} "
                    "khung Tổng giờ công/Mã đúng vị trí."
                )

            block_codes = tuple(str(block.employee_code).strip() for block in blocks)
            summary_codes = tuple(summaries.keys())
            if summaries and not has_private and set(block_codes) != set(summary_codes):
                missing_summary = sorted(set(block_codes) - set(summary_codes))
                extra_summary = sorted(set(summary_codes) - set(block_codes))
                detail = []
                if missing_summary:
                    detail.append(f"thiếu khung tổng cho mã {', '.join(missing_summary[:8])}")
                if extra_summary:
                    detail.append(f"khung tổng lệch mã {', '.join(extra_summary[:8])}")
                errors.append("Mã nhân viên giữa phần chấm công và phần tổng không khớp: " + "; ".join(detail))

            if summaries and not has_private:
                attendance_row_set = set(attendance_rows)
                for summary in summaries.values():
                    if summary.header_row - 1 not in attendance_row_set:
                        errors.append(
                            f"Khung mã {summary.code} tại dòng {summary.header_row} bị lệch: "
                            "không có dòng Att. Time ngay phía trên."
                        )
                        break

            if has_private and owner_records:
                errors = [error for error in errors if "mã nhân viên bị trùng" in error.lower()]
                if semantic_error:
                    errors.append(semantic_error)
                detected_kind = "output2"
                employee_codes = tuple(owner_records.keys())
                summary_count = len(owner_records)
            elif has_private:
                detected_kind = "output2"
                employee_codes = ()
                summary_count = 0
                errors.append(
                    semantic_error
                    or "Có cột lương nhưng không nhận diện được hồ sơ nhân viên theo Mã và tiêu đề."
                )
            elif summaries:
                detected_kind = "output1"
                employee_codes = block_codes
                summary_count = len(summaries)
            else:
                detected_kind = "raw"
                employee_codes = tuple(
                    str(worksheet.cell(row, 3).value or "").strip()
                    for row in raw_code_rows
                    if str(worksheet.cell(row, 3).value or "").strip()
                )
                summary_count = 0

            period = detect_period_from_sheet(worksheet)
            return WorkbookProfile(
                detected_kind=detected_kind,
                sheet_name=worksheet.title,
                month=period.get("month") if isinstance(period.get("month"), int) else None,
                year=period.get("year") if isinstance(period.get("year"), int) else None,
                employee_count=len(employee_codes),
                employee_codes=employee_codes,
                summary_count=summary_count,
                has_private_payroll=has_private,
                structural_errors=tuple(dict.fromkeys(errors)),
                ignored_trailing_style_rows=ignored_rows,
            )
        finally:
            workbook.close()


def validate_profile_for_role(profile: WorkbookProfile, role: WorkbookRole) -> None:
    if profile.structural_errors:
        raise ValueError(
            "Cấu trúc bảng không an toàn để xử lý. "
            + " ".join(profile.structural_errors)
            + " Hãy xuất lại file đúng mẫu, không tự xóa/chèn dòng giữa các nhân viên."
        )
    if not profile.employee_count:
        raise ValueError("Không nhận diện được mã nhân viên trong file.")

    if role == "analysis":
        if profile.detected_kind in {"output1", "output2"}:
            label = "Output 2/bảng chính thức" if profile.detected_kind == "output2" else "Output 1 đã tính"
            raise ValueError(
                f"File này có vẻ là {label}, không phải bảng chấm công đầu vào. "
                "Hãy chọn file gốc từ máy chấm công."
            )
        if profile.detected_kind not in {"raw", "factory2_raw"}:
            raise ValueError("File không giống bảng chấm công đầu vào của xưởng đã chọn.")
        return

    expected = {
        "mapping_current": "output1",
        "mapping_previous": "output2",
        "final_copy": "output2",
        "recalculate_output1": "output1",
        "recalculate_output2": "output2",
    }[role]
    if profile.detected_kind != expected:
        if role == "mapping_current" and profile.detected_kind == "output2":
            raise ValueError(
                "Ô Output 1 tháng mới đang chứa bảng chính thức/Output 2 có dữ liệu lương. "
                "Có thể bạn đã chọn nhầm hoặc đảo hai file."
            )
        if role == "mapping_previous" and profile.detected_kind == "output1":
            raise ValueError(
                "Ô Bảng chính thức tháng cũ đang chứa Output 1 chưa có dữ liệu lương. "
                "Có thể bạn đã chọn nhầm hoặc đảo hai file."
            )
        descriptions = {
            "mapping_current": "Output 1 tháng mới",
            "mapping_previous": "bảng chính thức/Output 2 tháng cũ",
            "final_copy": "bảng chính thức/Output 2 đã chốt",
            "recalculate_output1": "Output 1",
            "recalculate_output2": "Output 2",
        }
        raise ValueError(
            f"File không đúng loại cho mục này. Cần {descriptions[role]}, "
            f"nhưng hệ thống nhận diện là {_kind_label(profile.detected_kind)}."
        )


def validate_mapping_pair(
    current_path: Path,
    previous_path: Path,
    factory: str = "factory1",
) -> tuple[WorkbookProfile, WorkbookProfile]:
    if _sha256(current_path) == _sha256(previous_path):
        raise ValueError("Hai ô đang chứa cùng một file. Hãy chọn Output 1 tháng mới và bảng chính thức tháng cũ khác nhau.")

    current = profile_workbook(current_path, factory=factory)
    previous = profile_workbook(previous_path, factory=factory)

    if current.detected_kind == "output2" and previous.detected_kind == "output1":
        raise ValueError(
            "Có vẻ hai file đã bị chọn ngược: ô Output 1 tháng mới đang là bảng chính thức, "
            "còn ô Bảng chính thức tháng cũ đang là Output 1."
        )
    validate_profile_for_role(current, "mapping_current")
    validate_profile_for_role(previous, "mapping_previous")

    if not current.month or not current.year or not previous.month or not previous.year:
        raise ValueError("Không nhận diện đủ tháng/năm của hai file nên hệ thống dừng để tránh gán nhầm kỳ.")
    if (previous.year, previous.month) >= (current.year, current.month):
        raise ValueError(
            f"Bảng chính thức phải thuộc kỳ trước Output 1 tháng mới. "
            f"Hiện đang chọn tháng mới {current.period_label} và bảng cũ {previous.period_label}."
        )

    current_codes = set(current.employee_codes)
    previous_codes = set(previous.employee_codes)
    matched = current_codes & previous_codes
    minimum_match = 1 if len(current_codes) < 5 else min(10, max(3, int(len(current_codes) * 0.2)))
    if len(matched) < minimum_match:
        raise ValueError(
            f"Chỉ khớp {len(matched)}/{len(current_codes)} mã nhân viên giữa hai file. "
            "Có thể chọn nhầm xưởng, nhầm kỳ hoặc nhầm loại bảng."
        )
    return current, previous


def ensure_period_matches(profile: WorkbookProfile, month: int, year: int, label: str) -> None:
    if not profile.month or not profile.year:
        raise ValueError(f"Không nhận diện được tháng/năm trong {label}; hệ thống không lưu để tránh sai thư mục.")
    if (profile.month, profile.year) != (month, year):
        raise ValueError(
            f"{label} thuộc kỳ {profile.period_label}, nhưng ô lưu đang chọn {month:02d}/{year}. "
            "Hãy chọn đúng tháng/năm trước khi lưu."
        )


def _profile_factory2_or_unknown(path: Path, factory: str) -> WorkbookProfile:
    if factory == "factory2":
        try:
            analysis = analyze_factory2_workbook(path)
            period = analysis.get("period") or {}
            blocks = analysis.get("blocks") or []
            codes = tuple(str(block.get("employee_code") or "").strip() for block in blocks)
            return WorkbookProfile(
                detected_kind="factory2_raw",
                sheet_name=str(analysis.get("sheet_name") or ""),
                month=period.get("month") if isinstance(period.get("month"), int) else None,
                year=period.get("year") if isinstance(period.get("year"), int) else None,
                employee_count=len(codes),
                employee_codes=codes,
                summary_count=0,
                has_private_payroll=False,
                structural_errors=(),
            )
        except Exception:
            pass
    return WorkbookProfile(
        detected_kind="unknown",
        sheet_name="",
        month=None,
        year=None,
        employee_count=0,
        employee_codes=(),
        summary_count=0,
        has_private_payroll=False,
        structural_errors=("Không tìm thấy sheet chấm công phù hợp với xưởng đã chọn.",),
    )


def _select_attendance_sheet(workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        count = sum(
            1
            for row in range(1, worksheet.max_row + 1)
            if str(worksheet.cell(row, 1).value or "").strip() == "Att. Time"
        )
        if count:
            candidates.append((count, worksheet))
    return max(candidates, key=lambda item: item[0])[1] if candidates else None


def _spacing_errors(rows: list[int], label: str) -> list[str]:
    if len(rows) < 2:
        return []
    bad = [(left, right, right - left) for left, right in zip(rows, rows[1:]) if right - left != 8]
    if not bad:
        return []
    left, right, distance = bad[0]
    return [
        f"{label} bị lệch giữa dòng {left} và {right} (cách {distance} dòng thay vì 8). "
        "Có thể đã thiếu hoặc thừa dòng ngăn giữa hai nhân viên."
    ]


def _summary_header_labels(worksheet, summaries: dict) -> set[str]:
    labels: set[str] = set()
    for block in summaries.values():
        for col in range(block.total_col, block.data_end_col + 1):
            labels.add(_normalize_label(worksheet.cell(block.header_row, col).value))
    return labels


def _all_normalized_labels(worksheet) -> set[str]:
    labels: set[str] = set()
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip():
                labels.add(_normalize_label(cell.value))
    return labels


def _normalize_label(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("đ", "d")
    return " ".join(text.split())


def _kind_label(kind: str) -> str:
    return {
        "raw": "bảng chấm công đầu vào",
        "factory2_raw": "bảng chấm công đầu vào Xưởng 2",
        "output1": "Output 1",
        "output2": "Output 2/bảng chính thức",
        "unknown": "bảng không xác định",
    }.get(kind, kind)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
