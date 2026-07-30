import re
from pathlib import Path

from openpyxl import load_workbook


PERIOD_PATTERN = re.compile(r"(?P<year>20\d{2})[-/.](?P<month>\d{1,2})[-/.]\d{1,2}")
MONTH_YEAR_PATTERN = re.compile(r"(?:th[aá]ng\s*)?(?P<month>\d{1,2})\s*[/.-]\s*(?P<year>20\d{2})", re.IGNORECASE)


def detect_period_from_workbook(path: Path) -> dict[str, int | str | None]:
    wb = load_workbook(path, read_only=True, data_only=False)
    try:
        for ws in wb.worksheets:
            period = detect_period_from_sheet(ws)
            if period["month"] and period["year"]:
                return period
    finally:
        wb.close()

    return {"month": None, "year": None, "label": ""}


def detect_period_from_sheet(ws) -> dict[str, int | str | None]:
    for row in range(1, ws.max_row + 1):
        first_value = str(ws.cell(row=row, column=1).value or "")
        if "Att. Time" not in first_value:
            continue

        candidates = [str(ws.cell(row=row, column=col).value or "") for col in range(1, min(ws.max_column, 8) + 1)]
        period = detect_period_from_text(" ".join(candidates))
        if period["month"] and period["year"]:
            return period

    # Output files from different Excel versions can move or unmerge the
    # "Att. Time" label. Search the visible header area by content instead of
    # relying only on one fixed cell.
    candidates = []
    for row in range(1, min(ws.max_row, 40) + 1):
        for col in range(1, min(ws.max_column, 50) + 1):
            value = ws.cell(row=row, column=col).value
            if value not in (None, ""):
                candidates.append(str(value))
    period = detect_period_from_text(" ".join(candidates))
    if period["month"] and period["year"]:
        return period

    return {"month": None, "year": None, "label": ""}


def detect_period_from_text(text: str) -> dict[str, int | str | None]:
    match = PERIOD_PATTERN.search(text)
    if not match:
        reverse_match = MONTH_YEAR_PATTERN.search(text)
        if not reverse_match:
            return {"month": None, "year": None, "label": ""}
        match = reverse_match

    year = int(match.group("year"))
    month = int(match.group("month"))
    return {"month": month, "year": year, "label": f"{month:02d}/{year}"}
