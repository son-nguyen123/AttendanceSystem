from copy import copy
from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.block_detector import detect_employee_blocks
from app.services.punch_parser import parse_punches


YELLOW = "FFFF00"
WHITE = "FFFFFF"


@dataclass(frozen=True)
class WorkbookLayoutInspection:
    requires_normalization: bool
    raw_employee_count: int
    retained_employee_count: int
    discarded_empty_employee_count: int
    detected_block_count: int
    sheet_name: str
    missing_output1_summary: bool


@dataclass(frozen=True)
class RawEmployeeBlock:
    employee_row: int
    punch_rows: list[int]


def inspect_workbook_layout(path: Path) -> WorkbookLayoutInspection:
    wb = load_workbook(path, data_only=False)
    try:
        ws = _select_attendance_sheet(wb)
        code_rows = _employee_code_rows(ws)
        raw_blocks = _raw_employee_blocks_with_punches(ws, code_rows)
        blocks = detect_employee_blocks(ws)
        return WorkbookLayoutInspection(
            requires_normalization=_requires_normalization(ws, code_rows, raw_blocks, blocks),
            raw_employee_count=len(code_rows),
            retained_employee_count=len(raw_blocks),
            discarded_empty_employee_count=len(code_rows) - len(raw_blocks),
            detected_block_count=len(blocks),
            sheet_name=ws.title,
            missing_output1_summary=_missing_output1_summary(ws, blocks),
        )
    finally:
        wb.close()


def normalize_raw_attendance_workbook(source_path: Path, output_path: Path) -> Path:
    source_wb = load_workbook(source_path, data_only=False)
    ws = _select_attendance_sheet(source_wb)
    code_rows = _employee_code_rows(ws)
    raw_blocks = _raw_employee_blocks_with_punches(ws, code_rows)
    if not raw_blocks:
        raise ValueError("Khong tim thay dong ma nhan vien co du lieu cham cong")

    header_row = _first_attendance_header_row(ws)
    day_row = header_row + 1
    sunday_columns = _sunday_columns_from_period(ws.cell(row=header_row, column=3).value)
    target_wb = Workbook()
    target_ws = target_wb.active
    target_ws.title = ws.title

    _setup_clean_sheet(target_ws, ws)
    for index, block in enumerate(raw_blocks):
        target_header_row = 3 + index * 8
        _write_clean_block(target_ws, ws, header_row, day_row, block, target_header_row, sunday_columns)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(output_path)
    target_wb.close()
    source_wb.close()
    return output_path


def _requires_normalization(
    ws: Worksheet,
    code_rows: list[int],
    raw_blocks: list[RawEmployeeBlock],
    blocks: list,
) -> bool:
    if not code_rows:
        return False
    if len(raw_blocks) < len(code_rows):
        return True
    if len(blocks) < len(raw_blocks):
        return True

    code_row_set = set(code_rows)
    for block in blocks:
        if block.missing_row in code_row_set or block.late_row in code_row_set or block.result_row in code_row_set:
            return True
    return False


def _missing_output1_summary(ws: Worksheet, blocks: list) -> bool:
    if not blocks:
        return False

    for block in blocks:
        headers = {
            32: _plain_text(ws.cell(row=block.day_row, column=32).value),
            33: _plain_text(ws.cell(row=block.day_row, column=33).value),
            34: _plain_text(ws.cell(row=block.day_row, column=34).value),
        }
        has_summary = (
            headers[32] == "tong gio cong"
            and headers[33] == "ma"
            and "ten nhan vien" in headers[34]
            and "ghi chu" in headers[34]
        )
        if not has_summary:
            return True
    return False


def _plain_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return " ".join(text.split())


def _select_attendance_sheet(wb):
    best_sheet = None
    best_count = -1
    for ws in wb.worksheets:
        count = sum(1 for row in range(1, ws.max_row + 1) if _is_attendance_header(ws.cell(row=row, column=1).value))
        if count > best_count:
            best_sheet = ws
            best_count = count

    if best_sheet is None or best_count <= 0:
        raise ValueError("Khong tim thay sheet cham cong co dong Att. Time")

    return best_sheet


def _first_attendance_header_row(ws: Worksheet) -> int:
    for row in range(1, ws.max_row + 1):
        if _is_attendance_header(ws.cell(row=row, column=1).value):
            return row
    raise ValueError("Khong tim thay dong Att. Time")


def _employee_code_rows(ws: Worksheet) -> list[int]:
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        if _is_employee_info_row(ws, row):
            rows.append(row)
    return rows


def _is_attendance_header(value: object) -> bool:
    return str(value or "").strip() == "Att. Time"


def _is_code_label(value: object) -> bool:
    text = str(value or "").strip()
    return len(text) >= 2 and text[0].lower() == "m" and text.endswith(":")


def _is_employee_info_row(ws: Worksheet, row: int) -> bool:
    employee_code = str(ws.cell(row=row, column=3).value or "").strip()
    if not employee_code:
        return False

    name_label = _plain_text(ws.cell(row=row, column=9).value)
    department_label = _plain_text(ws.cell(row=row, column=19).value)
    if name_label.startswith("ten") and department_label.startswith("phong ban"):
        return True

    return _is_code_label(ws.cell(row=row, column=1).value)


def _raw_employee_blocks_with_punches(ws: Worksheet, code_rows: list[int]) -> list[RawEmployeeBlock]:
    blocks: list[RawEmployeeBlock] = []
    for index, employee_row in enumerate(code_rows):
        next_employee_row = code_rows[index + 1] if index + 1 < len(code_rows) else ws.max_row + 1
        punch_rows = [
            row
            for row in range(employee_row + 1, next_employee_row)
            if _row_has_punches(ws, row)
        ]
        if punch_rows:
            blocks.append(RawEmployeeBlock(employee_row=employee_row, punch_rows=punch_rows))
    return blocks


def _row_has_punches(ws: Worksheet, row: int) -> bool:
    max_col = min(ws.max_column, 31)
    return any(parse_punches(ws.cell(row=row, column=col).value) for col in range(1, max_col + 1))


def _setup_clean_sheet(target_ws: Worksheet, source_ws: Worksheet) -> None:
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.freeze_panes = None
    for col in range(1, 32):
        letter = get_column_letter(col)
        target_ws.column_dimensions[letter].width = _template_column_width(col)

    target_ws.row_dimensions[1].height = 28
    target_ws.row_dimensions[2].height = 28
    _safe_merge(target_ws, 1, 1, 2, 31)
    title_cell = target_ws.cell(row=1, column=1)
    title_cell.value = source_ws.cell(row=1, column=1).value or "Bao Cao Du Lieu Cham Cong"
    title_cell.font = Font(name="Arial", size=20, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _template_column_width(col: int) -> float:
    return 6.2


def _write_clean_block(
    target_ws: Worksheet,
    source_ws: Worksheet,
    source_header_row: int,
    source_day_row: int,
    source_block: RawEmployeeBlock,
    target_header_row: int,
    sunday_columns: set[int],
) -> None:
    target_day_row = target_header_row + 1
    target_employee_row = target_header_row + 2
    target_punch_row = target_header_row + 3
    target_missing_row = target_header_row + 4
    target_late_row = target_header_row + 5
    target_result_row = target_header_row + 6
    target_blank_row = target_header_row + 7
    source_employee_row = source_block.employee_row

    target_ws.row_dimensions[target_header_row].height = 14.5
    target_ws.row_dimensions[target_day_row].height = 18
    target_ws.row_dimensions[target_employee_row].height = 18
    target_ws.row_dimensions[target_punch_row].height = _punch_row_height(source_ws, source_block.punch_rows)
    for row in (target_missing_row, target_late_row, target_result_row, target_blank_row):
        target_ws.row_dimensions[row].height = 14.5

    for col in range(1, 32):
        is_sunday = col in sunday_columns
        _style_grid_cell(target_ws.cell(row=target_header_row, column=col), "header", False)
        _style_grid_cell(target_ws.cell(row=target_day_row, column=col), "day", is_sunday)
        _style_grid_cell(target_ws.cell(row=target_employee_row, column=col), "employee", is_sunday)
        _style_grid_cell(target_ws.cell(row=target_punch_row, column=col), "punch", is_sunday)
        _style_grid_cell(target_ws.cell(row=target_missing_row, column=col), "entry", is_sunday)
        _style_grid_cell(target_ws.cell(row=target_late_row, column=col), "entry", is_sunday)
        _style_grid_cell(target_ws.cell(row=target_result_row, column=col), "entry", is_sunday)
        _style_grid_cell(target_ws.cell(row=target_blank_row, column=col), "entry", is_sunday)

        target_ws.cell(row=target_day_row, column=col).value = source_ws.cell(row=source_day_row, column=col).value
        target_ws.cell(row=target_employee_row, column=col).value = source_ws.cell(row=source_employee_row, column=col).value
        target_ws.cell(row=target_punch_row, column=col).value = _format_punch_cell(
            _combined_punch_cell_value(source_ws, source_block.punch_rows, col)
        )
    target_ws.cell(row=target_employee_row, column=1).value = "Mã:"

    for col in range(1, 32):
        value = source_ws.cell(row=source_header_row, column=col).value
        if value is not None:
            target_ws.cell(row=target_header_row, column=col).value = value

    _apply_template_merges(target_ws, target_header_row)


def _apply_template_merges(ws: Worksheet, header_row: int) -> None:
    employee_row = header_row + 2
    tabulation = str(ws.cell(row=header_row, column=10).value or "").strip()
    tabulation_date = str(ws.cell(row=header_row, column=12).value or "").strip()
    department_label = str(ws.cell(row=employee_row, column=19).value or "").strip()
    department_name = str(ws.cell(row=employee_row, column=21).value or "").strip()

    if tabulation and tabulation_date:
        ws.cell(row=header_row, column=10).value = f"{tabulation} {tabulation_date}"
        ws.cell(row=header_row, column=12).value = None
    if department_label and department_name:
        ws.cell(row=employee_row, column=19).value = f"{department_label} {department_name}"
        ws.cell(row=employee_row, column=21).value = None

    _safe_merge(ws, header_row, 1, header_row, 2)
    _safe_merge(ws, header_row, 3, header_row, 8)
    _safe_merge(ws, header_row, 10, header_row, 15)
    if all(ws.cell(row=header_row, column=col).value in (None, "") for col in range(16, 32)):
        _safe_merge(ws, header_row, 16, header_row, 31)
    _safe_merge(ws, employee_row, 1, employee_row, 2)
    _safe_merge(ws, employee_row, 3, employee_row, 8)
    _safe_merge(ws, employee_row, 9, employee_row, 15)
    _safe_merge(ws, employee_row, 19, employee_row, 22)


def _format_punch_cell(value: object) -> str | None:
    punches = parse_punches(value)
    return "\n".join(punches) if punches else None


def _combined_punch_cell_value(ws: Worksheet, rows: list[int], col: int) -> str:
    values: list[str] = []
    for row in rows:
        value = ws.cell(row=row, column=col).value
        if parse_punches(value):
            values.append(str(value))
    return "\n".join(values)


def _style_grid_cell(cell, kind: str, is_sunday: bool = False) -> None:
    cell.border = _thin_border()
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor=YELLOW if is_sunday else WHITE)

    if kind == "header":
        cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    elif kind == "day":
        cell.font = Font(name="Arial", size=12)
    elif kind == "employee":
        cell.font = Font(name="Arial", size=12)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    elif kind == "punch":
        cell.font = Font(name="Arial", size=8)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.font = Font(name="Arial", size=11)


def _punch_row_height(ws: Worksheet, rows: list[int]) -> float:
    source_height = max((float(ws.row_dimensions[row].height or 90) for row in rows), default=90)
    return min(max(float(source_height), 100), 150)


def _sunday_columns_from_period(value: object) -> set[int]:
    match = re.search(r"(20\d{2})-(\d{1,2})-(\d{1,2})", str(value or ""))
    if not match:
        return set()

    year = int(match.group(1))
    month = int(match.group(2))
    result: set[int] = set()
    for day in range(1, 32):
        try:
            current = date(year, month, day)
        except ValueError:
            continue
        if current.weekday() == 6:
            result.add(day)
    return result


def _thin_border() -> Border:
    side = Side(style="thin", color="000000")
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_merge(ws: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if not (
            merged_range.max_row < start_row
            or merged_range.min_row > end_row
            or merged_range.max_col < start_col
            or merged_range.min_col > end_col
        ):
            ws.unmerge_cells(str(merged_range))
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)


def _copy_row(ws: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    _copy_row_format(ws, source_row, target_row, max_col)
    for col in range(1, max_col + 1):
        ws.cell(row=target_row, column=col).value = ws.cell(row=source_row, column=col).value
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _copy_row_format(ws: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=target_row, column=col)
        if source.has_style:
            target._style = copy(source._style)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _clear_row_values(ws: Worksheet, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).value = None


def _copy_single_row_merges(ws: Worksheet, source_row: int, target_row: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row != source_row or merged_range.max_row != source_row:
            continue

        target = (
            target_row,
            merged_range.min_col,
            target_row,
            merged_range.max_col,
        )
        if _has_merge(ws, *target):
            continue
        ws.merge_cells(start_row=target[0], start_column=target[1], end_row=target[2], end_column=target[3])


def _unmerge_attendance_area(ws: Worksheet, start_row: int, max_col: int) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.max_row < start_row or merged_range.min_col > max_col:
            continue
        ws.unmerge_cells(str(merged_range))


def _has_merge(ws: Worksheet, start_row: int, start_col: int, end_row: int, end_col: int) -> bool:
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row == start_row
            and merged_range.min_col == start_col
            and merged_range.max_row == end_row
            and merged_range.max_col == end_col
        ):
            return True
    return False
