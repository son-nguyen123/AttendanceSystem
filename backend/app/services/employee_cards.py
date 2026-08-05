import calendar
import html
import json
import os
import re
import subprocess
import zipfile
from datetime import date
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from app.services.block_detector import detect_employee_blocks
from app.services.owner_profile_sync import sync_latest_final_copy_profile
from app.services.payroll_workbook import export_payroll_workbook, preview_payroll
from app.services.period_detector import detect_period_from_sheet
from app.services.workbook_processor import analyze_workbook, export_processed_workbook


ROOT_DIR = Path(__file__).resolve().parents[3]
EXCEL_RANGE_EXPORT_SCRIPT = ROOT_DIR / "scripts" / "export-excel-ranges.ps1"


def export_employee_cards_zip(
    source_path: Path,
    output_path: Path,
    kind: str,
    review_overrides: list[dict] | None = None,
    factory: str = "factory1",
) -> Path:
    if kind not in {"output1", "output2"}:
        raise ValueError("Loại phiếu không hợp lệ")

    overrides = review_overrides or []
    profile_sync = sync_latest_final_copy_profile(source_path, factory)
    profile_codes = {
        str(code).strip()
        for code in profile_sync.get("profile_codes", [])
        if str(code).strip()
    } if profile_sync.get("status") == "ok" else set()
    with TemporaryDirectory(prefix="attendance-excel-screenshots-") as temporary_dir:
        workbook_path = Path(temporary_dir) / f"employee_cards_{kind}.xlsx"
        if kind == "output2":
            export_payroll_workbook(
                source_path,
                workbook_path,
                review_overrides=overrides,
                include_saved_data=True,
                profile_codes=profile_codes,
                factory=factory,
            )
        else:
            export_processed_workbook(
                source_path,
                workbook_path,
                review_overrides=overrides,
                factory=factory,
            )
        return export_employee_screenshots_from_workbook(workbook_path, output_path, kind)


def export_employee_screenshots_from_workbook(
    workbook_path: Path,
    output_path: Path,
    kind: str,
) -> Path:
    """Capture each employee block exactly as Excel renders it and zip the PNGs."""
    if kind not in {"output1", "output2"}:
        raise ValueError("Loại ảnh bảng công không hợp lệ")
    if not EXCEL_RANGE_EXPORT_SCRIPT.exists():
        raise FileNotFoundError("Thiếu công cụ chụp vùng Excel")

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)
    try:
        worksheet, blocks = _employee_screenshot_sheet_and_blocks(workbook)
        if not blocks:
            raise ValueError("Không tìm thấy khung bảng công nhân viên trong file Excel")
        period = detect_period_from_sheet(worksheet)
        period_label = _period_label(period)
        final_column = 44 if kind == "output2" else 35
        name_column = 35
        jobs: list[dict[str, str]] = []
        used_filenames: set[str] = set()
        for index, block in enumerate(blocks, start=1):
            employee_code = str(block.employee_code)
            name = str(worksheet.cell(block.result_row, name_column).value or "").strip()
            filename = _unique_card_filename(
                _card_filename(employee_code, {"name": name}, kind, period_label),
                used_filenames,
                index,
            )
            jobs.append(
                {
                    "sheet": worksheet.title,
                    "range": f"A{block.header_row}:{get_column_letter(final_column)}{block.header_row + 7}",
                    "filename": filename,
                }
            )
    finally:
        workbook.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with TemporaryDirectory(prefix="attendance-excel-png-") as image_dir:
        image_path = Path(image_dir)
        jobs_path = image_path / "jobs.json"
        jobs_path.write_text(json.dumps(jobs, ensure_ascii=False, indent=2), encoding="utf-8")
        _run_excel_range_export(workbook_path, jobs_path, image_path, len(jobs))
        with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for job in jobs:
                png_path = image_path / job["filename"]
                if not png_path.exists() or png_path.stat().st_size <= 0:
                    raise ValueError(f"Excel chưa tạo được ảnh {job['filename']}")
                archive.write(png_path, arcname=job["filename"])
    return output_path


def _employee_screenshot_sheet_and_blocks(workbook):
    best_sheet = None
    best_blocks = []
    for worksheet in workbook.worksheets:
        blocks = detect_employee_blocks(worksheet)
        if len(blocks) > len(best_blocks):
            best_sheet = worksheet
            best_blocks = blocks
    if best_sheet is None:
        best_sheet = workbook.active
    return best_sheet, best_blocks


def _run_excel_range_export(workbook_path: Path, jobs_path: Path, output_dir: Path, job_count: int) -> None:
    command = [
        "powershell.exe",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(EXCEL_RANGE_EXPORT_SCRIPT),
        "-WorkbookPath",
        str(workbook_path.resolve()),
        "-JobsPath",
        str(jobs_path.resolve()),
        "-OutputDir",
        str(output_dir.resolve()),
    ]
    creation_flags = 0x08000000 if os.name == "nt" else 0
    try:
        subprocess.run(
            command,
            check=True,
            capture_output=True,
            text=True,
            timeout=max(90, job_count * 8),
            creationflags=creation_flags,
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError("Excel chụp ảnh quá thời gian; hãy đóng các hộp thoại Excel đang mở và thử lại") from exc
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or exc.stdout or "").strip()
        raise RuntimeError(f"Không chụp được vùng Excel: {detail or 'Excel không phản hồi'}") from exc


def _unique_card_filename(filename: str, used: set[str], index: int) -> str:
    if filename not in used:
        used.add(filename)
        return filename
    path = Path(filename)
    unique = f"{path.stem}_{index}{path.suffix}"
    used.add(unique)
    return unique


def _render_employee_card_png(block: dict, payroll: dict | None, kind: str, period: dict) -> bytes:
    day_width = 56
    left_width = 72
    day_count = _period_day_count(period, block)
    payroll_width = 900 if kind == "output2" else 430
    top_height = 30
    day_row_height = 30
    info_row_height = 34
    punch_row_height = 118
    work_row_height = 36
    note_height = 28
    table_height = day_row_height + info_row_height + punch_row_height + work_row_height
    width = left_width + day_width * day_count + payroll_width + 2
    height = top_height + table_height + note_height + 2
    table_y = top_height
    right_x = left_width + day_width * day_count

    fonts = _load_fonts()
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)

    yellow = "#fff200"
    green = "#c5d99c"
    soft = "#f8fafc"
    black = "#111111"
    red = "#b91c1c"
    muted = "#4b5563"
    period_label = _period_range_label(period)
    tabulation_label = f"Tabulation {date.today().strftime('%Y-%m-%d')}"
    by_day = _block_results_by_day(block)
    sunday_days = _sunday_days(period)

    _draw_cell(draw, 0, 0, left_width, top_height, "white")
    draw.text((8, 8), "Att. Time", fill=black, font=fonts["small"])
    _draw_cell(draw, left_width, 0, day_width * min(day_count, 7), top_height, "white")
    draw.text((left_width + 8, 8), period_label, fill=black, font=fonts["small"])
    tab_x = left_width + day_width * min(day_count, 7)
    _draw_cell(draw, tab_x, 0, day_width * max(1, min(day_count - 7, 6)), top_height, "white")
    draw.text((tab_x + 8, 8), tabulation_label, fill=black, font=fonts["small"])
    if tab_x + day_width * 6 < right_x:
        _draw_cell(draw, tab_x + day_width * 6, 0, right_x - (tab_x + day_width * 6), top_height, "white")

    _draw_cell(draw, 0, table_y, left_width, day_row_height, soft)
    draw.text((8, table_y + 8), "Ngày", fill=black, font=fonts["small"])
    for day in range(1, day_count + 1):
        x = left_width + day_width * (day - 1)
        fill = yellow if day in sunday_days else "white"
        _draw_cell(draw, x, table_y, day_width, day_row_height, fill)
        _draw_wrapped_text(draw, str(day), x + 2, table_y + 7, day_width - 4, day_row_height - 4, fonts["label"], black, center=True)

    info_y = table_y + day_row_height
    _draw_cell(draw, 0, info_y, left_width, info_row_height, soft)
    draw.text((8, info_y + 9), "Mã:", fill=black, font=fonts["small"])
    _draw_cell(draw, left_width, info_y, day_width * day_count, info_row_height, "white")
    name = _employee_name(payroll)
    info_text = f"{block['employee_code']}"
    if name:
        info_text = f"{info_text}    Tên: {name}"
    draw.text((left_width + 8, info_y + 9), info_text, fill=black, font=fonts["small"])

    punch_y = info_y + info_row_height
    _draw_cell(draw, 0, punch_y, left_width, punch_row_height, soft)
    draw.text((8, punch_y + 10), "Giờ bấm", fill=black, font=fonts["small"])
    for day in range(1, day_count + 1):
        x = left_width + day_width * (day - 1)
        fill = yellow if day in sunday_days else "white"
        _draw_cell(draw, x, punch_y, day_width, punch_row_height, fill)
        value = _join_punches(by_day.get(day, {}).get("punches", []))
        _draw_wrapped_text(draw, value, x + 3, punch_y + 10, day_width - 6, punch_row_height - 14, fonts["tiny"], black, center=True)

    work_y = punch_y + punch_row_height
    _draw_cell(draw, 0, work_y, left_width, work_row_height, soft)
    draw.text((8, work_y + 10), "Công", fill=black, font=fonts["small"])
    for day in range(1, day_count + 1):
        x = left_width + day_width * (day - 1)
        fill = yellow if day in sunday_days else "white"
        _draw_cell(draw, x, work_y, day_width, work_row_height, fill)
        value = _format_value(by_day.get(day, {}).get("work_value"))
        color = red if day in sunday_days and value else black
        _draw_wrapped_text(draw, value, x + 3, work_y + 9, day_width - 6, work_row_height - 6, fonts["money"], color, center=True)

    if kind == "output2":
        _draw_payroll_section(
            draw,
            right_x,
            table_y,
            payroll_width,
            day_row_height + info_row_height,
            punch_row_height + work_row_height,
            note_height,
            payroll or {},
            fonts,
        )
    else:
        _draw_output1_info_section(
            draw,
            right_x,
            table_y,
            payroll_width,
            day_row_height + info_row_height,
            punch_row_height + work_row_height,
            note_height,
            block,
            payroll or {},
            fonts,
        )

    buffer = BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    return buffer.getvalue()


def _card_rows(block: dict) -> list[dict]:
    by_day: dict[int, dict] = {}
    for item in block.get("results", []):
        try:
            day = int(item.get("day"))
        except (TypeError, ValueError):
            continue
        if 1 <= day <= 31:
            by_day[day] = item

    days = range(1, 32)
    return [
        {"label": "Ngày", "height": 42, "values": [str(day) for day in days]},
        {
            "label": "Giờ bấm",
            "height": 86,
            "values": [_join_punches(by_day.get(day, {}).get("punches", [])) for day in days],
        },
        {
            "label": "Công",
            "height": 42,
            "values": [_format_value(by_day.get(day, {}).get("work_value")) for day in days],
        },
        {
            "label": "Quên / ?",
            "height": 42,
            "values": [_format_value(by_day.get(day, {}).get("missing_count")) for day in days],
        },
        {
            "label": "Trễ",
            "height": 42,
            "values": [_format_late(by_day.get(day, {}).get("late_minutes")) for day in days],
        },
    ]


def _block_results_by_day(block: dict) -> dict[int, dict]:
    by_day: dict[int, dict] = {}
    for item in block.get("results", []):
        try:
            day = int(item.get("day"))
        except (TypeError, ValueError):
            continue
        if 1 <= day <= 31:
            by_day[day] = item
    return by_day


def _draw_output1_info_section(
    draw,
    x: int,
    y: int,
    width: int,
    header_height: int,
    body_height: int,
    note_height: int,
    block: dict,
    payroll: dict,
    fonts: dict,
) -> None:
    yellow = "#fff200"
    green = "#c5d99c"
    black = "#111111"
    red = "#b91c1c"
    total_width = 120
    code_width = 80
    note_width = width - total_width - code_width
    total_hours = payroll.get("total_hours")
    if total_hours in {None, ""}:
        total_hours = _sum_block_work_hours(block)
    start_work_note = _format_start_work_note(payroll.get("start_work_note"))
    name_value = "\n".join(part for part in [_employee_name(payroll), start_work_note] if part)

    columns = [
        ("Tổng giờ công", total_width, _format_value(total_hours), green),
        ("Mã", code_width, str(block.get("employee_code") or ""), yellow),
        ("Tên / Ghi chú", note_width, name_value, yellow),
    ]

    offset = 0
    for title, col_width, value, fill in columns:
        _draw_cell(draw, x + offset, y, col_width, header_height, fill)
        _draw_wrapped_text(draw, title, x + offset + 6, y + 8, col_width - 12, header_height - 8, fonts["small"], red if title else black)
        _draw_cell(draw, x + offset, y + header_height, col_width, body_height, fill)
        value_y = y + header_height + 18 if title == "Tên / Ghi chú" else y + header_height + body_height - 34
        value_height = body_height - 22 if title == "Tên / Ghi chú" else 28
        _draw_wrapped_text(
            draw,
            value,
            x + offset + 6,
            value_y,
            col_width - 12,
            value_height,
            fonts["small"] if title == "Tên / Ghi chú" else fonts["money"],
            red,
            center=True,
        )
        offset += col_width

    if payroll.get("note"):
        _draw_cell(draw, x + total_width + code_width, y + header_height + body_height, note_width, note_height, yellow)
        _draw_wrapped_text(
            draw,
            str(payroll.get("note") or ""),
            x + total_width + code_width + 8,
            y + header_height + body_height + 9,
            note_width - 16,
            note_height - 8,
            fonts["money"],
            black,
        )


def _draw_payroll_section(
    draw,
    x: int,
    y: int,
    width: int,
    header_height: int,
    body_height: int,
    note_height: int,
    payroll: dict,
    fonts: dict,
) -> None:
    yellow = "#fff200"
    green = "#c5d99c"
    black = "#111111"
    red = "#b91c1c"
    name_value = _payroll_name_value(payroll)
    columns = [
        ("Tên / Ghi chú", 250, name_value),
        ("Tổng giờ công", 85, _format_value(payroll.get("total_hours"))),
        ("Ngày công", 95, _format_value(payroll.get("work_days"))),
        ("Mức lương", 115, _money(payroll.get("monthly_salary"))),
        ("Thưởng", 95, _money(payroll.get("bonus"))),
        ("Ứng + phạt", 105, _money(payroll.get("advance_or_penalty"))),
        ("Lương tháng", 155, _money(payroll.get("final_salary"))),
    ]
    offset = 0
    for title, col_width, value in columns:
        fill = yellow if title == "Tên / Ghi chú" else green
        _draw_cell(draw, x + offset, y, col_width, header_height, fill)
        _draw_wrapped_text(draw, title, x + offset + 6, y + 8, col_width - 12, header_height - 8, fonts["small"], black)
        _draw_cell(draw, x + offset, y + header_height, col_width, body_height, fill)
        _draw_wrapped_text(
            draw,
            str(value or ""),
            x + offset + 6,
            y + header_height + 18,
            col_width - 12,
            body_height - 20,
            fonts["money"],
            red if title == "Lương tháng" else black,
            center=True,
        )
        offset += col_width

    note = str(payroll.get("note") or "")
    if note:
        _draw_cell(draw, x, y + header_height + body_height, width, note_height, yellow)
        _draw_wrapped_text(draw, note, x + 8, y + header_height + body_height + 9, width - 16, note_height - 8, fonts["small"], red)


def _payroll_name_value(payroll: dict) -> str:
    name = str(payroll.get("name") or "").strip()
    start_work_note = _format_start_work_note(payroll.get("start_work_note"))
    return "\n".join(part for part in [name, start_work_note] if part)


def _draw_cell(draw, x: int, y: int, width: int, height: int, fill: str) -> None:
    draw.rectangle([x, y, x + width, y + height], fill=fill, outline="#111111", width=1)


def _draw_wrapped_text(draw, value: str, x: int, y: int, width: int, height: int, font, fill: str, center: bool = False) -> None:
    if not value:
        return

    lines: list[str] = []
    for raw_line in str(value).split("\n"):
        words = raw_line.split(" ")
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if _text_width(draw, candidate, font) <= width or not current:
                current = candidate
            else:
                lines.append(current)
                current = word
        if current:
            lines.append(current)

    line_height = max(10, font.size + 2)
    max_lines = max(1, height // line_height)
    lines = lines[:max_lines]
    for index, line in enumerate(lines):
        text_x = x
        if center:
            text_x = x + max(0, (width - _text_width(draw, line, font)) / 2)
        draw.text((text_x, y + index * line_height), line, fill=fill, font=font)


def _text_width(draw, text: str, font) -> int:
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0]


def _load_fonts() -> dict:
    font_path = _find_font_path()
    if font_path:
        return {
            "title": ImageFont.truetype(font_path, 22),
            "label": ImageFont.truetype(font_path, 16),
            "small": ImageFont.truetype(font_path, 14),
            "tiny": ImageFont.truetype(font_path, 11),
            "money": ImageFont.truetype(font_path, 15),
        }
    default = ImageFont.load_default()
    return {"title": default, "label": default, "small": default, "tiny": default, "money": default}


def _find_font_path() -> str | None:
    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    return None


def _render_employee_card_svg(block: dict, payroll: dict | None, kind: str, period_label: str) -> str:
    day_width = 48
    row_height = 34
    left_width = 110
    day_count = 31
    payroll_width = 760 if kind == "output2" else 0
    width = left_width + day_width * day_count + payroll_width + 2
    height = 286 if kind == "output2" else 250

    rows = [(row["label"], row["values"]) for row in _card_rows(block)]

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        "<style>"
        "text{font-family:Arial,'DejaVu Sans',sans-serif;fill:#111827}"
        ".small{font-size:12px}.tiny{font-size:10px}.label{font-size:13px;font-weight:700}"
        ".title{font-size:18px;font-weight:800}.money{font-size:13px;font-weight:800}"
        ".muted{fill:#4b5563}.red{fill:#b91c1c}.cell{stroke:#111;stroke-width:1;fill:#fff}"
        ".head{fill:#fff200}.green{fill:#c5d99c}.soft{fill:#f8fafc}"
        "</style>",
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="#ffffff"/>',
        f'<rect x="0" y="0" width="{width}" height="38" class="head" stroke="#111"/>',
        f'<text x="12" y="25" class="title">Phiếu nhân viên {html.escape(str(block["employee_code"]))}</text>',
        f'<text x="280" y="25" class="label muted">{html.escape(period_label)}</text>',
    ]

    start_y = 38
    parts.append(f'<rect x="0" y="{start_y}" width="{left_width}" height="{row_height}" class="cell soft"/>')
    parts.append(f'<text x="10" y="{start_y + 22}" class="label">Mã</text>')
    parts.append(f'<rect x="{left_width}" y="{start_y}" width="{day_width * day_count}" height="{row_height}" class="cell"/>')
    parts.append(
        f'<text x="{left_width + 10}" y="{start_y + 22}" class="label red">{html.escape(str(block["employee_code"]))}</text>'
    )

    for row_index, (label, values) in enumerate(rows):
        y = start_y + row_height * (row_index + 1)
        fill_class = "soft" if row_index % 2 == 0 else ""
        parts.append(f'<rect x="0" y="{y}" width="{left_width}" height="{row_height}" class="cell {fill_class}"/>')
        parts.append(f'<text x="10" y="{y + 22}" class="label">{html.escape(label)}</text>')
        for index in range(day_count):
            x = left_width + day_width * index
            weekend_class = "head" if index in {2, 9, 16, 23, 30} else ""
            parts.append(f'<rect x="{x}" y="{y}" width="{day_width}" height="{row_height}" class="cell {weekend_class}"/>')
            value = values[index] if index < len(values) else ""
            parts.extend(_multiline_text(value, x + day_width / 2, y + 13, day_width - 4, "tiny"))

    if kind == "output2":
        payroll_x = left_width + day_width * day_count
        parts.extend(_render_payroll_section(payroll_x, start_y, payroll_width, row_height, payroll or {}))

    parts.append("</svg>")
    return "\n".join(parts)


def _render_payroll_section(x: int, y: int, width: int, row_height: int, payroll: dict) -> list[str]:
    name_value = _payroll_name_value(payroll)
    columns = [
        ("Tên / Ghi chú", 210, name_value),
        ("Tổng giờ công", 75, _format_value(payroll.get("total_hours"))),
        ("Ngày công", 85, _format_value(payroll.get("work_days"))),
        ("Mức lương", 95, _money(payroll.get("monthly_salary"))),
        ("Thưởng", 80, _money(payroll.get("bonus"))),
        ("Ứng + phạt", 90, _money(payroll.get("advance_or_penalty"))),
        ("Lương tháng", 125, _money(payroll.get("final_salary"))),
    ]
    parts = []
    offset = 0
    for title, col_width, value in columns:
        fill = "green" if title not in {"Tên / Ghi chú"} else "head"
        parts.append(f'<rect x="{x + offset}" y="{y}" width="{col_width}" height="{row_height}" class="cell {fill}"/>')
        parts.append(f'<text x="{x + offset + 6}" y="{y + 22}" class="label">{html.escape(title)}</text>')
        parts.append(
            f'<rect x="{x + offset}" y="{y + row_height}" width="{col_width}" height="{row_height * 5}" class="cell {fill}"/>'
        )
        parts.extend(_multiline_text(str(value or ""), x + offset + col_width / 2, y + row_height + 24, col_width - 10, "money"))
        offset += col_width

    note = str(payroll.get("note") or "")
    if note:
        parts.append(f'<rect x="{x}" y="{y + row_height * 6}" width="{width}" height="{row_height}" class="cell head"/>')
        parts.extend(_multiline_text(note, x + 10, y + row_height * 6 + 20, width - 20, "small red", anchor="start"))
    return parts


def _apply_review_overrides_to_analysis(analysis: dict, review_overrides: list[dict]) -> None:
    overrides = _review_overrides_by_employee_day(review_overrides)
    for block in analysis.get("blocks", []):
        employee_code = str(block["employee_code"])
        for result in block.get("results", []):
            override = overrides.get((employee_code, result["day"]), {})
            if "missing_count" in override:
                result["missing_count"] = override["missing_count"]
            if "late_minutes" in override:
                result["late_minutes"] = override["late_minutes"]
            if "work_value" in override:
                result["work_value"] = override["work_value"]


def _review_overrides_by_employee_day(items: list[dict]) -> dict[tuple[str, int], dict]:
    result: dict[tuple[str, int], dict] = {}
    for item in items:
        employee_code = str(item.get("employee_code", "")).strip()
        day = item.get("day")
        if not employee_code or not isinstance(day, int):
            continue

        target = result.setdefault((employee_code, day), {})
        if "missing_count" in item:
            target["missing_count"] = item.get("missing_count")
        if "late_minutes" in item:
            target["late_minutes"] = item.get("late_minutes")
        if "work_value" in item:
            target["work_value"] = item.get("work_value")
    return result


def _multiline_text(value: str, x: float, y: float, max_width: int, class_name: str, anchor: str = "middle") -> list[str]:
    if not value:
        return []

    max_chars = max(4, max_width // 7)
    lines = []
    for raw_line in str(value).split("\n"):
        line = raw_line.strip()
        while len(line) > max_chars:
            lines.append(line[:max_chars])
            line = line[max_chars:]
        lines.append(line)
    lines = lines[:5]

    return [
        f'<text x="{x}" y="{y + index * 12}" class="{class_name}" text-anchor="{anchor}">{html.escape(line)}</text>'
        for index, line in enumerate(lines)
        if line
    ]


def _join_punches(punches: list[str]) -> str:
    return "\n".join(str(item) for item in punches)


def _format_late(value: object) -> str:
    if value in {None, ""}:
        return ""
    return f"{_format_value(value)} phút"


def _format_value(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sum_block_work_hours(block: dict) -> int | float:
    total = 0.0
    for item in block.get("results", []):
        value = item.get("work_value")
        if isinstance(value, (int, float)):
            total += float(value)
    rounded = round(total, 2)
    return int(rounded) if rounded.is_integer() else rounded


def _format_start_work_note(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.lower()
    if normalized.startswith("bắt đầu") or normalized.startswith("bat dau"):
        return text
    return f"Bắt đầu làm {text}"


def _employee_name(payroll: dict | None) -> str:
    return str((payroll or {}).get("name") or "").strip()


def _money(value: object) -> str:
    if value is None or value == "":
        return ""
    try:
        return f"{float(value):,.0f}"
    except (TypeError, ValueError):
        return str(value)


def _period_label(period: dict) -> str:
    month = period.get("month")
    year = period.get("year")
    if month and year:
        return f"Tháng {int(month):02d}/{int(year)}"
    return str(period.get("label") or "")


def _period_year_month(period: dict) -> tuple[int | None, int | None]:
    month = period.get("month")
    year = period.get("year")
    if isinstance(month, int) and isinstance(year, int):
        return year, month
    try:
        return int(year), int(month)
    except (TypeError, ValueError):
        return None, None


def _period_day_count(period: dict, block: dict) -> int:
    year, month = _period_year_month(period)
    if year and month:
        try:
            return calendar.monthrange(year, month)[1]
        except ValueError:
            pass

    days = []
    for item in block.get("results", []):
        try:
            days.append(int(item.get("day")))
        except (TypeError, ValueError):
            continue
    return max(days, default=31)


def _sunday_days(period: dict) -> set[int]:
    year, month = _period_year_month(period)
    if not year or not month:
        return set()
    try:
        day_count = calendar.monthrange(year, month)[1]
    except ValueError:
        return set()
    return {day for day in range(1, day_count + 1) if date(year, month, day).weekday() == 6}


def _period_range_label(period: dict) -> str:
    year, month = _period_year_month(period)
    if year and month:
        try:
            last_day = calendar.monthrange(year, month)[1]
            return f"{year}-{month:02d}-01 ~ {year}-{month:02d}-{last_day:02d}"
        except ValueError:
            pass
    return _period_label(period)


def _card_filename(employee_code: str, payroll: dict | None, kind: str, period_label: str) -> str:
    name = str((payroll or {}).get("name") or "").strip()
    suffix = "_".join(part for part in ["bang_cong", employee_code, name, kind, period_label] if part)
    safe = re.sub(r"[^\w.\-]+", "_", suffix, flags=re.UNICODE).strip("_")
    return f"{safe or employee_code}.png"
