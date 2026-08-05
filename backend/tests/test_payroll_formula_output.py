from pathlib import Path
from types import SimpleNamespace
import sys
import unittest
from unittest.mock import patch

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.payroll_workbook import (
    _build_employee_preview,
    _write_monthly_grand_total,
    _write_payroll_block,
)


class PayrollFormulaOutputTests(unittest.TestCase):
    def test_output2_uses_formula_layout_and_separate_note_areas(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        block = SimpleNamespace(header_row=3, result_row=9, employee_code="1006")
        preview = {
            "name": "Employee Name",
            "bank_account": "56010001632781",
            "monthly_salary": 3_330_000,
            "bonus": 200_000,
            "advance_or_penalty": 500_000,
            "start_work_note": "02/2023",
            "note": "Attendance comment",
        }
        try:
            sheet["C3"] = "2026-06-01 ~ 2026-06-30"
            _write_payroll_block(sheet, block, preview)

            expected_formulas = {
                "AF9": "=SUM(A9:AE9)",
                "AK9": "=AJ9/26",
                "AL9": "=AJ9/208",
                "AM9": "=SUM(A9:AE9)/8",
                "AN9": "=SUM(A10:AE10)",
                "AP9": "=IF(ISNUMBER(AG9),AG9*AF9,0)",
                "AR9": "=AK9*AM9+(AN9*AL9*1.5)-IF(ISNUMBER(AQ9),AQ9,0)-AP9+IF(ISNUMBER(AO9),AO9,0)",
            }
            for coordinate, formula in expected_formulas.items():
                self.assertEqual(sheet[coordinate].value, formula)

            self.assertEqual(sheet["AH9"].value, "1006")
            self.assertEqual(sheet["AI9"].value, "Employee Name")
            self.assertEqual(sheet["AI8"].value, "56010001632781")
            self.assertEqual(sheet["AI8"].number_format, "@")
            self.assertEqual(sheet["AJ9"].value, 3_330_000)
            self.assertEqual(sheet["AJ9"].font.color.rgb, "00000000")
            self.assertEqual(sheet["AO9"].value, 200_000)
            self.assertEqual(sheet["AQ9"].value, 500_000)
            self.assertIn("AJ10:AR10", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual(sheet["AI10"].value, "Bắt đầu làm 02/2023")
            self.assertEqual(sheet["AI10"].fill.fgColor.rgb, "00DDEBF7")
            self.assertTrue(sheet["AI10"].font.bold)
            self.assertEqual(sheet["AI10"].font.color.rgb, "00000000")
            self.assertEqual(sheet["AJ10"].value, "Attendance comment")
            self.assertEqual(sheet["AJ10"].fill.fgColor.rgb, "00FFF4CC")
            self.assertEqual(sheet["AJ10"].font.color.rgb, "00C00000")
        finally:
            workbook.close()

    def test_formula_only_output_omits_saved_payroll_data(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        block = SimpleNamespace(header_row=3, result_row=9, employee_code="1006")
        saved_entry = SimpleNamespace(
            name="Employee Name",
            start_work_note="2025-01-01",
            note="Local data",
            bonus=200_000,
            advance_or_penalty=500_000,
        )
        try:
            sheet["A9"] = 8
            sheet["A10"] = 2
            with patch("app.services.payroll_workbook.get_payroll_entry", return_value=saved_entry):
                preview = _build_employee_preview(sheet, block, include_saved_data=False)

            self.assertEqual(preview["employee_code"], "1006")
            self.assertEqual(preview["total_hours"], 8)
            self.assertEqual(preview["work_days"], 1)
            self.assertEqual(preview["overtime_hours"], 2)
            self.assertEqual(preview["nq_penalty"], 0)
            self.assertEqual(preview["final_salary"], 0)
            self.assertEqual(preview["name"], "")
            self.assertIsNone(preview["monthly_salary"])
            self.assertIsNone(preview["bonus"])
            self.assertIsNone(preview["advance_or_penalty"])
            self.assertEqual(preview["start_work_note"], "")
            self.assertEqual(preview["note"], "")
        finally:
            workbook.close()

    def test_output2_adds_monthly_grand_total(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        blocks = [
            SimpleNamespace(header_row=3, result_row=9),
            SimpleNamespace(header_row=11, result_row=17),
        ]
        try:
            sheet["C3"] = "2026-06-01 ~ 2026-06-30"
            _write_monthly_grand_total(sheet, blocks)

            self.assertIn("E19:AQ19", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual(sheet["E19"].value, "Tổng tháng 6/2026")
            self.assertEqual(sheet["AR19"].value, "=SUM(AR9,AR17)")
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
