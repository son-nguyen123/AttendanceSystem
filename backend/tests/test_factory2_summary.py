from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import sys
import unittest

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.factory2_workbook import analyze_factory2_workbook, export_factory2_output2, write_factory2_standard_source


class Factory2SummaryTests(unittest.TestCase):
    def test_counts_active_and_empty_employee_codes_separately(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "factory2.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Mã NV", "Tên NV", "Ngày", "Lần 1", "Lần 2"])
            sheet.append(["1001", "Nguyễn A", date(2026, 7, 1), "07:30", "17:00"])
            sheet.append(["1001", "Nguyễn A", date(2026, 7, 2), None, None])
            sheet.append(["1002", "Nguyễn B", date(2026, 7, 1), None, None])
            workbook.save(path)
            workbook.close()

            result = analyze_factory2_workbook(path)

            self.assertEqual(result["summary"]["blocks"], 1)
            self.assertEqual(result["summary"]["source_employee_count"], 2)
            self.assertEqual(result["summary"]["empty_employee_count"], 1)

    def test_uses_first_date_column_in_legacy_factory2_sheet(self) -> None:
        with TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "factory2-legacy.xlsx"
            output = Path(temp_dir) / "factory2-horizontal.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([None, None, None, None, None, None, "07:30"])
            sheet.append(["Maõ NV", "Teân NV", "Ngaøy", "Ngày", "Tối", "Phút thêm", "Laàn 1", "Laàn 2"])
            sheet.append([2, "Nguyễn A", date(2026, 6, 1), 8, 0, 0, "07:30", "17:00"])
            sheet.append([2, "Nguyễn A", date(2026, 6, 2), 0, 0, 0, None, None])
            sheet.append([3, "Nguyễn B", date(2026, 6, 1), 0, 0, 0, None, None])
            workbook.save(source)
            workbook.close()

            result = analyze_factory2_workbook(source)
            write_factory2_standard_source(source, output)

            self.assertEqual(result["period"]["label"], "06/2026")
            self.assertEqual(result["summary"]["blocks"], 1)
            self.assertEqual(result["summary"]["source_employee_count"], 2)
            self.assertTrue(output.exists())

    def test_output2_conversion_keeps_same_period_payroll_values_and_new_formulas(self) -> None:
        with TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "factory2-old-form.xlsx"
            output = Path(temp_dir) / "factory2-output2.xlsx"
            workbook = Workbook()
            attendance = workbook.active
            attendance.title = "Cham cong"
            attendance.append([None, None, None, None, None, None, "07:30"])
            attendance.append(["Maõ NV", "Teân NV", "Ngaøy", "Ngày", "Tối", "Phút thêm", "Laàn 1", "Laàn 2"])
            attendance.append([2, "Nguyễn A", date(2026, 6, 1), 8, 0, 0, "07:30", "17:00"])
            payroll = workbook.create_sheet("Luong cu")
            payroll.append(["Mã", "Tên nhân viên", "Mức Lương", "Thưởng", "Phạt NQ", "Ứng Lương"])
            payroll.append([2, "Nguyễn A", 8_000_000, 500_000, 120_000, 300_000])
            workbook.save(source)
            workbook.close()

            export_factory2_output2(
                source,
                output,
                include_saved_data=True,
                carry_source_payroll_data=True,
            )

            converted = load_workbook(output, data_only=False)
            sheet = converted.active
            self.assertEqual(sheet["AI9"].value, "NGUYEN A")
            self.assertEqual(sheet["AJ9"].value, 8_000_000)
            self.assertEqual(sheet["AO9"].value, 500_000)
            self.assertEqual(sheet["AP9"].value, 120_000)
            self.assertEqual(sheet["AQ9"].value, 300_000)
            self.assertTrue(str(sheet["AK9"].value).startswith("=IF("))
            self.assertTrue(str(sheet["AR9"].value).startswith("=IF("))
            converted.close()

    def test_output2_conversion_carries_old_factory2_summary_block_as_snapshot(self) -> None:
        with TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "factory2-old-summary.xlsx"
            output = Path(temp_dir) / "factory2-output2.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cham cong"
            sheet.append([None, None, None, None, None, None, "07:30"])
            sheet.append(["Maõ NV", "Teân NV", "Ngaøy", "Ngày", "Tối", "Phút thêm", "Laàn 1", "Laàn 2"])
            sheet.append([2, "2", date(2026, 6, 1), 8, 0, 0, "07:30", "17:00"])
            sheet.append([None, None, "Tổng h làm", 215.7])
            sheet.append([
                2,
                "Họ Và Tên",
                "Số Ngày đi làm",
                "Mức Lương",
                "Lương 1 Ngày",
                "Lương 1 Giờ",
                "Số H tăng ca",
                "Thưởng",
                "Ứng lương",
                "Lương Tháng 06/2026",
            ])
            sheet.append([2, "Nguyễn A", 26.9625, 3_120_000, 120_000, 15_000, 7.7, 500_000, 300_000, 3_294_000])
            workbook.save(source)
            workbook.close()

            export_factory2_output2(
                source,
                output,
                include_saved_data=True,
                carry_source_payroll_data=True,
            )

            converted = load_workbook(output, data_only=False)
            result = converted.active
            self.assertEqual(result["AF9"].value, 215.7)
            self.assertEqual(result["AH9"].value, "2")
            self.assertEqual(result["AI9"].value, "NGUYEN A")
            self.assertEqual(result["AJ9"].value, 3_120_000)
            self.assertEqual(result["AK9"].value, '=IF(AJ9>0,AJ9/26,"")')
            self.assertEqual(result["AL9"].value, '=IF(AJ9>0,AJ9/208,"")')
            self.assertEqual(result["AM9"].value, "=SUM(A9:AE9)/8")
            self.assertEqual(result["AN9"].value, 7.7)
            self.assertEqual(result["AO9"].value, 500_000)
            self.assertEqual(result["AQ9"].value, 300_000)
            self.assertEqual(result["AR9"].value, 3_294_000)
            self.assertTrue(str(result["AK9"].value).startswith("="))
            self.assertTrue(str(result["AL9"].value).startswith("="))
            self.assertTrue(str(result["AM9"].value).startswith("="))
            self.assertFalse(str(result["AR9"].value).startswith("="))
            converted.close()

    def test_output2_conversion_carries_legacy_start_and_free_form_note(self) -> None:
        """The old Xưởng 2 row stores profile text beside ``Tổng h làm``."""
        with TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "factory2-old-summary-with-note.xlsx"
            output = Path(temp_dir) / "factory2-output2.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Cham cong"
            sheet.append([None, None, None, None, None, None, "07:30"])
            sheet.append(["Maõ NV", "Teân NV", "Ngaøy", "Ngày", "Tối", "Phút thêm", "Laàn 1", "Laàn 2"])
            sheet.append([4, "Pham Thi Tuyet Trinh", date(2026, 6, 1), 8, 0, 0, "07:30", "17:00"])
            sheet.append([
                4,
                "Bắt đầu làm T3/2023",
                "Tổng h làm",
                196.6,
                "làm 2 ca, cố gắng làm chuyên cần nha, để ko mất thưởng.",
            ])
            sheet.append([
                4,
                "Họ Và Tên",
                "Số Ngày đi làm",
                "Mức Lương",
                "Lương 1 Ngày",
                "Lương 1 Giờ",
                "Số H tăng ca",
                "Thưởng",
                "Ứng lương",
                "Lương Tháng 06/2026",
            ])
            sheet.append([4, "Pham Thi Tuyet Trinh", 24.575, 3_120_000, 120_000, 15_000, 0, 0, 0, 2_949_000])
            workbook.save(source)
            workbook.close()

            export_factory2_output2(
                source,
                output,
                include_saved_data=True,
                carry_source_payroll_data=True,
            )

            converted = load_workbook(output, data_only=False)
            result = converted.active
            self.assertEqual(result["AI10"].value, "Bắt đầu làm T3/2023")
            self.assertEqual(
                result["AJ10"].value,
                "làm 2 ca, cố gắng làm chuyên cần nha, để ko mất thưởng.",
            )
            converted.close()


if __name__ == "__main__":
    unittest.main()
