import json
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
import zipfile

from openpyxl import Workbook

from app.services.employee_cards import export_employee_screenshots_from_workbook


class EmployeeExcelScreenshotTests(unittest.TestCase):
    def test_exports_one_exact_excel_range_per_employee(self) -> None:
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            workbook_path = root / "attendance.xlsx"
            output_path = root / "employee-images.zip"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Bang cong"
            for header_row, code, name in ((3, "1006", "Ho Thi Tu Uyen"), (11, "1229", "Truong Ngoc Vu")):
                worksheet.cell(header_row, 1, "Att. Time")
                worksheet.cell(header_row + 2, 1, "M\u00e3:")
                worksheet.cell(header_row + 2, 3, code)
                worksheet.cell(header_row + 6, 34, name)
            workbook.save(workbook_path)

            captured_jobs: list[dict[str, str]] = []

            def fake_excel_export(_workbook_path, jobs_path, output_dir, job_count):
                jobs = json.loads(jobs_path.read_text(encoding="utf-8"))
                self.assertEqual(job_count, 2)
                captured_jobs.extend(jobs)
                for job in jobs:
                    (output_dir / job["filename"]).write_bytes(b"valid-png-placeholder")

            with patch("app.services.employee_cards._run_excel_range_export", side_effect=fake_excel_export):
                export_employee_screenshots_from_workbook(workbook_path, output_path, "output1")

            self.assertEqual([job["range"] for job in captured_jobs], ["A3:AI10", "A11:AI18"])
            with zipfile.ZipFile(output_path) as archive:
                self.assertEqual(len(archive.namelist()), 2)
                self.assertTrue(all(name.endswith(".png") for name in archive.namelist()))

            output2_path = root / "employee-images-output2.zip"
            captured_jobs.clear()
            with patch("app.services.employee_cards._run_excel_range_export", side_effect=fake_excel_export):
                export_employee_screenshots_from_workbook(workbook_path, output2_path, "output2")
            self.assertEqual([job["range"] for job in captured_jobs], ["A3:AR10", "A11:AR18"])


if __name__ == "__main__":
    unittest.main()
