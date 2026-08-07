from pathlib import Path
from tempfile import TemporaryDirectory
import sys
import unittest
from unittest.mock import patch

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.models.attendance import EmployeeBlock
from app.services import bank_account_store, drive_backup, owner_profile_sync, payroll_store
from app.services.owner_profile_sync import sync_owner_profiles_from_workbook
from app.services.payroll_workbook import _build_employee_preview, _write_payroll_block


class FinalCopyProfileSyncTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory()
        root = Path(self.temp_dir.name)
        self.data_path = root / "payroll_data.json"
        self.sources_path = root / "payroll_profile_sources.json"
        self.bank_accounts_path = root / "bank_accounts.json"
        self.original_data_path = payroll_store.PAYROLL_DATA_PATH
        self.original_sources_path = payroll_store.PAYROLL_PROFILE_SOURCES_PATH
        self.original_bank_accounts_path = bank_account_store.REGISTRY_PATH
        payroll_store.PAYROLL_DATA_PATH = self.data_path
        payroll_store.PAYROLL_PROFILE_SOURCES_PATH = self.sources_path
        bank_account_store.REGISTRY_PATH = self.bank_accounts_path

    def tearDown(self) -> None:
        payroll_store.PAYROLL_DATA_PATH = self.original_data_path
        payroll_store.PAYROLL_PROFILE_SOURCES_PATH = self.original_sources_path
        bank_account_store.REGISTRY_PATH = self.original_bank_accounts_path
        self.temp_dir.cleanup()

    def test_final_copy_replaces_manual_profile_only_when_user_confirms(self) -> None:
        root = Path(self.temp_dir.name)
        source = root / "final.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        headers = {
            "AF4": "Tổng giờ công", "AH4": "Mã", "AI4": "Tên nhân viên / Ghi chú",
            "AJ4": "Mức Lương", "AK4": "Lương 1 Ngày Công", "AL4": "Lương 1 Giờ Công",
            "AO4": "Thưởng", "AQ4": "Ứng Lương",
        }
        for cell, value in headers.items():
            sheet[cell] = value
        sheet["AH9"] = "1006"
        sheet["AI8"] = "56010001632781"
        sheet["AI9"] = "Tên từ bản chốt"
        sheet["AJ9"] = 5_200_000
        sheet["AK9"] = 200_000
        sheet["AL9"] = 25_000
        sheet["AO9"] = 250_000
        sheet["AQ9"] = 999_999
        sheet["AI10"] = "Bắt đầu làm 02/2023 | Ghi chú cần đồng bộ"
        workbook.save(source)
        workbook.close()

        payroll_store.save_payroll_entry(
            "1006",
            payroll_store.PayrollEntry(
                name="Tên local bẩn", hourly_salary=1, bonus=1,
                advance_or_penalty=123_456, note="Ghi chú local bẩn",
            ),
        )

        result = sync_owner_profiles_from_workbook(
            source,
            factory="factory1",
            month=7,
            year=2026,
            source_kind="final_copy",
            overwrite_manual=True,
        )
        entry = payroll_store.get_payroll_entry("1006")

        self.assertEqual(result["updated_codes"], ["1006"])
        self.assertEqual(entry.name, "TEN TU BAN CHOT")
        self.assertEqual(entry.start_work_note, "02/2023")
        self.assertEqual(entry.note, "Ghi chú cần đồng bộ")
        self.assertEqual(entry.hourly_salary, 25_000)
        self.assertEqual(entry.bonus, 250_000)
        self.assertEqual(entry.advance_or_penalty, 123_456)
        self.assertEqual(bank_account_store.get_saved_account_number("factory1", "1006"), "56010001632781")
        self.assertEqual(bank_account_store.get_saved_account_number("factory2", "1006"), "")

        preview_workbook = Workbook()
        preview_sheet = preview_workbook.active
        preview_sheet["A9"] = 8
        preview = _build_employee_preview(
            preview_sheet,
            EmployeeBlock(
                sheet_name=preview_sheet.title, header_row=3, day_row=4, employee_row=5,
                punch_row=6, missing_row=7, late_row=8, result_row=9, employee_code="1006",
            ),
            profile_codes={"1006"},
            factory="factory1",
        )
        self.assertEqual(preview["bank_account"], "56010001632781")
        preview_workbook.close()

    def test_final_copies_are_partitioned_by_factory(self) -> None:
        root = Path(self.temp_dir.name)
        source = root / "final.xlsx"
        _write_reformed_final_copy_source(source)
        config = {"drive_backup_dir": str(root / "drive")}
        with patch("app.services.drive_backup.sync_owner_profiles_from_workbook", return_value={"status": "ok"}):
            factory1 = drive_backup.create_final_excel_copy(config, source, source.name, 7, 2026, factory="factory1")
            factory2 = drive_backup.create_final_excel_copy(config, source, source.name, 7, 2026, factory="factory2")

        self.assertIn("Xuong1", Path(factory1["path"]).parts)
        self.assertIn("Xuong2", Path(factory2["path"]).parts)
        self.assertNotEqual(Path(factory1["folder"]), Path(factory2["folder"]))
        self.assertEqual(len(drive_backup.list_final_excel_copies(config, factory="factory1")), 1)
        self.assertEqual(len(drive_backup.list_final_excel_copies(config, factory="factory2")), 1)

    def test_routine_final_copy_sync_keeps_manual_profile(self) -> None:
        root = Path(self.temp_dir.name)
        source = root / "final.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        for cell, value in {
            "AF4": "Tổng giờ công", "AH4": "Mã", "AI4": "Tên nhân viên / Ghi chú",
            "AJ4": "Mức Lương", "AK4": "Lương 1 Ngày Công", "AL4": "Lương 1 Giờ Công",
        }.items():
            sheet[cell] = value
        sheet["AH9"] = "1006"
        sheet["AI9"] = "Tên từ bản chốt"
        sheet["AL9"] = 25_000
        workbook.save(source)
        workbook.close()

        payroll_store.save_payroll_entry(
            "1006",
            payroll_store.PayrollEntry(name="Tên nhập tay", hourly_salary=20_000),
        )
        result = sync_owner_profiles_from_workbook(
            source, factory="factory1", month=7, year=2026, source_kind="final_copy",
        )

        entry = payroll_store.get_payroll_entry("1006")
        self.assertIn("1006", result["conflict_codes"])
        self.assertEqual(entry.name, "TEN NHAP TAY")
        self.assertEqual(entry.hourly_salary, 20_000)

    def test_latest_final_copy_is_authoritative_even_when_newer_than_current_attendance(self) -> None:
        root = Path(self.temp_dir.name)
        older = root / "older-final.xlsx"
        newest = root / "newest-final.xlsx"
        _write_simple_final_copy_source(older, "Tên cũ", 20_000)
        _write_simple_final_copy_source(newest, "Tên mới", 30_000)
        payroll_store.save_payroll_entry(
            "1006",
            payroll_store.PayrollEntry(name="Tên nhập tay", hourly_salary=1),
        )

        copies = [
            {"path": str(older), "month": 6, "year": 2026, "modified_at": "2026-06-30T10:00:00"},
            {"path": str(newest), "month": 7, "year": 2026, "modified_at": "2026-07-31T10:00:00"},
        ]
        with patch("app.services.cloud_sync.list_drive_final_copies", return_value=copies):
            result = owner_profile_sync.sync_latest_final_copy_profile(root / "attendance-june.xlsx", "factory1")

        entry = payroll_store.get_payroll_entry("1006")
        self.assertEqual(result["source_month"], 7)
        self.assertEqual(entry.name, "TEN MOI")
        self.assertEqual(entry.hourly_salary, 30_000)

    def test_history_or_analysis_file_cannot_refresh_profiles(self) -> None:
        source = Path(self.temp_dir.name) / "anything.xlsx"
        Workbook().save(source)

        result = sync_owner_profiles_from_workbook(source, source_kind="history_output2")

        self.assertEqual(result["reason"], "final_copy_only")
        self.assertFalse(self.data_path.exists())

    def test_note_is_written_only_in_name_and_note_area_and_expands_row(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        block = EmployeeBlock(
            sheet_name=sheet.title,
            header_row=3,
            day_row=4,
            employee_row=5,
            punch_row=6,
            missing_row=7,
            late_row=8,
            result_row=9,
            employee_code="1006",
        )
        sheet["AH10"] = "Nội dung cũ ở cột Mã"
        preview = {
            "name": "Nhân viên", "monthly_salary": 5_200_000, "daily_salary": 200_000,
            "hourly_salary": 25_000, "work_days": 26, "bonus": 0, "advance_or_penalty": 0,
            "start_work_note": "02/2023", "note": " ".join(["ghi chú dài"] * 80),
        }

        _write_payroll_block(sheet, block, preview)

        self.assertIsNone(sheet["AH10"].value)
        self.assertTrue(str(sheet["AI10"].value).startswith("Bắt đầu làm 02/2023"))
        self.assertTrue(str(sheet["AJ10"].value).startswith("ghi chú dài"))
        self.assertGreater(sheet.row_dimensions[10].height, 20)
        workbook.close()


def _write_simple_final_copy_source(path: Path, name: str, hourly_salary: int) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["AF4"] = "Tong gio cong"
    sheet["AH4"] = "Ma"
    sheet["AI4"] = "Ten nhan vien / Ghi chu"
    sheet["AJ4"] = "Muc Luong"
    sheet["AL4"] = "Luong 1 Gio Cong"
    sheet["AH9"] = "1006"
    sheet["AI9"] = name
    sheet["AJ9"] = hourly_salary * 26 * 8
    sheet["AL9"] = hourly_salary
    workbook.save(path)
    workbook.close()


def _write_reformed_final_copy_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A3"] = "Att. Time"
    sheet["C3"] = "2026-07-01 ~ 2026-07-31"
    sheet["A5"] = "Mã:"
    sheet["C5"] = "1006"
    headers = [
        "Tổng giờ công",
        "Mức tiền phạt NQ trên giờ công (đ)",
        "Mã",
        "Tên nhân viên / Ghi chú",
        "Mức Lương",
        "Lương 1 Ngày Công",
        "Lương 1 Giờ Công",
        "Số Ngày Đi Làm",
        "Giờ làm thêm",
        "Thưởng",
        "Phạt NQ",
        "Ứng Lương",
        "Lương Tháng 7/2026",
    ]
    for column, value in enumerate(headers, start=32):
        sheet.cell(row=4, column=column).value = value
    sheet["AF9"] = 160
    sheet["AH9"] = "1006"
    sheet["AI9"] = "Nhân viên 1006"
    sheet["AJ9"] = 5_200_000
    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
