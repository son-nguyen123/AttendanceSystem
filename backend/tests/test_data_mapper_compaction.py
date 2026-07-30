from pathlib import Path
from tempfile import TemporaryDirectory
import sys
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.data_mapper import (
    SummaryBlock,
    _owner_records_by_code,
    _prepare_mapping_source_workbook,
    _write_monthly_grand_total,
    _write_reformed_owner_area,
    map_owner_data_to_current_workbook,
)


class DataMapperCompactionTests(unittest.TestCase):
    def test_ignores_large_trailing_style_only_area(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "inflated.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "real data"
            sheet["A1000"].fill = PatternFill("solid", fgColor="FFFF00")
            workbook.save(source)
            workbook.close()

            prepared, ignored_rows = _prepare_mapping_source_workbook(source, root / "compact")

            self.assertNotEqual(prepared, source)
            self.assertEqual(ignored_rows, 1)
            compacted = load_workbook(prepared)
            try:
                self.assertEqual(compacted.active.max_row, 1)
                self.assertEqual(compacted.active["A1"].value, "real data")
            finally:
                compacted.close()

    def test_keeps_small_intentional_formatted_tail(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "normal.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "real data"
            sheet["A20"].fill = PatternFill("solid", fgColor="FFFF00")
            workbook.save(source)
            workbook.close()

            prepared, ignored_rows = _prepare_mapping_source_workbook(source, root / "compact")

            self.assertEqual(prepared, source)
            self.assertEqual(ignored_rows, 0)


class DataMapperNumberFormatTests(unittest.TestCase):
    def test_semantic_mapping_reads_old_layout_by_code_and_header_meaning(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            current_path = root / "current.xlsx"
            previous_path = root / "previous.xlsx"
            output_path = root / "mapped.xlsx"

            current_workbook = Workbook()
            current_sheet = current_workbook.active
            current_sheet["C3"] = "2026-07-01 ~ 2026-07-31"
            current_sheet["AF4"] = "Tổng giờ công"
            current_sheet["AH4"] = "Mã"
            current_sheet["AH9"] = "1006"
            current_sheet["AI9"] = "Tên tháng mới"
            current_workbook.save(current_path)
            current_workbook.close()

            previous_workbook = Workbook()
            previous_sheet = previous_workbook.active
            previous_sheet["C2"] = "2026-06-01 ~ 2026-06-30"
            previous_sheet["J5"] = "Mã"
            previous_sheet["O5"] = "Mức Lương"
            previous_sheet["V5"] = "Thưởng"
            previous_sheet["AB5"] = "Ứng Lương + Phạt"
            previous_sheet["AD5"] = "Lương Tháng 6/2026"
            previous_sheet["J12"] = "1006"
            previous_sheet["K12"] = "Nguyễn Văn A"
            previous_sheet["O12"] = 3_850_000
            previous_sheet["V12"] = 200_000
            previous_sheet["AB12"] = 500_000
            previous_sheet["K13"] = "Bắt đầu làm từ tháng 6"
            previous_workbook.save(previous_path)
            previous_workbook.close()

            check_workbook = load_workbook(previous_path, data_only=False)
            try:
                records = _owner_records_by_code(check_workbook.active, check_workbook.active)
                self.assertEqual(records["1006"].salary, 3_850_000)
                self.assertEqual(records["1006"].bonus, 200_000)
                self.assertTrue(records["1006"].has_previous_deduction)
            finally:
                check_workbook.close()

            summary = map_owner_data_to_current_workbook(
                current_path,
                previous_path,
                output_path,
                mode="output2",
            )
            mapped = load_workbook(output_path, data_only=False)
            try:
                sheet = mapped.active
                self.assertEqual(summary["matched_codes"], ["1006"])
                self.assertEqual(sheet["AI9"].value, "Nguyễn Văn A")
                self.assertEqual(sheet["AJ9"].value, 3_850_000)
                self.assertEqual(sheet["AO9"].value, 200_000)
                self.assertEqual(sheet["AQ9"].value, "?")
                self.assertEqual(sheet["AQ9"].fill.fgColor.rgb, "00C4D79B")
                self.assertEqual(sheet["AI10"].value, "Bắt đầu làm từ tháng 6")
            finally:
                mapped.close()

    def test_reformed_flexible_numbers_use_general_format(self) -> None:
        source_workbook = Workbook()
        source_values_workbook = Workbook()
        target_workbook = Workbook()
        target = SummaryBlock(
            code="1006",
            header_row=4,
            result_row=9,
            total_col=32,
            code_col=34,
            data_start_col=35,
            data_end_col=44,
        )
        try:
            _write_reformed_owner_area(
                source_workbook.active,
                source_values_workbook.active,
                target_workbook.active,
                None,
                target,
            )

            sheet = target_workbook.active
            for coordinate in ("AF9", "AG9", "AM9", "AN9"):
                self.assertEqual(sheet[coordinate].number_format, "General")
            for coordinate in ("AJ9", "AK9", "AL9", "AO9", "AP9", "AQ9", "AR9"):
                self.assertEqual(sheet[coordinate].number_format, "#,##0")
            self.assertIn("AI10:AR10", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual(sheet["AI10"].fill.fgColor.rgb, "00FFF4CC")
        finally:
            source_workbook.close()
            source_values_workbook.close()
            target_workbook.close()

    def test_previous_deduction_uses_green_question_marker(self) -> None:
        source_workbook = Workbook()
        source_values_workbook = Workbook()
        target_workbook = Workbook()
        source = SummaryBlock(
            code="1006",
            header_row=4,
            result_row=9,
            total_col=32,
            code_col=34,
            data_start_col=35,
            data_end_col=43,
        )
        target = SummaryBlock(
            code="1006",
            header_row=4,
            result_row=9,
            total_col=32,
            code_col=34,
            data_start_col=35,
            data_end_col=44,
        )
        try:
            source_workbook.active["AQ4"] = "Ứng Lương + Phạt"
            source_workbook.active["AQ9"] = 500_000
            source_values_workbook.active["AQ4"] = "Ứng Lương + Phạt"
            source_values_workbook.active["AQ9"] = 500_000

            needs_review = _write_reformed_owner_area(
                source_workbook.active,
                source_values_workbook.active,
                target_workbook.active,
                source,
                target,
            )

            self.assertTrue(needs_review)
            self.assertEqual(target_workbook.active["AQ9"].value, "?")
            self.assertEqual(target_workbook.active["AQ9"].fill.fgColor.rgb, "00C4D79B")
        finally:
            source_workbook.close()
            source_values_workbook.close()
            target_workbook.close()

    def test_mapping_output2_adds_monthly_grand_total(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        block = SummaryBlock(
            code="1006",
            header_row=4,
            result_row=9,
            total_col=32,
            code_col=34,
            data_start_col=35,
            data_end_col=44,
        )
        try:
            sheet["C3"] = "2026-06-01 ~ 2026-06-30"
            _write_monthly_grand_total(sheet, [block])

            self.assertIn("E11:AQ11", {str(item) for item in sheet.merged_cells.ranges})
            self.assertEqual(sheet["E11"].value, "Tổng tháng 6/2026")
            self.assertEqual(sheet["AR11"].value, "=SUM(AR9)")
            self.assertEqual(sheet["AR11"].fill.fgColor.rgb, "00FFFF00")
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
