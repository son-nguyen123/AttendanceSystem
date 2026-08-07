from pathlib import Path
from tempfile import TemporaryDirectory
import json
import sys
import unittest

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services import bank_payroll, bank_account_store
from app.services.data_mapper import map_owner_data_to_current_workbook


class BankPayrollScannerTests(unittest.TestCase):
    def test_scan_prefers_employee_summary_over_account_row(self) -> None:
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "output2.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            headers = {
                34: "Ma",
                35: "Ten nhan vien",
                36: "Muc luong",
                37: "Luong 1 Ngay Cong",
                38: "Luong 1 Gio Cong",
                39: "So Ngay Di Lam",
                40: "Gio lam them",
                41: "Thuong",
                42: "Phat NQ",
                43: "Ung Luong",
                44: "Luong Thang 6/2026",
            }
            for column, value in headers.items():
                sheet.cell(4, column).value = value

            # Output 2 stores the account one row above the real employee
            # summary. This value must never be treated as a code.
            sheet.cell(8, 34).value = "5601884980"
            sheet.cell(9, 34).value = "1006"
            sheet.cell(9, 35).value = "Alice"
            sheet.cell(9, 36).value = 3_120_000
            sheet.cell(9, 39).value = 20
            sheet.cell(9, 44).value = 3_000_000
            workbook.save(source)
            workbook.close()

            original_registry_path = bank_payroll.REGISTRY_PATH
            original_session_dir = bank_payroll.SESSION_DIR
            bank_payroll.REGISTRY_PATH = root / "bank_accounts.json"
            bank_payroll.SESSION_DIR = root / "bank_payroll"
            bank_payroll.REGISTRY_PATH.write_text(
                json.dumps({
                    "factory1:1006": {
                        "factory": "factory1",
                        "employee_code": "1006",
                        "account_number": "5601884980",
                    },
                }),
                encoding="utf-8",
            )
            try:
                result = bank_payroll.scan_official_workbook(source, "factory1")
            finally:
                bank_payroll.REGISTRY_PATH = original_registry_path
                bank_payroll.SESSION_DIR = original_session_dir

            self.assertEqual([item["employee_code"] for item in result["employees"]], ["1006"])
            self.assertEqual(result["employees"][0]["account_number"], "5601884980")
            self.assertEqual(result["employees"][0]["salary"], 3_000_000)


class BankAccountMappingLayoutTests(unittest.TestCase):
    def test_output2_account_is_not_written_into_code_column(self) -> None:
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            current = root / "current.xlsx"
            previous = root / "previous.xlsx"
            output = root / "output2.xlsx"

            current_workbook = Workbook()
            current_sheet = current_workbook.active
            current_sheet["AF4"] = "Tong gio cong"
            current_sheet["AG4"] = "Ma"
            current_sheet["AH4"] = "Ten nhan vien"
            current_sheet["AG9"] = "1006"
            current_sheet["AH9"] = "Alice"
            current_workbook.save(current)
            current_workbook.close()

            previous_workbook = Workbook()
            previous_sheet = previous_workbook.active
            previous_sheet["J5"] = "Ma"
            previous_sheet["O5"] = "Muc luong"
            previous_sheet["V5"] = "Thuong"
            previous_sheet["AD5"] = "Luong Thang 6/2026"
            previous_sheet["J12"] = "1006"
            previous_sheet["K12"] = "Alice"
            previous_sheet["O12"] = 3_850_000
            previous_workbook.save(previous)
            previous_workbook.close()

            original_registry_path = bank_account_store.REGISTRY_PATH
            bank_account_store.REGISTRY_PATH = root / "bank_accounts.json"
            bank_account_store.REGISTRY_PATH.write_text(
                json.dumps({
                    "factory1:1006": {
                        "factory": "factory1",
                        "employee_code": "1006",
                        "account_number": "5601884980",
                    },
                }),
                encoding="utf-8",
            )
            try:
                map_owner_data_to_current_workbook(current, previous, output, mode="output2")
            finally:
                bank_account_store.REGISTRY_PATH = original_registry_path

            workbook = load_workbook(output, data_only=False)
            try:
                sheet = workbook.active
                self.assertIsNone(sheet["AH8"].value)
                self.assertEqual(sheet["AI8"].value, "5601884980")
                self.assertEqual(sheet["AH9"].value, "1006")
            finally:
                workbook.close()


class OldSalaryExcelImportTests(unittest.TestCase):
    def test_import_reads_salary_template_and_ignores_other_month_rows(self) -> None:
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "salary-old.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A2"] = "THIỆN TRÍ THÁNG 6/2026"
            sheet.append([])
            sheet.append(["STT", "Mã Nhân Viên", "HỌ TÊN", "SỐ TK", "SỐ TIỀN", "NỘI DUNG"])
            sheet.append([1, 1006, "Alice", "5601884980", 1000000, "T6/2026"])
            sheet.append([2, 1007, "Bob", "5601884981", 1000000, "T5/2026"])
            workbook.save(source)
            workbook.close()

            original_registry_path = bank_payroll.REGISTRY_PATH
            bank_payroll.REGISTRY_PATH = root / "bank_accounts.json"
            try:
                self.assertEqual(bank_payroll.detect_excel_salary_period(source), (6, 2026))
                result = bank_payroll.import_accounts_from_excel_salary(source, "factory1", 6, 2026, "replace")
                registry = bank_payroll._load_registry()
            finally:
                bank_payroll.REGISTRY_PATH = original_registry_path

            self.assertEqual(result["imported"], 1)
            self.assertEqual(result["month"], 6)
            self.assertEqual(registry["factory1:1006"]["account_number"], "5601884980")
            self.assertNotIn("factory1:1007", registry)

    def test_old_salary_import_only_adds_missing_and_reports_mismatches(self) -> None:
        with TemporaryDirectory() as temporary_dir:
            root = Path(temporary_dir)
            source = root / "salary-old.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["A2"] = "THIỆN TRÍ THÁNG 6/2026"
            sheet.append([])
            sheet.append(["STT", "Mã Nhân Viên", "HỌ TÊN", "SỐ TK", "SỐ TIỀN", "NỘI DUNG"])
            sheet.append([1, 1006, "Alice", "5601884981", 1000000, "T6/2026"])
            sheet.append([2, 1007, "Bob", "5601884982", 1000000, "T6/2026"])
            sheet.append([3, 1008, "Cara", "5601884983", 1000000, "T6/2026"])
            workbook.save(source)
            workbook.close()

            original_registry_path = bank_payroll.REGISTRY_PATH
            bank_payroll.REGISTRY_PATH = root / "bank_accounts.json"
            bank_payroll.REGISTRY_PATH.write_text(
                json.dumps({
                    "factory1:1006": {
                        "factory": "factory1",
                        "employee_code": "1006",
                        "name": "Alice",
                        "account_number": "5601884900",
                    },
                    "factory1:1007": {
                        "factory": "factory1",
                        "employee_code": "1007",
                        "name": "Bob",
                        "account_number": "5601884982",
                    },
                }),
                encoding="utf-8",
            )
            try:
                result = bank_payroll.import_accounts_from_excel_salary(source, "factory1", 6, 2026, "fill_missing")
                registry = bank_payroll._load_registry()
                bank_payroll.save_accounts("factory1", [{"employee_code": "1006", "name": "Alice", "account_number": "5601884981"}])
                updated_registry = bank_payroll._load_registry()
            finally:
                bank_payroll.REGISTRY_PATH = original_registry_path

            self.assertEqual(result["imported"], 1)
            self.assertEqual(result["skipped_existing"], ["1007"])
            self.assertEqual(result["conflicts"][0]["employee_code"], "1006")
            self.assertEqual(result["conflicts"][0]["existing_account"], "5601884900")
            self.assertEqual(result["conflicts"][0]["file_accounts"], ["5601884981"])
            self.assertEqual(registry["factory1:1006"]["account_number"], "5601884900")
            self.assertEqual(registry["factory1:1008"]["account_number"], "5601884983")
            self.assertEqual(updated_registry["factory1:1006"]["account_number"], "5601884981")


if __name__ == "__main__":
    unittest.main()
