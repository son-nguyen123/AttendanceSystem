from pathlib import Path
import unittest

from openpyxl import Workbook, load_workbook

from app.services.factory1_workbook import export_factory1_legacy_output2


def _legacy_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "May"
    sheet["A1"] = "Att. Time"
    sheet["C1"] = "2026-05-01 ~ 2026-05-31"
    sheet["A3"] = "Mã:"
    sheet["C3"] = "1001"
    sheet["D4"] = "07:2707:2711:3512:5817:01"
    sheet["D7"] = 8

    sheet.cell(3, 32).value = "Tổng giờ công"
    sheet.cell(3, 33).value = "Mã"
    sheet.cell(3, 34).value = "Tên nhân viên"
    sheet.cell(3, 35).value = "Mức lương"
    sheet.cell(3, 36).value = "Lương 1 Ngày Công"
    sheet.cell(3, 37).value = "Lương 1 Giờ Công"
    sheet.cell(3, 38).value = "Số Ngày Đi Làm"
    sheet.cell(3, 39).value = "Giờ làm thêm"
    sheet.cell(3, 40).value = "Thưởng"
    sheet.cell(3, 41).value = "Ứng Lương"
    sheet.cell(3, 42).value = "Lương Tháng 05/2026"
    sheet.cell(8, 33).value = "1001"
    sheet.cell(8, 34).value = "Nguyễn Thị Mẫu"
    sheet.cell(8, 35).value = 3120000
    sheet.cell(8, 39).value = 2.5
    sheet.cell(8, 40).value = 100000
    sheet.cell(8, 42).value = 3500000
    sheet.cell(9, 34).value = "T3/2023"
    sheet.cell(9, 35).value = "Ghi chú cũ"
    workbook.save(path)


class Factory1LegacyConversionTests(unittest.TestCase):
    def test_keeps_notes_and_writes_formulas(self):
        import tempfile

        temp_dir = Path(tempfile.mkdtemp(prefix="factory1-test-"))
        try:
            source = temp_dir / "legacy.xlsx"
            output = temp_dir / "converted.xlsx"
            _legacy_workbook(source)

            export_factory1_legacy_output2(source, output)
            sheet = load_workbook(output, data_only=False).active

            self.assertEqual(sheet["D4"].value, "07:27\n11:35\n12:58\n17:01")
            self.assertEqual(sheet["D7"].alignment.horizontal, "center")
            self.assertEqual(sheet.column_dimensions["A"].width, 6.2)
            self.assertEqual(sheet.column_dimensions["AE"].width, 6.2)
            self.assertEqual(sheet["AH7"].value, "1001")
            self.assertEqual(sheet["AI7"].value, "NGUYEN THI MAU")
            self.assertEqual(sheet["AJ7"].value, 3120000)
            self.assertTrue(sheet["AK7"].value.startswith("=IF("))
            self.assertTrue(sheet["AL7"].value.startswith("=IF("))
            self.assertTrue(sheet["AM7"].value.startswith("=SUM("))
            self.assertEqual(sheet["AN7"].value, 2.5)
            self.assertTrue(str(sheet["AI8"].value).startswith("Bắt đầu làm"))
            self.assertEqual(sheet["AJ8"].value, "Ghi chú cũ")
            self.assertIn(sheet["AI6"].value, (None, ""))
            self.assertTrue(sheet["AR7"].value.startswith("=IF("))
        finally:
            import shutil

            shutil.rmtree(temp_dir, ignore_errors=True)
