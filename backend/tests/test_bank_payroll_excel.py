from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from app.services.bank_payroll import _build_bank_workbook


class BankPayrollExcelTests(unittest.TestCase):
    def test_workbook_matches_bank_six_column_layout(self) -> None:
        employees = [
            {
                "employee_code": "1006",
                "name": "Hồ Thị Tú Uyên",
                "account_number": "5601884980",
                "salary": 3330000,
            },
            {
                "employee_code": "0012",
                "name": "Nguyễn Văn A",
                "account_number": "0012345678",
                "salary": 2000000,
            },
        ]
        with TemporaryDirectory() as temporary_dir:
            output = Path(temporary_dir) / "bank-payroll.xlsx"
            _build_bank_workbook(output, employees, 6, 2026)
            workbook = load_workbook(output, data_only=True)

        worksheet = workbook["BangLuongNganHang"]
        self.assertEqual(worksheet["A2"].value, " THIỆN TRÍ THÁNG 6/2026")
        self.assertEqual(
            [worksheet.cell(3, column).value for column in range(1, 7)],
            ["STT", "Mã Nhân Viên", "HỌ TÊN", "SỐ TK", "SỐ TIỀN", "NỘI DUNG"],
        )
        self.assertEqual(
            [worksheet.cell(4, column).value for column in range(1, 7)],
            [1, "1006", "HO THI TU UYEN", "5601884980", 3330000, "T6/2026"],
        )
        self.assertEqual(worksheet["B5"].value, "0012")
        self.assertEqual(worksheet["D5"].value, "0012345678")
        self.assertEqual(worksheet["A6"].value, "TỔNG")
        self.assertEqual(worksheet["E6"].value, 5330000)


if __name__ == "__main__":
    unittest.main()
