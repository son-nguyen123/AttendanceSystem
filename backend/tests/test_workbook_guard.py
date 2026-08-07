from pathlib import Path
from tempfile import TemporaryDirectory
import sys
import unittest

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.workbook_guard import (
    inspect_workbook_for_role,
    profile_workbook,
    validate_mapping_pair,
)


class WorkbookGuardTests(unittest.TestCase):
    def test_classifies_output_roles_and_accepts_valid_mapping_pair(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            current = root / "current.xlsx"
            previous = root / "previous.xlsx"
            _write_workbook(current, month=7, private=False, codes=["1006", "1194"])
            _write_workbook(previous, month=6, private=True, codes=["1006", "1194"])

            current_profile, previous_profile = validate_mapping_pair(current, previous)

            self.assertEqual(current_profile.detected_kind, "output1")
            self.assertEqual(previous_profile.detected_kind, "output2")
            self.assertEqual(current_profile.period_label, "07/2026")
            self.assertEqual(previous_profile.period_label, "06/2026")

    def test_rejects_swapped_mapping_files(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            official = root / "official.xlsx"
            output1 = root / "output1.xlsx"
            _write_workbook(official, month=6, private=True, codes=["1006"])
            _write_workbook(output1, month=7, private=False, codes=["1006"])

            with self.assertRaisesRegex(ValueError, "chọn ngược"):
                validate_mapping_pair(official, output1)

    def test_rejects_missing_separator_row(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "shifted.xlsx"
            _write_workbook(path, month=7, private=False, codes=["1006", "1194"], spacing=7)

            profile = profile_workbook(path)

            self.assertTrue(profile.structural_errors)
            with self.assertRaisesRegex(ValueError, "thiếu hoặc thừa dòng"):
                inspect_workbook_for_role(path, "mapping_current")

    def test_accepts_old_official_layout_with_missing_separator_row(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "old_official_shifted.xlsx"
            _write_workbook(path, month=6, private=True, codes=["1006", "1194"], spacing=7)

            profile = inspect_workbook_for_role(path, "mapping_previous")

            self.assertEqual(profile.detected_kind, "output2")
            self.assertEqual(set(profile.employee_codes), {"1006", "1194"})
            self.assertFalse(profile.structural_errors)

    def test_analysis_rejects_already_processed_output(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "output1.xlsx"
            _write_workbook(path, month=7, private=False, codes=["1006"])

            with self.assertRaisesRegex(ValueError, "Output 1 đã tính"):
                inspect_workbook_for_role(path, "analysis")

    def test_public_penalty_rate_does_not_make_output1_private(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "output1_with_penalty_rate.xlsx"
            _write_workbook(path, month=7, private=False, codes=["1006"])
            workbook = load_workbook(path)
            workbook.active.cell(4, 33).value = "Muc tien phat NQ tren gio cong (d)"
            workbook.save(path)
            workbook.close()

            profile = inspect_workbook_for_role(path, "recalculate_output1")

            self.assertEqual(profile.detected_kind, "output1")

    def test_final_copy_rejects_legacy_output2_layout(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "legacy_output2.xlsx"
            _write_output2(path, month=6, reformed=False)

            profile = profile_workbook(path)

            self.assertEqual(profile.detected_kind, "output2")
            self.assertFalse(profile.has_reformed_output2_layout)
            with self.assertRaisesRegex(ValueError, "khung mới"):
                inspect_workbook_for_role(path, "final_copy")

    def test_final_copy_accepts_reformed_output2_layout(self) -> None:
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "reformed_output2.xlsx"
            _write_output2(path, month=6, reformed=True)

            profile = inspect_workbook_for_role(path, "final_copy")

            self.assertEqual(profile.detected_kind, "output2")
            self.assertTrue(profile.has_reformed_output2_layout)


def _write_workbook(
    path: Path,
    *,
    month: int,
    private: bool,
    codes: list[str],
    spacing: int = 8,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    for index, code in enumerate(codes):
        header_row = 3 + index * spacing
        day_row = header_row + 1
        employee_row = header_row + 2
        result_row = header_row + 6
        sheet.cell(header_row, 1).value = "Att. Time"
        sheet.cell(header_row, 3).value = f"2026-{month:02d}-01 ~ 2026-{month:02d}-30"
        sheet.cell(day_row, 32).value = "Tổng giờ công"
        sheet.cell(day_row, 34).value = "Mã"
        sheet.cell(employee_row, 1).value = "Mã:"
        sheet.cell(employee_row, 3).value = code
        sheet.cell(result_row, 34).value = code
        sheet.cell(result_row, 35).value = f"Nhân viên {code}"
        if private:
            sheet.cell(day_row, 36).value = "Mức Lương"
            sheet.cell(day_row, 41).value = "Thưởng"
            sheet.cell(day_row, 43).value = "Ứng Lương"
            sheet.cell(day_row, 44).value = f"Lương Tháng {month}/2026"
            sheet.cell(result_row, 36).value = 3_000_000
    workbook.save(path)
    workbook.close()


def _write_output2(path: Path, *, month: int, reformed: bool) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet["A3"] = "Att. Time"
    sheet["C3"] = f"2026-{month:02d}-01 ~ 2026-{month:02d}-30"
    sheet["A5"] = "Mã:"
    sheet["C5"] = "1006"

    headers = {
        32: "Tổng giờ công",
        33: "Mức tiền phạt NQ trên giờ công (đ)",
        34: "Mã",
        35: "Tên nhân viên / Ghi chú",
        36: "Mức Lương",
        37: "Lương 1 Ngày Công",
        38: "Lương 1 Giờ Công",
        39: "Số Ngày Đi Làm",
        40: "Giờ làm thêm",
        41: "Thưởng",
        42: "Phạt NQ",
        43: "Ứng Lương",
        44: f"Lương Tháng {month}/2026",
    }
    if not reformed:
        headers = {
            32: "Tổng giờ công",
            33: "Mã",
            34: "Tên nhân viên / Ghi chú",
            37: "Mức Lương",
            38: "Lương 1 Ngày Công",
            39: "Lương 1 Giờ Công",
            40: "Số Ngày Đi Làm",
            41: "Thưởng",
            42: "Ứng Lương + Phạt",
            43: f"Lương Tháng {month}/2026",
        }
    for col, value in headers.items():
        sheet.cell(row=4, column=col).value = value

    result_row = 9
    code_col = 34 if reformed else 33
    name_col = code_col + 1
    sheet.cell(row=result_row, column=32).value = 160
    sheet.cell(row=result_row, column=code_col).value = "1006"
    sheet.cell(row=result_row, column=name_col).value = "Nhân viên 1006"
    sheet.cell(row=result_row, column=36 if reformed else 37).value = 5_200_000
    workbook.save(path)
    workbook.close()


if __name__ == "__main__":
    unittest.main()
