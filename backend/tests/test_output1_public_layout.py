from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.final_copy_export import export_final_copy_output1
from app.services.workbook_processor import _write_output1_summary_block


class Output1PublicLayoutTests(unittest.TestCase):
    def test_session_output1_uses_public_columns_through_ai(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        block = SimpleNamespace(
            header_row=3,
            day_row=4,
            result_row=9,
            employee_code="1229",
        )
        sheet["A9"] = 8
        entry = SimpleNamespace(name="TRUONG NGOC VU", start_work_note="T3/2024")
        try:
            with (
                patch("app.services.workbook_processor.get_payroll_entry", return_value=entry),
                patch("app.services.workbook_processor.get_saved_account_number", return_value="8812007717"),
            ):
                _write_output1_summary_block(sheet, block)

            self.assertEqual(sheet["AF4"].value, "Tổng giờ công")
            self.assertEqual(sheet["AG4"].value, "Mức tiền phạt NQ trên giờ công (đ)")
            self.assertEqual(sheet["AH4"].value, "Mã")
            self.assertEqual(sheet["AI4"].value, "Tên nhân viên / Ghi chú")
            self.assertEqual(sheet["AH9"].value, "1229")
            self.assertEqual(sheet["AI8"].value, "8812007717")
            self.assertEqual(sheet["AI9"].value, "TRUONG NGOC VU")
            self.assertEqual(sheet["AI10"].value, "Bắt đầu làm T3/2024")
        finally:
            workbook.close()

    def test_final_copy_output1_physically_ends_at_ai(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "final.xlsx"
            output = root / "output1.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet["AF4"] = "Tổng giờ công"
            sheet["AH4"] = "Mã"
            sheet["AI4"] = "Tên nhân viên / Ghi chú"
            sheet["AJ4"] = "Mức Lương"
            sheet["AH9"] = "1229"
            sheet["AI9"] = "TRUONG NGOC VU"
            sheet["AJ9"] = 3_330_000
            sheet["AR9"] = 3_297_981
            workbook.save(source)
            workbook.close()

            export_final_copy_output1(source, output)

            exported = load_workbook(output, data_only=False)
            try:
                self.assertEqual(exported.active.max_column, 35)
                self.assertEqual(exported.active["AI9"].value, "TRUONG NGOC VU")
                self.assertEqual(exported.active.print_area, "'Sheet'!$A$1:$AI$9")
            finally:
                exported.close()


if __name__ == "__main__":
    unittest.main()
